#can_battery_info.py
import customtkinter as ctk
import tkinter as tk
import os
import pandas as pd
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from PIL import Image, ImageTk
Image.CUBIC = Image.BICUBIC  # Image resizing method
import datetime
import threading
import time
import asyncio
from tkinter import messagebox
from openpyxl import Workbook
from pcan_api.custom_pcan_methods import *  # Custom PCAN methods for CAN communication
from load_api.chroma_api import *  # Custom API for Chroma communication
import helpers.config as config  # Config helper


class CanBatteryInfo:
    def __init__(self, master, main_window=None):
        """Initialize the CanBatteryInfo class."""
        self.master = master  # Parent window
        self.main_window = main_window  # Reference to the main application window

        # Setup the window and initialize variables
        self.selected_button = None  # Track the currently selected side menu button
        self.master.resizable(True, False)  # Allow horizontal resizing only

        # Calculate and set the side menu width as a ratio of the window width
        self.side_menu_width_ratio = 0.20  # 20% for side menu
        self.update_frame_sizes()  # Set initial frame sizes

        # Boolean variables to track battery states
        self.check_discharge_max = tk.BooleanVar()  # Track if max discharge is selected
        self.show_dynamic_fields_var = tk.BooleanVar()  # Track if dynamic fields are shown

        # Initialize Chroma device connection state
        self.is_connected = True
        self.skyla_device_status = tk.BooleanVar(value=False)  # Testing mode specific state
        self.skyla_mcb_status = tk.BooleanVar(value=False)  # Maintenance mode specific
        self.chroma_device_status = tk.BooleanVar(value=False)  # Testing mode specific state
        self.chroma_mcb_status = tk.BooleanVar(value=False)  # Maintenance mode specific
        self.agree_charging_status = tk.BooleanVar(value=False)  # Custom mode specific state
        self.agree_discharging_status = tk.BooleanVar(value=False)  # Initially OFF
        self.load_status = tk.BooleanVar(value=False)

        # Track connection status string
        self.connection_status_var = tk.StringVar(value="Disconnected")

        # Battery selection control variables
        self.charger_control_var_battery_1 = tk.BooleanVar(value=False)  # Charger for Battery 1
        self.charger_control_var_battery_2 = tk.BooleanVar(value=False)  # Charger for Battery 2
        self.discharger_control_var_battery_1 = tk.BooleanVar(value=False)  # Discharger for Battery 1
        self.discharger_control_var_battery_2 = tk.BooleanVar(value=False)  # Discharger for Battery 2

        # Set the default battery and tracking variables
        self.selected_battery = "Battery 1"  # Default battery selection
        self.charger_control_var = self.charger_control_var_battery_1
        self.discharger_control_var = self.discharger_control_var_battery_1

        self.logging_active = False  # Track whether logging is active
        self.limited = False  # Flag to show limited or full data
        self.first_time_dashboard = False  # Track if it's the first time the dashboard is shown
        self.mode_var = tk.StringVar(value="Testing")  # Mode variable (Testing by default)
        self.control_var = tk.StringVar(value="Charging")  # control variable (Charging by default)

        # Set the path for images and assets
        base_path = os.path.dirname(os.path.abspath(__file__))
        self.assets_path = os.path.join(base_path, "../assets/images/")
        self.assets_icon_path = os.path.join(base_path, "../assets/icons/")

        # Style configuration for GUI components
        style = ttk.Style()
        style.configure("TLabelframe.Label", font=("Helvetica", 10, "bold"))  # Bold Labelframe titles

        style = ttk.Style()
        style.configure("Custom.Treeview",
                        background="#f0f0f0",  # Set background color
                        foreground="black",    # Set text color
                        rowheight=25,          # Set row height
                        fieldbackground="#f0f0f0")  # Set field background color
        style.configure("Custom.Treeview.Heading", 
                        font=("Helvetica", 12, "bold"), 
                        background="#007f4e",   # Set header background
                        foreground="white")     # Set header text color
        # Define a custom style for the Radiobutton
        style.configure('Custom.TRadiobutton', font=('Helvetica', 12, 'bold'), foreground="black")
        self.caution_alert_shown = False
        self.alarm_alert_shown = False
        self.critical_alert_shown = False
        self.device_data = None  # Variable to store device data
        self.battery_status_flags = None  # Variable to store battery status flags
        self.device_data = device_data_battery_1  # Initially set to Battery 1 data
        self.battery_status_flags = battery_1_status_flags  # Battery 1 status flags

        self.check_battery_log_for_cycle_count()  # Check if the battery log contains cycle count

        # Create the main frame container for the GUI
        self.main_frame = ttk.Frame(self.master)
        self.main_frame.pack(fill="both", expand=True)

        # Create the side menu and content frames
        self.side_menu_frame = ttk.Frame(self.main_frame, bootstyle="info")  # Side menu frame
        self.side_menu_frame.place(x=0, y=0, width=self.side_menu_width, relheight=1.0)

        self.content_frame = ttk.Frame(self.main_frame, bootstyle="light")  # Content frame for displaying data
        self.content_frame.place(x=self.side_menu_width, y=0, width=self.content_frame_width, relheight=1.0)

        # Load icons for side menu buttons
        self.disconnect_icon = self.load_icon_menu(os.path.join(self.assets_path, "disconnect_button.png"))
        self.control_icon = self.load_icon_menu(os.path.join(self.assets_path, "load_icon.png"))
        self.config_icon = self.load_icon_menu(os.path.join(self.assets_path, "settings.png"))
        self.info_icon = self.load_icon_menu(os.path.join(self.assets_path, "info.png"))
        self.help_icon = self.load_icon_menu(os.path.join(self.assets_path, "help.png"))
        self.dashboard_icon = self.load_icon_menu(os.path.join(self.assets_path, "menu.png"))
        self.download_icon = self.load_icon_menu(os.path.join(self.assets_path, "download.png"))

        # Load or create images for checked and unchecked states
        self.unchecked_image = self.load_icon_menu(os.path.join(self.assets_icon_path, "blank-check-box.png"))
        self.checked_image = self.load_icon_menu(os.path.join(self.assets_icon_path, "square.png"))
        self.switch_on_image = self.load_icon_menu(os.path.join(self.assets_icon_path, "switch-on.png"),size=(64, 64))
        self.switch_off_image = self.load_icon_menu(os.path.join(self.assets_icon_path, "switch-off.png"),size=(64, 64))

        # Create side menu labels and buttons
        self.side_menu_heading = ttk.Label(
            self.side_menu_frame,
            text="ADE BMS",
            font=("Courier New", 18, "bold"),
            bootstyle="inverse-info",  # Change color style
            anchor="center"
        )
        self.side_menu_heading.pack(fill="x", pady=(10, 20))

        # Add dashboard, control, info, report, and help buttons to the side menu
        self.dashboard_button = ttk.Button(
            self.side_menu_frame,
            text=" Dashboard     ",
            command=lambda: self.select_button(self.dashboard_button, self.show_dashboard),
            image=self.dashboard_icon,
            compound="left",  # Icon on the left
            bootstyle="info"
        )
        self.dashboard_button.pack(fill="x", pady=5)

        # Control button in the side menu
        self.control_button = ttk.Button(
            self.side_menu_frame,
            text=" Control          ",
            command=lambda: self.select_button(self.control_button, self.show_new_control),
            image=self.control_icon,
            compound="left",  # Icon on the left
            bootstyle="info"
        )
        self.control_button.pack(fill="x", pady=2)

        # Info button in the side menu
        self.info_button = ttk.Button(
            self.side_menu_frame,
            text=" Info                ",
            command=lambda: self.select_button(self.info_button, self.show_info),
            image=self.info_icon,
            compound="left",  # Icon on the left
            bootstyle="info"
        )
        self.info_button.pack(fill="x", pady=5)

        # Report button in the side menu
        self.report_button = ttk.Button(
            self.side_menu_frame,
            text=" Report            ",
            command=lambda: self.select_button(self.report_button, self.show_report),
            image=self.download_icon,
            compound="left",  # Icon on the left
            bootstyle="info"
        )
        self.report_button.pack(fill="x", pady=2)

        # Additional configuration and help buttons (conditionally displayed)
        if self.limited:
            self.config_button = ttk.Button(
                self.side_menu_frame,
                text=" Configuration",
                command=lambda: self.select_button(self.config_button, self.show_config),
                image=self.config_icon,
                compound="left",  # Icon on the left
                bootstyle="info"
            )
            self.config_button.pack(fill="x", pady=2)

        self.help_button = ttk.Button(
            self.side_menu_frame,
            text=" Help                ",
            command=lambda: self.select_button(self.help_button),
            image=self.help_icon,
            compound="left",  # Icon on the left
            bootstyle="info"
        )
        self.help_button.pack(fill="x", pady=(0, 10))

        # Mode selection dropdown
        self.mode_label = ttk.Label(self.side_menu_frame, bootstyle="inverse-info", text="Mode:")
        self.mode_label.pack(pady=(110, 5), padx=(0, 5))

        # Mode Dropdown
        self.mode_dropdown = ttk.Combobox(
            self.side_menu_frame,
            textvariable=self.mode_var,
            values=["Testing", "Maintenance"],
            state="readonly"
        )
        self.mode_dropdown.pack(padx=(5, 20))
        self.mode_dropdown.bind("<<ComboboxSelected>>", self.update_mode)

        # Battery selection dropdown (conditionally displayed based on the data)
        self.battery_selection_label = ttk.Label(self.side_menu_frame, text="Select Battery:", bootstyle="inverse-info")
        self.battery_var = tk.StringVar(value="Battery 1")  # Default to Battery 1
        self.battery_dropdown = ttk.Combobox(
            self.side_menu_frame,
            textvariable=self.battery_var,
            values=["Battery 1", "Battery 2"],
            state="readonly"
        )
        self.battery_dropdown.bind("<<ComboboxSelected>>", self.update_battery_selection)

        if device_data_battery_2['serial_number'] == 0:
            # Hide battery selection if only Battery 1 is available
            self.battery_selection_label.pack_forget()
            self.battery_dropdown.pack_forget()
        else:
            # Show battery selection if both batteries are available
            self.battery_selection_label.pack(pady=(10, 5))
            self.battery_dropdown.pack(padx=(5, 20), pady=(0, 10))

        # Disconnect button in the side menu
        self.disconnect_button = ttk.Button(
            self.side_menu_frame,
            text=" Disconnect",
            command=lambda: self.select_button(self.disconnect_button, self.on_disconnect),
            image=self.disconnect_icon,
            compound="left",  # Icon on the left
            bootstyle="danger"
        )
        self.disconnect_button.pack(side="bottom", pady=20)

        # Check battery status flags and current to set charge and discharge FET statuses
        if self.battery_status_flags.get('charge_fet_test') == 0:
            self.charge_fet_status = True
        else:
            self.charge_fet_status = False
        if self.device_data['current'] > 0:
            self.discharge_fet_status = True
        else:
            self.discharge_fet_status = False

        # Show dashboard by default and bind window resize event
        self.show_dashboard()
        self.master.bind('<Configure>', self.on_window_resize)
    
    def update_battery_selection(self, event=None):
        """Update the controls based on the selected battery."""
        self.selected_battery = self.battery_var.get()
        if self.selected_battery == "Battery 1":
            self.device_data = device_data_battery_1
            self.battery_status_flags = battery_1_status_flags
            self.charger_control_var = self.charger_control_var_battery_1
            self.discharger_control_var = self.discharger_control_var_battery_1
        elif self.selected_battery == "Battery 2":
            self.device_data = device_data_battery_2
            self.battery_status_flags = battery_2_status_flags
            self.charger_control_var = self.charger_control_var_battery_2
            self.discharger_control_var = self.discharger_control_var_battery_2
            self.check_battery_log_for_cycle_count()

        if self.selected_button == self.dashboard_button:
            self.show_dashboard()
        elif self.selected_button == self.info_button:
            self.show_info()
        elif self.selected_button == self.control_button:
            self.show_new_control()
        elif self.selected_button == self.report_button:
            self.show_report()
    
    def clear_control_content_frame(self):
        """Clear the content frame for new content."""
        for widget in self.control_content_frame.winfo_children():
            widget.destroy()

    def show_new_control(self):
        """Show the load control screen."""
        self.clear_content_frame()

        control_panel_frame = ttk.Frame(self.content_frame, borderwidth=10, relief="solid")
        control_panel_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

         # Function to update the display based on the control variable
        def update_control_frame():
            if self.control_var.get() == "Charging":
                self.chroma_device_status.set(False)
                self.chroma_mcb_status.set(False)
                self.agree_discharging_status.set(False)
                self.charging_button.config(style="Selected.TButton")
                self.discharging_button.config(style="TButton")
                self.show_charging_frame()
               
            elif self.control_var.get() == "Discharging":
                self.skyla_device_status.set(False)
                self.skyla_mcb_status.set(False)
                self.agree_charging_status.set(False)
                self.discharging_button.config(style="Selected.TButton")
                self.charging_button.config(style="TButton")
                self.show_discharging_frame()

        # Function to set control_var to Charging and update the frame
        def select_charging():
            self.control_var.set("Charging")
            update_control_frame()
        
        # Function to set control_var to Discharging and update the frame
        def select_discharging():
            self.control_var.set("Discharging")
            update_control_frame()

        # Set up styles for selected and unselected buttons
        style = ttk.Style()
        style.configure("TButton", font=("Arial", 12), padding=10)
        style.configure("Selected.TButton", font=("Arial", 12, "bold"), background="lightblue", padding=10)

        # Charging Button (set as default)
        self.charging_button = ttk.Button(control_panel_frame, text="Charging", command=select_charging, style="Selected.TButton")
        self.charging_button.grid(row=0, column=0, columnspan=6, padx=195, pady=10, sticky="ew")

        # Discharging Button
        self.discharging_button = ttk.Button(control_panel_frame, text="Discharging", command=select_discharging)
        self.discharging_button.grid(row=0, column=6, columnspan=6, padx=195, pady=10, sticky="ew")


        # Content frame for checkboxes and agree button
        self.control_content_frame = ttk.Frame(control_panel_frame)
        self.control_content_frame.grid(row=4, column=0, columnspan=12, rowspan=8, padx=10, pady=10, sticky="nsew")

        # Show charging frame by default
        update_control_frame()

    def show_charging_frame(self):
        """Display the Charging section."""
        self.clear_control_content_frame()

        # Create a tk.Label to act as the top border
        top_border = tk.Label(self.control_content_frame, background="black", width=100, height=1)
        top_border.grid(row=0, column=0, columnspan=12, sticky="ew")

        # Configure grid weights for centering
        self.control_content_frame.grid_columnconfigure(0, weight=1)  # Left spacer
        self.control_content_frame.grid_columnconfigure(6, weight=1)  # Center column
        self.control_content_frame.grid_columnconfigure(12, weight=1)  # Right spacer
        
        title_label = ttk.Label(self.control_content_frame, text="Charging", font=("Arial", 24, "bold"))
        title_label.grid(row=0, column=4, padx=250, pady=5, sticky="nsew")
        sub_title_label = ttk.Label(self.control_content_frame, text="Manual Intervention Checks", font=("Arial", 18, "bold"))
        sub_title_label.grid(row=1, column=4, padx=150, pady=5, sticky="nsew")

        # Checkbox variables stored in self to maintain their state across refreshes
        self.charger_status = tk.BooleanVar(value=self.skyla_device_status.get())  # For "Turn ON Skyla Charger"
        self.charger_mcb_status = tk.BooleanVar(value=self.skyla_mcb_status.get())  # For "Turn ON MCB Switch"

        checkbox1 = tk.Checkbutton(self.control_content_frame, text="Turn ON Skyla Charger", font=("Helvetica", 16), variable=self.charger_status, image=self.unchecked_image,  # Set the unchecked image
        selectimage=self.checked_image,  # Set the checked image
        indicatoron=False,  # Hide the default checkbox indicator
        compound='left',
        padx=10, pady=0,  # Remove internal padding
        bd=0,  # Remove border
        anchor="w"  # Align text and image to the left
        )
        checkbox1.grid(row=2, column=2, columnspan=4, padx=10, pady=20, sticky="ew")

        checkbox2 = tk.Checkbutton(self.control_content_frame, text="Turn ON MCB Switch", font=("Helvetica", 16), variable=self.charger_mcb_status, image=self.unchecked_image,  # Set the unchecked image
        selectimage=self.checked_image,  # Set the checked image
        indicatoron=False,  # Hide the default checkbox indicator
        compound='left',
        padx=10, pady=0,  # Remove internal padding
        bd=0,  # Remove border
        anchor="w"  # Align text and image to the left
        )
        checkbox2.grid(row=4, column=2, columnspan=4, padx=10, pady=20, sticky="ew")

        # Agree Button (disabled by default)
        self.charging_agree_button = ttk.Button(self.control_content_frame, text="Agree", state="disabled", command=lambda:self.show_charger_control())
        self.charging_agree_button.grid(row=6, column=2,padx=20, pady=30)
        if self.charger_status.get() and self.charger_mcb_status.get():
                self.charging_agree_button.config(state="normal")

        # Function to enable/disable agree button based on checkbox states
        def update_agree_button(*args):
            self.skyla_device_status.set(self.charger_status.get())
            self.skyla_mcb_status.set(self.charger_mcb_status.get())
            if self.charger_status.get() and self.charger_mcb_status.get():
                self.charging_agree_button.config(state="normal")
            else:
                self.charging_agree_button.config(state="disabled")

        # Track changes in checkbox states
        self.charger_status.trace_add("write", update_agree_button)
        self.charger_mcb_status.trace_add("write", update_agree_button)

        # Create a style for the Checkbutton
        style = ttk.Style()
        style.configure("CustomCheckbutton.TCheckbutton", font=("Helvetica", 20))  # Set desired font size

        # Initialize charger control label and button (hidden initially)
        self.charge_button_battery_1 = tk.Checkbutton(self.control_content_frame, text="Charging On/Off", font=("Helvetica", 16),
                                                        variable=self.charger_control_var_battery_1, 
                                                        image=self.switch_off_image,  # Set the unchecked image
                                                        selectimage=self.switch_on_image,  # Set the checked image
                                                        indicatoron=False,  # Hide the default checkbox indicator
                                                        command=lambda: self.toggle_button_style(self.charger_control_var_battery_1, self.charge_button_battery_1, 'charge', self.selected_battery),
                                                        compound='left',
                                                        padx=10, pady=0,  # Remove internal padding
                                                        bd=0,  # Remove border
                                                        anchor="w"  # Align text and image to the left
                                                        )
        self.charge_button_battery_1.grid(row=8, column=2, columnspan=4, rowspan= 2,  padx=10, pady=20, sticky="ew")
        agree_charging_status = self.agree_charging_status.get() 
        if agree_charging_status == False:
            self.charge_button_battery_1.grid_remove()

        test_label = ttk.Label(self.control_content_frame, text="")
        test_label.grid(row=10, column=2, rowspan= 2, padx=10, pady=5, sticky="nsew")

    def show_charger_control(self):
        # Show charger control elements
        self.agree_charging_status.set(True)
        self.charge_button_battery_1.grid(row=8, column=2, columnspan=4,  rowspan= 2, padx=10, pady=20, sticky="ew")

    def show_discharging_frame(self):
        """Display the Charging section."""
        self.clear_control_content_frame()

        title_label = ttk.Label(self.control_content_frame, text="Dicharging", font=("Arial", 24, "bold"))
        title_label.grid(row=0, column=4, padx=250, pady=5, sticky="nsew")
        sub_title_label = ttk.Label(self.control_content_frame, text="Manual Intervention Checks", font=("Arial", 18, "bold"))
        sub_title_label.grid(row=1, column=4, padx=150, pady=5, sticky="nsew")

        # Checkbox variables stored in self to maintain their state across refreshes
        self.chroma_load_status = tk.BooleanVar(value=self.chroma_device_status.get())  # For "Turn ON Chroma Load"
        self.load_mcb_status = tk.BooleanVar(value=self.chroma_mcb_status.get())  # For "Turn ON MCB Switch"

        checkbox1 = tk.Checkbutton(self.control_content_frame, text="Turn ON Chroma Load", font=("Helvetica", 16), variable=self.chroma_load_status, image=self.unchecked_image,  # Set the unchecked image
        selectimage=self.checked_image,  # Set the checked image
        indicatoron=False,  # Hide the default checkbox indicator
        compound='left',
        padx=10, pady=0,  # Remove internal padding
        bd=0,  # Remove border
        anchor="w"  # Align text and image to the left
        )
        checkbox1.grid(row=2, column=2, columnspan=4, padx=10, pady=10, sticky="ew")

        checkbox2 = tk.Checkbutton(self.control_content_frame, text="Turn ON MCB Switch", font=("Helvetica", 16), variable=self.load_mcb_status, image=self.unchecked_image,  # Set the unchecked image
        selectimage=self.checked_image,  # Set the checked image
        indicatoron=False,  # Hide the default checkbox indicator
        compound='left',
        padx=10, pady=0,  # Remove internal padding
        bd=0,  # Remove border
        anchor="w"  # Align text and image to the left
        )
        checkbox2.grid(row=4, column=2, columnspan=4, padx=10, pady=10, sticky="ew")

        # Agree Button (disabled by default)
        self.discharging_agree_button = ttk.Button(self.control_content_frame, text="Agree", state="disabled", command=lambda:self.show_load_control())
        self.discharging_agree_button.grid(row=6, column=2, padx=20, pady=20)

        if self.chroma_load_status.get() and self.load_mcb_status.get():
                self.discharging_agree_button.config(state="normal")

        # Initialize style for custom Checkbutton
        style = ttk.Style()

        # Create a new style for the Checkbutton
        style.configure('Custom.TCheckbutton', font=("Helvetica", 18), padding=(20, 10))
       
        self.discharge_button_battery_1 = tk.Checkbutton(self.control_content_frame, text="Discharging On/Off", font=("Helvetica", 16),
                                                        variable=self.discharger_control_var_battery_1, 
                                                        image=self.switch_off_image,  # Set the unchecked image
                                                        selectimage=self.switch_on_image,  # Set the checked image
                                                        indicatoron=False,  # Hide the default checkbox indicator
                                                        command=lambda: self.toggle_button_style(self.discharger_control_var_battery_1, self.discharge_button_battery_1, 'discharge', self.selected_battery),
                                                        compound='left',
                                                        padx=10, pady=0,  # Remove internal padding
                                                        bd=0,  # Remove border
                                                        anchor="w"  # Align text and image to the left
                                                        )

        # Section: Right Control Frame (Change to Grid Layout)
        self.load_control_frame = ttk.Labelframe(self.control_content_frame, text="Load", bootstyle="dark", borderwidth=5, relief="solid")
        self.load_control_frame.grid(row=9, column=2, rowspan=1, columnspan=10, padx=5, pady=5, sticky="nsew")
        
        # Function to switch between Charging and Discharging
        def select_discharging_default_testing():
            self.load_default_button.configure(style="Selected.TButton")
            self.load_custom_button.configure(style="TButton")
            show_default_controls()
            # Show default control buttons and hide custom controls
            self.testing_default_fields_frame.grid()
            hide_custom_controls()

        def select_discharging_custom_testing():
            self.load_custom_button.configure(style="Selected.TButton")
            self.load_default_button.configure(style="TButton")
            # Call the show_custom_controls to initialize the custom elements if they don't exist
            show_custom_controls()
            # Show custom controls and hide default controls
            self.testing_custom_fields_frame.grid()
            hide_default_controls()

        # Set up styles for selected and unselected buttons
        style = ttk.Style()
        style.configure("TButton", font=("Arial", 12), padding=10)
        style.configure("Selected.TButton", font=("Arial", 12, "bold"), background="lightblue", padding=10)
        # Check the selected mode
        selected_mode = self.mode_var.get()
        if selected_mode == "Testing":
            # Content frame for checkboxes and agree button
            control_testing_frame = ttk.Frame(self.load_control_frame)
            control_testing_frame.grid(row=4, column=0, columnspan=12, rowspan=8, padx=10, pady=10, sticky="nsew")

            # Function to show default controls
            def show_default_controls():
                # Frame for dynamic fields (L1, L2, T1, T2, Repeat, Save button)
                self.testing_default_fields_frame = ttk.Frame(control_testing_frame)
                self.testing_default_fields_frame.grid(row=1, column=0, columnspan=6, padx=10, pady=10, sticky="nsew")
                self.load_default_button_control = ttk.Button(self.testing_default_fields_frame, text="Set 50A & Turn ON", command=set_l1_50a_and_turn_on)
                self.load_default_button_control.grid(row=1, column=0, columnspan=6, padx=150, pady=10, sticky="ew")
                self.load_custom_button_control = ttk.Button(self.testing_default_fields_frame, text="Turn OFF", command=turn_load_off)
                self.load_custom_button_control.grid(row=1, column=6, columnspan=6, padx=190, pady=10, sticky="ew")

            def hide_default_controls():
                if hasattr(self, 'testing_default_fields_frame') and self.testing_default_fields_frame.winfo_exists():
                    self.testing_default_fields_frame.grid_remove()
                
            # Function to show custom controls
            def show_custom_controls():
                # Frame for dynamic fields (L1, L2, T1, T2, Repeat, Save button)
                self.testing_custom_fields_frame = ttk.Frame(control_testing_frame)
                self.testing_custom_fields_frame.grid(row=1, column=0, columnspan=6, padx=10, pady=10, sticky="nsew")
                self.custom_label = ttk.Label(self.testing_custom_fields_frame, text="Set Custom Current (A):", font=("Helvetica", 12), width=20, anchor="e")
                self.custom_label.grid(row=1, column=2, columnspan=2, padx=50, pady=5, sticky="w")
                self.custom_current_entry = ttk.Entry(self.testing_custom_fields_frame)
                self.custom_current_entry.grid(row=1, column=4, columnspan=2, padx=50, pady=5, sticky="ew")
                self.custom_button = ttk.Button(self.testing_custom_fields_frame, text="Save", command=lambda: save_custom_value())
                self.custom_button.grid(row=1, column=6, columnspan=2, padx=50, pady=5, sticky="ew")
                self.load_toggle_button = ttk.Button(self.testing_custom_fields_frame, text="Turn ON Load", command=toggle_load, bootstyle='success')
                self.load_toggle_button.grid(row=1, column=8, columnspan=2, padx=50, pady=5, sticky="ew")

            # Function to hide custom controls
            def hide_custom_controls():
                if hasattr(self, 'testing_custom_fields_frame') and self.testing_custom_fields_frame.winfo_exists():
                    self.testing_custom_fields_frame.grid_remove()

            # Set up the main buttons for selection between Default and Custom
            self.load_default_button = ttk.Button(self.load_control_frame, text="Default", command=lambda:select_discharging_default_testing(), style="Selected.TButton")
            self.load_default_button.grid(row=0, column=0, columnspan=6, padx=190, pady=10, sticky="ew")

            self.load_custom_button = ttk.Button(self.load_control_frame, text="Custom", command=lambda:select_discharging_custom_testing(), style="TButton")
            self.load_custom_button.grid(row=0, column=6, columnspan=6, padx=190, pady=10, sticky="ew")

             # Initially, show default controls
            select_discharging_default_testing()

            # Custom value save function
            def save_custom_value():
                """Saves the custom current value and sets it to the Chroma device."""
                custom_value = self.custom_current_entry.get()
                print(f"{custom_value} load value")
                if custom_value.isdigit():  # Basic validation to ensure it's a number
                    set_custom_l1_value(int(custom_value))
                else:
                    print("Invalid input: Please enter a valid number.")

            # Toggle Load On/Off function
            def toggle_load():
                if self.load_status.get():
                    turn_load_off()
                    self.load_toggle_button.config(text="Turn ON Load", bootstyle="success")  # Change to green when OFF
                    self.load_status.set(False)  # Update state to OFF
                else:
                    turn_load_on()
                    self.load_toggle_button.config(text="Turn OFF Load", bootstyle="danger")  # Change to red when ON
                    self.load_status.set(True)  # Update state to ON

        elif selected_mode == "Maintenance":
            # Content frame for checkboxes and agree button
            control_maintance_frame = ttk.Frame(self.load_control_frame)
            control_maintance_frame.grid(row=4, column=0, columnspan=12, rowspan=8, padx=10, pady=10, sticky="nsew")
            # Function to switch between Charging and Discharging
            def select_discharging_default_maintenance():
                self.load_default_maintance_button.configure(style="Selected.TButton")
                self.load_custom_maintance_button.configure(style="TButton")
                self.load_dynamic_button.configure(style="TButton")
                show_maintance_default_controls()
                self.maintance_default_fields_frame.grid()
                hide_maintance_custom_controls()
                hide_maintance_dynamic_controls()

            def select_discharging_custom_maintenance():
                self.load_custom_maintance_button.configure(style="Selected.TButton")
                self.load_default_maintance_button.configure(style="TButton")
                self.load_dynamic_button.configure(style="TButton")
                # Call the show_custom_controls to initialize the custom elements if they don't exist
                show_maintance_custom_controls()
                # Show custom controls and hide default controls
                # Initially hide the dynamic fields frame
                self.maintance_custom_fields_frame.grid()
                hide_maintance_default_controls()
                hide_maintance_dynamic_controls()

            def select_discharging_dynamic_maintenance():
                self.load_dynamic_button.configure(style="Selected.TButton")
                self.load_default_maintance_button.configure(style="TButton")
                self.load_custom_maintance_button.configure(style="TButton")
                # Call the show_custom_controls to initialize the custom elements if they don't exist
                show_maintance_dynamic_controls()
                # Show custom controls and hide default controls
                self.dynamic_fields_frame.grid()
                hide_maintance_custom_controls()
                hide_maintance_default_controls()

            # Function to show default controls
            def show_maintance_default_controls():
                # Frame for dynamic fields (L1, L2, T1, T2, Repeat, Save button)
                self.maintance_default_fields_frame = ttk.Frame(control_maintance_frame)
                self.maintance_default_fields_frame.grid(row=1, column=0, columnspan=6, padx=10, pady=5, sticky="nsew")

                # Initially hide the dynamic fields frame
                self.maintance_default_fields_frame.grid_remove()

                self.load_default_maintance_button_control = ttk.Button( self.maintance_default_fields_frame, text="Set 100A & Turn ON", command=set_l1_100a_and_turn_on)
                self.load_default_maintance_button_control.grid(row=1, column=0, columnspan=6, padx=140, pady=5, sticky="ew")

                self.load_custom_maintance_button_control = ttk.Button( self.maintance_default_fields_frame, text="Turn OFF", command=turn_load_off)
                self.load_custom_maintance_button_control.grid(row=1, column=6, columnspan=6, padx=180, pady=5, sticky="ew")

            def hide_maintance_default_controls():
                if hasattr(self, 'maintance_default_fields_frame') and self.maintance_default_fields_frame.winfo_exists():
                    self.maintance_default_fields_frame.grid_remove()

            # Function to show custom controls
            def show_maintance_custom_controls():
               # Frame for dynamic fields (L1, L2, T1, T2, Repeat, Save button)
                self.maintance_custom_fields_frame = ttk.Frame(control_maintance_frame)
                self.maintance_custom_fields_frame.grid(row=1, column=0, columnspan=6, padx=10, pady=10, sticky="nsew")

                # Initially hide the dynamic fields frame
                self.maintance_custom_fields_frame.grid_remove()

                self.maintance_custom_label = ttk.Label(self.maintance_custom_fields_frame, text="Set Custom Current (A):", font=("Helvetica", 12), width=20, anchor="e")
                self.maintance_custom_label.grid(row=1, column=2, columnspan=2, padx=50, pady=5, sticky="w")
                self.maintance_custom_current_entry = ttk.Entry(self.maintance_custom_fields_frame)
                self.maintance_custom_current_entry.grid(row=1, column=4, columnspan=2, padx=50, pady=5, sticky="ew")
                self.maintance_custom_button = ttk.Button(self.maintance_custom_fields_frame, text="Save", command=lambda: save_custom_value())
                self.maintance_custom_button.grid(row=1, column=6, columnspan=2, padx=50, pady=5, sticky="ew")
                self.maintance_load_toggle_button = ttk.Button(self.maintance_custom_fields_frame, text="Turn ON Load", command=toggle_load, bootstyle='success')
                self.maintance_load_toggle_button.grid(row=1, column=8, columnspan=2, padx=50, pady=5, sticky="ew")

            # Function to hide custom controls
            def hide_maintance_custom_controls():
                if hasattr(self, 'maintance_custom_fields_frame') and self.maintance_custom_fields_frame.winfo_exists():
                    # Initially hide the dynamic fields frame
                    self.maintance_custom_fields_frame.grid_remove()

             # Function to show custom controls
            
            def show_maintance_dynamic_controls():
                # Frame for dynamic fields (L1, L2, T1, T2, Repeat, Save button)
                self.dynamic_fields_frame = ttk.Frame(control_maintance_frame)
                self.dynamic_fields_frame.grid(row=1, column=0, columnspan=6, padx=10, pady=5, sticky="nsew")

                # Initially hide the dynamic fields frame
                self.dynamic_fields_frame.grid_remove()

                # L1 Entry
                ttk.Label(self.dynamic_fields_frame, text="L1 :").grid(row=0, column=0, padx=10, pady=5, sticky="e")
                self.l1_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
                self.l1_entry.grid(row=0, column=1, padx=20, pady=10, sticky="ew")

                # L2 Entry
                ttk.Label(self.dynamic_fields_frame, text="L2 :").grid(row=0, column=2, padx=10, pady=5, sticky="e")
                self.l2_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
                self.l2_entry.grid(row=0, column=3, padx=20, pady=10, sticky="ew")

                # T1 Entry
                ttk.Label(self.dynamic_fields_frame, text="T1 :").grid(row=0, column=4, padx=10, pady=5, sticky="e")
                self.t1_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
                self.t1_entry.grid(row=0, column=5, padx=20, pady=10, sticky="ew")

                # T2 Entry
                ttk.Label(self.dynamic_fields_frame, text="T2 :").grid(row=0, column=6, padx=10, pady=5, sticky="e")
                self.t2_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
                self.t2_entry.grid(row=0, column=7, padx=20, pady=10, sticky="ew")

                # Repeat Entry
                ttk.Label(self.dynamic_fields_frame, text="Repeat :").grid(row=0, column=8, padx=10, pady=5, sticky="e")
                self.repeat_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
                self.repeat_entry.grid(row=0, column=9, padx=20, pady=10, sticky="ew")

                # Save Button
                save_button = ttk.Button(self.dynamic_fields_frame, text="Save", command=lambda: save_dynamic_fields())
                save_button.grid(row=0, column=10, padx=25, pady=5, sticky="ew")

                def toggle_load_on_and_off():
                    if self.load_status.get():
                        turn_load_off()
                        toggle_button_on.config(text="Turn ON Load", bootstyle="success")  # Change to green when OFF
                        self.load_status.set(False)  # Update state to OFF
                    else:
                        turn_load_on()
                        toggle_button_on.config(text="Turn OFF Load", bootstyle="danger")  # Change to red when ON
                        self.load_status.set(True)  # Update state to ON

                # Turn ON Load Button (in the same row as dynamic fields)
                toggle_button_on = ttk.Button(self.dynamic_fields_frame, text="Turn ON Load", command=lambda:toggle_load_on_and_off(), bootstyle='success')
                toggle_button_on.grid(row=0, column=11, columnspan=2, padx=5, pady=5, sticky="ew")


                def save_dynamic_fields():
                    # Collect values from the entry boxes
                    l1_value = self.l1_entry.get()
                    l2_value = self.l2_entry.get()
                    t1_value = self.t1_entry.get()
                    t2_value = self.t2_entry.get()
                    repeat_value = self.repeat_entry.get()

                    # Call the method to set values in Chroma
                    set_dynamic_values(l1_value, l2_value, t1_value, t2_value, repeat_value)

                # Expand columns for even distribution
                for i in range(11):
                    self.dynamic_fields_frame.grid_columnconfigure(i, weight=1)
    
            # Function to hide custom controls
            def hide_maintance_dynamic_controls():
                if hasattr(self, 'dynamic_fields_frame') and self.dynamic_fields_frame.winfo_exists():
                     self.dynamic_fields_frame.grid_remove()  # Hide the dynamic fields


            # Set up the main buttons for selection between Default and Custom
            self.load_default_maintance_button = ttk.Button(self.load_control_frame, text="Default", command=lambda:select_discharging_default_maintenance(), style="Selected.TButton")
            self.load_default_maintance_button.grid(row=0, column=0, columnspan=4, padx=85, pady=10, sticky="ew")

            self.load_custom_maintance_button = ttk.Button(self.load_control_frame, text="Custom", command=lambda:select_discharging_custom_maintenance(), style="TButton")
            self.load_custom_maintance_button.grid(row=0, column=4, columnspan=4, padx=85, pady=10, sticky="ew")

            self.load_dynamic_button = ttk.Button(self.load_control_frame, text="Dynamic", command=lambda:select_discharging_dynamic_maintenance(), style="TButton")
            self.load_dynamic_button.grid(row=0, column=8, columnspan=4, padx=75, pady=10, sticky="ew")

            select_discharging_default_maintenance()
            
           # Custom value save function
            def save_custom_value():
                """Saves the custom current value and sets it to the Chroma device."""
                custom_value = self.custom_current_entry.get()
                print(f"{custom_value} load value")
                if custom_value.isdigit():  # Basic validation to ensure it's a number
                    set_custom_l1_value(int(custom_value))
                else:
                    print("Invalid input: Please enter a valid number.")

            # Toggle Load On/Off function
            def toggle_load():
                if self.load_status.get():
                    turn_load_off()
                    self.load_toggle_button.config(text="Turn ON Load", bootstyle="success")  # Change to green when OFF
                    self.load_status.set(False)  # Update state to OFF
                else:
                    turn_load_on()
                    self.load_toggle_button.config(text="Turn OFF Load", bootstyle="danger")  # Change to red when ON
                    self.load_status.set(True)  # Update state to ON

        # Function to enable/disable agree button based on checkbox states
        def update_agree_button(*args):
            self.chroma_device_status.set(self.chroma_load_status.get())
            self.chroma_mcb_status.set(self.load_mcb_status.get())
            if self.chroma_load_status.get() and self.load_mcb_status.get():
                self.discharging_agree_button.config(state="normal")
            else:
                self.discharging_agree_button.config(state="disabled")

        # Track changes in checkbox states
        self.chroma_load_status.trace_add("write", update_agree_button)
        self.load_mcb_status.trace_add("write", update_agree_button)

        self.discharge_button_battery_1.grid(row=7, column=2, columnspan=4, rowspan= 2,  padx=10, pady=5, sticky="ew")
        agree_discharging_status = self.agree_discharging_status.get()
        if agree_discharging_status == False:
            self.discharge_button_battery_1.grid_remove()
            self.load_control_frame.grid_remove()
    
    def show_load_control(self):
        # Show charger control elements
        # self.connect_device()
        self.agree_discharging_status.set(True)
        self.load_control_frame.grid(row=9, column=2, rowspan=1, columnspan=10, padx=5, pady=5, sticky="nsew")
        self.discharge_button_battery_1.grid(row=7, column=2, columnspan=4, rowspan= 2,  padx=10, pady=5, sticky="ew")

    def show_control(self):
        """Show the load control screen."""
        self.clear_content_frame()

        load_frame = ttk.Frame(self.content_frame, borderwidth=10, relief="solid")
        load_frame.grid(row=0, column=0, columnspan=8, rowspan=8, padx=10, pady=10, sticky="nsew")

        # Section: Right Control Frame (Change to Grid Layout)
        right_control_frame = ttk.Labelframe(load_frame, text="Controls", bootstyle="dark", borderwidth=5, relief="solid")
        right_control_frame.grid(row=0, column=1, rowspan=1, padx=5, pady=5, sticky="nsew")  # Positioned to the right

        # Load icons
        self.reset_icon = self.load_icon(os.path.join(self.assets_path, "reset.png"))
        self.activate_icon = self.load_icon(os.path.join(self.assets_path, "activate.png"))

        if self.selected_battery == 'Battery 1':
            # Charger Control
            charger_control_label = ttk.Label(right_control_frame, text="Charging On/Off", font=("Helvetica", 12))
            charger_control_label.grid(row=0, column=0, padx=10, pady=5)
            self.charge_button_battery_1 = ttk.Checkbutton(
                right_control_frame,
                variable=self.charger_control_var,
                bootstyle="success-round-toggle" if self.charger_control_var_battery_1.get() else "danger-round-toggle",
                command=lambda: self.toggle_button_style(self.charger_control_var_battery_1, self.charge_button_battery_1, 'charge', self.selected_battery)
            )
            self.charge_button_battery_1.grid(row=0, column=1, padx=10, pady=5)

            # Discharger Control
            discharger_control_label = ttk.Label(right_control_frame, text="Discharging On/Off", font=("Helvetica", 12))
            discharger_control_label.grid(row=1, column=0, padx=10, pady=5)
            self.discharge_button_battery_1 = ttk.Checkbutton(
                right_control_frame,
                variable=self.discharger_control_var,
                bootstyle="success-round-toggle" if self.discharger_control_var_battery_1.get() else "danger-round-toggle",
                command=lambda: self.toggle_button_style(self.discharger_control_var_battery_1, self.discharge_button_battery_1, 'discharge', self.selected_battery)
            )
            self.discharge_button_battery_1.grid(row=1, column=1, padx=10, pady=5)
        elif self.selected_battery == 'Battery 2':
            # Charger Control
            charger_control_label = ttk.Label(right_control_frame, text="Charging On/Off", font=("Helvetica", 12))
            charger_control_label.grid(row=0, column=0, padx=10, pady=5)
            self.charge_button_battery_2 = ttk.Checkbutton(
                right_control_frame,
                variable=self.charger_control_var,
                bootstyle="success-round-toggle" if self.charger_control_var_battery_2.get() else "danger-round-toggle",
                command=lambda: self.toggle_button_style(self.charger_control_var_battery_2, self.charge_button_battery_2, 'charge', self.selected_battery)
            )
            self.charge_button_battery_2.grid(row=0, column=1, padx=10, pady=5)
    
            # Discharger Control
            discharger_control_label = ttk.Label(right_control_frame, text="Discharging On/Off", font=("Helvetica", 12))
            discharger_control_label.grid(row=1, column=0, padx=10, pady=5)
            self.discharge_button_battery_2 = ttk.Checkbutton(
                right_control_frame,
                variable=self.discharger_control_var,
                bootstyle="success-round-toggle" if self.discharger_control_var_battery_2.get() else "danger-round-toggle",
                command=lambda: self.toggle_button_style(self.discharger_control_var_battery_2, self.discharge_button_battery_2, 'discharge', self.selected_battery)
            )
            self.discharge_button_battery_2.grid(row=1, column=1, padx=10, pady=5)

        # Row 2: BMS Reset Button
        self.bms_reset_button = ctk.CTkButton(
            right_control_frame, text="Reset", image=self.reset_icon, compound="left",
            fg_color="#ff6361", hover_color="#d74a49"
        )
        self.bms_reset_button.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        # Row 3: Activate Heater Button
        self.activate_heater_button = ctk.CTkButton(
            right_control_frame, text="Activate Heater", image=self.activate_icon, compound="left",
            command=self.activate_heater, fg_color="#ff6361", hover_color="#d74a49"
        )
        self.activate_heater_button.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        # Section:  Device Connection
        device_connect_frame = ttk.Labelframe(load_frame, bootstyle='dark', text="Connection", padding=10, borderwidth=10, relief="solid")
        device_connect_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Connect Button
        self.connect_button = ttk.Button(device_connect_frame, text="Connect to Chroma", command=self.connect_device)
        self.connect_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        # Status Label
        self.status_label = ttk.Label(device_connect_frame, text="Status: Disconnected")
        self.status_label.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Make columns expand to fill space
        device_connect_frame.grid_columnconfigure(0, weight=1)  # First column (Connect button)
        device_connect_frame.grid_columnconfigure(1, weight=1)  # Second column (Status Label)
        device_connect_frame.grid_columnconfigure(2, weight=1)  # Third column (Download Report button)

        # Check the selected mode
        selected_mode = self.mode_var.get()

        if selected_mode == "Testing":
            # Section: Testing Mode
            testing_frame = ttk.Labelframe(load_frame, text="Testing", padding=10, borderwidth=10, relief="solid", bootstyle='dark')
            testing_frame.grid(row=1, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

            test_button = ttk.Button(testing_frame, text="Set 50A and Turn ON Load", command=set_l1_50a_and_turn_on)
            test_button.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

            # Turn Off Load Button
            turn_off_button = ttk.Button(testing_frame, text="Turn OFF Load", command=turn_load_off)
            turn_off_button.grid(row=0, column=1, columnspan=2, padx=10, pady=10, sticky="ew")

        elif selected_mode == "Maintenance":
            # Section: Maintenance Mode
            maintenance_frame = ttk.Labelframe(load_frame, text="Maintenance", padding=10, borderwidth=10, relief="solid", bootstyle='dark')
            maintenance_frame.grid(row=1, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

            # Add a checkbox for dynamic fields
            dynamic_fields_checkbox = ttk.Checkbutton(
                maintenance_frame,
                text="Dynamic",
                variable=self.show_dynamic_fields_var,
                command=self.toggle_dynamic_fields  # Method to show/hide fields
            )
            dynamic_fields_checkbox.grid(row=0, column=0, padx=10, pady=10, sticky="w")

            # The "Set 100A and Turn ON Load" button
            self.maintenance_button = ttk.Button(maintenance_frame, text="Set 100A and Turn ON Load", command=set_l1_100a_and_turn_on)
            self.maintenance_button.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

            # The "Turn OFF Load" button
            self.turn_off_button = ttk.Button(maintenance_frame, text="Turn OFF Load", command=turn_load_off)
            self.turn_off_button.grid(row=1, column=2, padx=10, pady=10, sticky="ew")

            # Frame for dynamic fields (L1, L2, T1, T2, Repeat, Save button)
            self.dynamic_fields_frame = ttk.Frame(maintenance_frame)
            self.dynamic_fields_frame.grid(row=1, column=0, columnspan=6, padx=10, pady=10, sticky="nsew")

            # Initially hide the dynamic fields frame
            self.dynamic_fields_frame.grid_remove()

            # L1 Entry
            ttk.Label(self.dynamic_fields_frame, text="L1 :").grid(row=0, column=0, padx=10, pady=10, sticky="e")
            self.l1_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
            self.l1_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

            # L2 Entry
            ttk.Label(self.dynamic_fields_frame, text="L2 :").grid(row=0, column=2, padx=10, pady=10, sticky="e")
            self.l2_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
            self.l2_entry.grid(row=0, column=3, padx=10, pady=10, sticky="ew")

            # T1 Entry
            ttk.Label(self.dynamic_fields_frame, text="T1 :").grid(row=0, column=4, padx=10, pady=10, sticky="e")
            self.t1_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
            self.t1_entry.grid(row=0, column=5, padx=10, pady=10, sticky="ew")

            # T2 Entry
            ttk.Label(self.dynamic_fields_frame, text="T2 :").grid(row=0, column=6, padx=10, pady=10, sticky="e")
            self.t2_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
            self.t2_entry.grid(row=0, column=7, padx=10, pady=10, sticky="ew")

            # Repeat Entry
            ttk.Label(self.dynamic_fields_frame, text="Repeat :").grid(row=0, column=8, padx=10, pady=10, sticky="e")
            self.repeat_entry = ttk.Entry(self.dynamic_fields_frame, width=5)
            self.repeat_entry.grid(row=0, column=9, padx=10, pady=10, sticky="ew")

            # Save Button
            save_button = ttk.Button(self.dynamic_fields_frame, text="Save", command=lambda: save_dynamic_fields())
            save_button.grid(row=0, column=10, padx=10, pady=10, sticky="ew")

            def toggle_load_on_and_off():
                if self.load_status.get():
                    turn_load_off()
                    toggle_button_on.config(text="Turn ON Load", bootstyle="success")  # Change to green when OFF
                    self.load_status.set(False)  # Update state to OFF
                else:
                    turn_load_on()
                    toggle_button_on.config(text="Turn OFF Load", bootstyle="danger")  # Change to red when ON
                    self.load_status.set(True)  # Update state to ON

            # Turn ON Load Button (in the same row as dynamic fields)
            toggle_button_on = ttk.Button(self.dynamic_fields_frame, text="Turn ON Load", command=lambda:toggle_load_on_and_off(), bootstyle='success')
            toggle_button_on.grid(row=0, column=11, columnspan=2, padx=5, pady=5, sticky="ew")

            
            def save_dynamic_fields():
                # Collect values from the entry boxes
                l1_value = self.l1_entry.get()
                l2_value = self.l2_entry.get()
                t1_value = self.t1_entry.get()
                t2_value = self.t2_entry.get()
                repeat_value = self.repeat_entry.get()

                # Call the method to set values in Chroma
                set_dynamic_values(l1_value, l2_value, t1_value, t2_value, repeat_value)

            # Expand columns for even distribution
            for i in range(11):
                self.dynamic_fields_frame.grid_columnconfigure(i, weight=1)
    

        # Section: Custom Mode
        custom_frame = ttk.Labelframe(load_frame, text="Custom", padding=10, borderwidth=10, relief="solid", bootstyle='dark')
        custom_frame.grid(row=3, column=0,columnspan=4, padx=5, pady=5, sticky="nsew")

        custom_label = ttk.Label(custom_frame, text="Set Custom Current (A):", font=("Helvetica", 12), width=20, anchor="e")
        custom_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        def save_custom_value():
            """Saves the custom current value and sets it to the Chroma device."""
            custom_value = self.custom_current_entry.get()
            print(f"{custom_value} load value")
            if custom_value.isdigit():  # Basic validation to ensure it's a number
                set_custom_l1_value(int(custom_value))
            else:
                print("Invalid input: Please enter a valid number.")

        self.custom_current_entry = ttk.Entry(custom_frame)
        self.custom_current_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        custom_button = ttk.Button(custom_frame, text="Save", command=lambda:save_custom_value())
        custom_button.grid(row=0, column=2, columnspan=1, padx=5, pady=5, sticky="ew")

        # Create a toggle button for turning load ON and OFF
        self.load_status = tk.BooleanVar()  # Boolean variable to track load status (ON/OFF)
        self.load_status.set(False)  # Initial state is OFF

        def toggle_load():
            if self.load_status.get():
                turn_load_off()
                toggle_button.config(text="Turn ON Load", bootstyle="success")  # Change to green when OFF
                self.load_status.set(False)  # Update state to OFF
            else:
                turn_load_on()
                toggle_button.config(text="Turn OFF Load", bootstyle="danger")  # Change to red when ON
                self.load_status.set(True)  # Update state to ON

        # Toggle button for Load ON/OFF
        toggle_button = ttk.Button(custom_frame, text="Turn ON Load", command=toggle_load, bootstyle='success')
        toggle_button.grid(row=0, column=3, columnspan=2, padx=5, pady=5, sticky="ew")

        # Make the rows and columns expand as needed for content_frame and load_frame
        self.content_frame.grid_rowconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(1, weight=1)  # Ensures the load_frame expands fully
        self.content_frame.grid_columnconfigure(0, weight=1)

        load_frame.grid_rowconfigure(0, weight=1)  # Device Connection frame
        load_frame.grid_rowconfigure(1, weight=1)  # Testing frame
        load_frame.grid_rowconfigure(2, weight=1)  # Maintenance frame
        load_frame.grid_rowconfigure(3, weight=1)  # Custom frame
        load_frame.grid_columnconfigure(0, weight=1)  # Ensure columns expand to fill space

        # Highlight the control button in the side menu
        self.select_button(self.control_button)

    def connect_device(self):
        """Connect to the Chroma device."""
        result = find_and_connect()  # This can be the same method or a separate one for specific device connection

    def show_config(self):
        """Show the configuration settings page where users can view and update configurations."""
        self.clear_content_frame()  # Clear the content area before loading the config page

        # Create a frame for the config page content
        config_frame = ttk.Frame(self.content_frame, borderwidth=10, relief="solid")
        config_frame.pack(fill="x")

        # Create a new Labelframe for configuration settings
        config_labelframe = ttk.Labelframe(self.content_frame, text="Configuration Settings", bootstyle="dark", borderwidth=10, relief="solid")
        config_labelframe.pack(fill="x", padx=10, pady=10)

        # Validation functions for float and int
        def validate_numeric_input(value):
            if value == "" or value.isdigit() or value == "-" or value.replace(".", "", 1).isdigit():
                return True
            return False

        vcmd = (self.content_frame.register(validate_numeric_input), "%P")

        # Define configuration fields, using values from the config file
        config_values = {
            "Logging Time (Mins)": config.config_values['can_config'].get('logging_time'),
            "Discharge Max Current (A)": config.config_values['can_config'].get('discharging_current_max'),  # Get value from config file
            "Discharge Cutoff Voltage (V)": config.config_values['can_config'].get('discharge_cutoff_volt'),
            "Charge Cutoff Current (A)": config.config_values['can_config'].get('charging_cutoff_curr'),
            "Charge Cutoff Voltage (V)": config.config_values['can_config'].get('charging_cutoff_volt'),
            "Charge Cutoff Capacity (%)": config.config_values['can_config'].get('charging_cutoff_capacity'),
            "Temperature Caution (°C)": config.config_values['can_config'].get('temperature_caution'),
            "Temperature Alarm (°C)": config.config_values['can_config'].get('temperature_alarm'),
            "Temperature Critical (°C)": config.config_values['can_config'].get('temperature_critical')
        }

        # Create labels and entry fields for each config
        row = 0
        self.config_entries = {}  # Store entry fields to update values later
        for config_name, config_value in config_values.items():
            # Label for each configuration field
            ttk.Label(config_labelframe, text=f"{config_name}:", font=("Helvetica", 10, "bold")).grid(row=row, column=0, padx=10, pady=5, sticky="e")

            # Entry field for each configuration value
            config_entry = ttk.Entry(config_labelframe, validate="key", validatecommand=vcmd)
            config_entry.insert(0, config_value)  # Insert the existing value into the entry
            config_entry.grid(row=row, column=1, padx=10, pady=5, sticky="ew")

            # Save the entry fields in a dictionary for later access
            self.config_entries[config_name] = config_entry

            row += 1

        # Add an "Update" button to save the modified values inside the labelframe
        update_button = ttk.Button(config_labelframe, text="Update Configurations", command=self.update_config_values, bootstyle='success')
        update_button.grid(row=row, column=0, columnspan=2, padx=10, pady=20, sticky="ew")

        # Make the columns stretch to fill available space
        config_labelframe.grid_columnconfigure(0, weight=1)
        config_labelframe.grid_columnconfigure(1, weight=2)

        # Create a new Labelframe for project management
        project_frame = ttk.Labelframe(self.content_frame, text="Project Management", bootstyle="dark", borderwidth=10, relief="solid")
        project_frame.pack(fill="x", padx=10, pady=10)

        row += 1

        # Add the Combobox for the Project Dropdown
        ttk.Label(project_frame, text="Select Project:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=10, pady=5, sticky="e")
        self.project_var = tk.StringVar()
        self.project_combobox = ttk.Combobox(project_frame, textvariable=self.project_var, values=config.config_values['can_config']['projects'], state="readonly")
        self.project_combobox.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        self.project_combobox.set("Select a Project")

        # Add buttons for Add, Edit, and Delete project
        add_project_button = ttk.Button(project_frame, text="Add Project", command=self.add_project, bootstyle='success')
        add_project_button.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        edit_project_button = ttk.Button(project_frame, text="Edit Project", command=self.edit_project)
        edit_project_button.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        delete_project_button = ttk.Button(project_frame, text="Delete Project", command=self.delete_project, bootstyle='danger')
        delete_project_button.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

        # Make columns stretch to fill available space
        project_frame.grid_columnconfigure(0, weight=1)
        project_frame.grid_columnconfigure(1, weight=2)

    def update_config_values(self):
        """Update the configuration values and save them to the JSON config file."""

        def determine_number_type(value):
            """Determine if the value should be an int or float."""
            try:
                value_str = str(value)
                # First, try to convert it to an int
                if '.' in value_str:
                    return float(value)  # Return as float
                else:
                    return int(value)  # Return as int
            except ValueError:
                # If conversion fails, return None
                return None

        # Retrieve the updated values from the entry fields
        updated_values = {}
        for config_name, entry in self.config_entries.items():
            updated_values[config_name] = entry.get()

        # Update the configuration values in the config file with dynamic type handling
        config.config_values['can_config']['logging_time'] = determine_number_type(updated_values['Logging Time (Mins)'])
        config.config_values['can_config']['discharging_current_max'] = determine_number_type(updated_values['Discharge Max Current (A)'])
        config.config_values['can_config']['discharge_cutoff_volt'] = determine_number_type(updated_values['Discharge Cutoff Voltage (V)'])
        config.config_values['can_config']['charging_cutoff_curr'] = determine_number_type(updated_values['Charge Cutoff Current (A)'])
        config.config_values['can_config']['charging_cutoff_volt'] = determine_number_type(updated_values['Charge Cutoff Voltage (V)'])
        config.config_values['can_config']['charging_cutoff_capacity'] = determine_number_type(updated_values['Charge Cutoff Capacity (%)'])
        config.config_values['can_config']['temperature_caution'] = determine_number_type(updated_values['Temperature Caution (°C)'])
        config.config_values['can_config']['temperature_alarm'] = determine_number_type(updated_values['Temperature Alarm (°C)'])
        config.config_values['can_config']['temperature_critical'] = determine_number_type(updated_values['Temperature Critical (°C)'])

        # Save the updated configuration to the config file
        config.save_config()

        # Optionally show a message to the user indicating that the config has been updated
        messagebox.showinfo("Success", "Configuration updated and saved successfully.")
    
    def add_project(self):
        """Add a new project to the CAN project list using a custom input window."""

        # Create a new top-level window for input
        input_window = tk.Toplevel(self.master)
        input_window.title("Add Project")

        # Calculate the position for the window to be centered
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        window_width = 300
        window_height = 150
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)

        # Set the geometry of the window directly in the center
        input_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        input_window.transient(self.master)  # Set to always be on top of the parent window
        input_window.grab_set()  # Make the input window modal

        # Create and pack the widgets
        label = ttk.Label(input_window, text="Enter New Project Name:", font=("Helvetica", 12))
        label.pack(pady=10)

        project_name_var = tk.StringVar()  # Use StringVar to store the project name
        entry = ttk.Entry(input_window, textvariable=project_name_var)
        entry.pack(pady=5)

        # Function to handle OK button click
        def handle_ok():
            new_project_name = project_name_var.get().strip()
            if new_project_name and new_project_name not in config.config_values['can_config']['projects']:
                config.config_values['can_config']['projects'].append(new_project_name)
                config.save_config()  # Save the config with the new project
                self.project_combobox['values'] = config.config_values['can_config']['projects']  # Update the combobox options
                self.project_combobox.set(new_project_name)  # Select the newly added project
                input_window.destroy()  # Close the input window
            elif new_project_name in config.config_values['can_config']['projects']:
                messagebox.showwarning("Duplicate Project", f"Project '{new_project_name}' already exists.")
            else:
                messagebox.showwarning("Invalid Input", "Project name cannot be empty.")

        # OK Button
        ok_button = ttk.Button(input_window, text="OK", command=handle_ok)
        ok_button.pack(pady=10)

        # Focus the input window and entry
        input_window.focus()
        entry.focus_set()

        # Bind the Return key to trigger the OK button
        input_window.bind('<Return>', lambda event: handle_ok())

    def edit_project(self):
        """Edit the selected project name using a custom input window."""
        selected_project = self.project_combobox.get()

        # Check if a project is selected
        if not selected_project or selected_project == "Select a Project":
            messagebox.showwarning("No Project Selected", "Please select a project to edit.")
            return

        # Create a new top-level window for input
        input_window = tk.Toplevel(self.master)
        input_window.title("Edit Project")

        # Calculate the position for the window to be centered
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        window_width = 300
        window_height = 150
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)

        # Set the geometry of the window directly in the center
        input_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        input_window.transient(self.master)  # Set to always be on top of the parent window
        input_window.grab_set()  # Make the input window modal

        # Create and pack the widgets
        label = ttk.Label(input_window, text=f"Edit Project Name ({selected_project}):", font=("Helvetica", 12))
        label.pack(pady=10)

        project_name_var = tk.StringVar(value=selected_project)  # Pre-fill with current project name
        entry = ttk.Entry(input_window, textvariable=project_name_var)
        entry.pack(pady=5)

        # Function to handle OK button click
        def handle_ok():
            new_project_name = project_name_var.get().strip()
            if new_project_name and new_project_name not in config.config_values['can_config']['projects']:
                index = config.config_values['can_config']['projects'].index(selected_project)
                config.config_values['can_config']['projects'][index] = new_project_name.strip()
                config.save_config()  # Save the config with the edited project
                self.project_combobox['values'] = config.config_values['can_config']['projects']  # Update the combobox options
                self.project_combobox.set(new_project_name.strip())  # Set the newly edited project as selected
                input_window.destroy()  # Close the input window
            elif new_project_name in config.config_values['can_config']['projects']:
                messagebox.showwarning("Duplicate Project", f"Project '{new_project_name}' already exists.")
            else:
                messagebox.showwarning("Invalid Input", "Project name cannot be empty.")

        # OK Button
        ok_button = ttk.Button(input_window, text="OK", command=handle_ok)
        ok_button.pack(pady=10)

        # Focus the input window and entry
        input_window.focus()
        entry.focus_set()

        # Bind the Return key to trigger the OK button
        input_window.bind('<Return>', lambda event: handle_ok())
    
    def delete_project(self):
        """Delete the selected project from the CAN project list."""
        selected_project = self.project_combobox.get()

        # Check if a project is selected
        if not selected_project or selected_project == "Select a Project":
            messagebox.showwarning("No Project Selected", "Please select a project to delete.")
            return

        # Proceed with deleting if a project is selected
        if selected_project in config.config_values['can_config']['projects']:
            config.config_values['can_config']['projects'].remove(selected_project)
            config.save_config()  # Save the config with the removed project
            self.project_combobox['values'] = config.config_values['can_config']['projects']  # Update the combobox options
            self.project_combobox.set('Select a Project')  # Reset the combobox selection
            messagebox.showinfo("Deleted project", f"Project {selected_project} deleted successfully")

    def show_dashboard(self):
        self.clear_content_frame()

        # Create a Frame for the battery info details at the bottom
        info_frame = ttk.Labelframe(self.content_frame, text="Battery Information", bootstyle="dark", borderwidth=10, relief="solid")
        info_frame.pack(fill="x", expand=True, padx=5, pady=5)

        # Set consistent padding values
        padx_val = 2
        pady_val = 2

        # Device Name
        ttk.Label(info_frame, text="Device Name :", font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.device_name_label = ttk.Label(info_frame, text=self.device_data.get('device_name', 'N/A'), font=("Helvetica", 10))
        self.device_name_label.grid(row=0, column=1, sticky="w")

        # Serial Number
        ttk.Label(info_frame, text="Serial Number :", font=("Helvetica", 10, "bold")).grid(row=0, column=2, padx=padx_val, pady=pady_val, sticky="w")
        self.serial_number_label = ttk.Label(info_frame, text=self.device_data.get('serial_number', 'N/A'), font=("Helvetica", 10))
        self.serial_number_label.grid(row=0, column=3, sticky="w")

        # Manufacturer Name
        ttk.Label(info_frame, text="Manufacturer :", font=("Helvetica", 10, "bold")).grid(row=0, column=4, padx=padx_val, pady=pady_val, sticky="w")
        self.manufacturer_label = ttk.Label(info_frame, text=self.device_data.get('manufacturer_name', 'N/A'), font=("Helvetica", 10))
        self.manufacturer_label.grid(row=0, column=5, padx=padx_val, pady=pady_val, sticky="w")

        # Cycle Count
        ttk.Label(info_frame, text="Cycle Count :", font=("Helvetica", 10, "bold")).grid(row=0, column=6, padx=padx_val, pady=pady_val, sticky="w")
        self.cycle_count_label = ttk.Label(info_frame, text=self.device_data.get('cycle_count', 'N/A'), font=("Helvetica", 10))
        self.cycle_count_label.grid(row=0, column=7, padx=padx_val, pady=pady_val, sticky="w")

        # Charging Status
        ttk.Label(info_frame, text="Charging Status :", font=("Helvetica", 10, "bold")).grid(row=0, column=8, padx=padx_val, pady=pady_val, sticky="w")

        self.status_var = tk.StringVar()
        charging_bootstyle = 'dark'
        discharging_bootstyle = "dark"

        # Determine the status and set the corresponding style
        charging_status = self.device_data.get('charging_battery_status', 'Off').lower()
        if charging_status == "charging":
            bootstyle = "success"
            charging_bootstyle = 'success'
            self.status_var.set("charging")
        elif charging_status == "discharging":
            bootstyle = "danger"
            discharging_bootstyle = "danger"
            self.status_var.set("discharging")
        else:  # Assuming 'Idle' or any other status means the device is off
            bootstyle = "dark"
            charging_bootstyle = 'dark'
            discharging_bootstyle = "dark"
            self.status_var.set("off")

        # Create a Radiobutton to visually represent the status
        self.status_indicator = ttk.Radiobutton(
            info_frame,
            text=charging_status.capitalize(),  # Display the charging status text
            variable=self.status_var,  # Use the StringVar
            value=charging_status,  # Set the value that matches the current status
            bootstyle=bootstyle,  # Apply the bootstyle based on the charging status
            style="Custom.TRadiobutton",
        )
        self.status_indicator.grid(row=0, column=9, padx=padx_val, pady=pady_val, sticky="w")

        # Set the column weights for even expansion
        for col in range(10):
            info_frame.grid_columnconfigure(col, weight=1)
        
        # Create a main frame to hold Charging, Discharge, and Battery Health sections
        main_frame = ttk.Frame(self.content_frame)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Charging Section
        self.charging_frame = ttk.Labelframe(main_frame, text="Charging", bootstyle=charging_bootstyle, borderwidth=10, relief="solid")
        self.charging_frame.grid(row=0, column=0, columnspan=2, padx=5, pady=2, sticky="nsew")

        # Charging Current (Row 0, Column 0)
        charging_current_frame = ttk.Frame(self.charging_frame)
        charging_current_frame.grid(row=0, column=0, padx=70, pady=2, sticky="nsew", columnspan=1)
        charging_current_label = ttk.Label(charging_current_frame, text="Charging Current", font=("Helvetica", 10, "bold"))
        charging_current_label.pack(pady=(5, 5))
        self.charging_current_meter = ttk.Meter(
            master=charging_current_frame,
            metersize=180,
            amountused=self.device_data['charging_current'],
            meterthickness=10,
            metertype="semi",
            subtext="Charging Current",
            textright="A",
            amounttotal=120,
            bootstyle=self.get_gauge_style(self.device_data['charging_current'], "charging_current"),
            stripethickness=10,
            subtextfont='-size 10'
        )
        self.charging_current_meter.pack()

        # Charging Voltage (Row 0, Column 1)
        charging_voltage_frame = ttk.Frame(self.charging_frame)
        charging_voltage_frame.grid(row=0, column=1, padx=70, pady=2, sticky="nsew", columnspan=1)
        charging_voltage_label = ttk.Label(charging_voltage_frame, text="Charging Voltage", font=("Helvetica", 10, "bold"))
        charging_voltage_label.pack(pady=(5, 5))
        self.charging_voltage_meter = ttk.Meter(
            master=charging_voltage_frame,
            metersize=180,
            amountused=self.device_data['charging_voltage'],
            meterthickness=10,
            metertype="semi",
            subtext="Charging Voltage",
            textright="V",
            amounttotal=35,
            bootstyle=self.get_gauge_style(self.device_data['charging_voltage'], "charging_voltage"),
            stripethickness=10,
            subtextfont='-size 10'
        )
        self.charging_voltage_meter.pack()

        # Battery Health Section
        battery_health_frame = ttk.Labelframe(main_frame, text="Battery Health", bootstyle='dark', borderwidth=10, relief="solid")
        battery_health_frame.grid(row=0, column=2, rowspan=2, padx=5, pady=5, sticky="nsew")

        # Temperature (Top of Battery Health)
        temp_frame = ttk.Frame(battery_health_frame)
        temp_frame.pack(padx=10, pady=5, fill="both", expand=True)
        temp_label = ttk.Label(temp_frame, text="Temperature", font=("Helvetica", 10, "bold"))
        temp_label.pack(pady=(10, 10))
        self.temp_meter = ttk.Meter(
            master=temp_frame,
            metersize=180,
            amountused=self.device_data['temperature'],
            meterthickness=10,
            metertype="semi",
            subtext="Temperature",
            textright="°C",
            amounttotal=100,
            bootstyle=self.get_gauge_style(self.device_data['temperature'], "temperature"),
            stripethickness=8,
            subtextfont='-size 10'
        )
        self.temp_meter.pack()

        # Capacity (Bottom of Battery Health)
        capacity_frame = ttk.Frame(battery_health_frame)
        capacity_frame.pack(padx=10, pady=5, fill="both", expand=True)
        capacity_label = ttk.Label(capacity_frame, text="Capacity", font=("Helvetica", 10, "bold"))
        capacity_label.pack(pady=(10, 10))
        remaining_capacity = self.device_data.get('remaining_capacity')
        full_charge_capacity = self.device_data.get('full_charge_capacity')
        capacity = (remaining_capacity / full_charge_capacity) * 100
        self.capacity_meter = ttk.Meter(
            master=capacity_frame,
            metersize=180,
            amountused=round(capacity, 2),
            meterthickness=10,
            metertype="semi",
            subtext="Capacity",
            textright="%",
            amounttotal=100,
            bootstyle=self.get_gauge_style(capacity, "capacity"),
            stripethickness=8,
            subtextfont='-size 10'
        )
        self.capacity_meter.pack()


        # Discharging Section
        self.discharging_frame = ttk.Labelframe(main_frame, text="Discharging", bootstyle=discharging_bootstyle, borderwidth=10, relief="solid")
        self.discharging_frame .grid(row=1, column=0, columnspan=2, padx=5, pady=2, sticky="nsew")

        def update_max_value():
            if self.check_discharge_max.get():
                self.discharge_max_value.set(450)
            else:
                self.discharge_max_value.set(100)
            self.discharging_current_meter.configure(amounttotal=self.discharge_max_value.get())

        # Discharging Current (Row 1, Column 0)
        discharging_current_frame = ttk.Frame(self.discharging_frame )
        discharging_current_frame.grid(row=1, column=0, padx=70, pady=2, sticky="nsew", columnspan=1)
        discharging_current_label = ttk.Label(discharging_current_frame, text="Discharging Current", font=("Helvetica", 10, "bold"))
        discharging_current_label.pack(pady=(5, 5))
        # Variable to control the max value of the discharging current gauge
        self.discharge_max_value = tk.IntVar(value=100)  # Default max value

        # Checkbox to toggle the max value of the discharging current gauge
        self.discharge_max_checkbox = ttk.Checkbutton(
            discharging_current_frame, 
            text="Set Max", 
            variable=self.check_discharge_max, 
            command=update_max_value
        )
        selected_mode = self.mode_var.get()
        if selected_mode == "Maintenance":
            self.discharge_max_checkbox.pack(pady=(5, 10))

        # Discharging Current Gauge
        self.discharging_current_meter = ttk.Meter(
            master=discharging_current_frame,
            metersize=180,
            amountused=self.device_data['current'],  # Assuming this is the discharging current
            meterthickness=10,
            metertype="semi",
            subtext="Discharging Current",
            textright="A",
            amounttotal=self.discharge_max_value.get(),
            bootstyle=self.get_gauge_style(self.device_data['current'], "current"),
            stripethickness=10,
            subtextfont='-size 10'
        )
        self.discharging_current_meter.pack()

        # Discharging Voltage (Row 1, Column 1)
        discharging_voltage_frame = ttk.Frame(self.discharging_frame )
        discharging_voltage_frame.grid(row=1, column=1, padx=70, pady=2, sticky="nsew", columnspan=1)
        discharging_voltage_label = ttk.Label(discharging_voltage_frame, text="Battery Voltage", font=("Helvetica", 10, "bold"))
        discharging_voltage_label.pack(pady=(5, 35))
        self.discharging_voltage_meter = ttk.Meter(
            master=discharging_voltage_frame,
            metersize=180,
            amountused=self.device_data['voltage'],  # Assuming this is the discharging voltage
            meterthickness=10,
            metertype="semi",
            subtext="Battery Voltage",
            textright="V",
            amounttotal=100,
            bootstyle=self.get_gauge_style(self.device_data['voltage'], "voltage"),
            stripethickness=10,
            subtextfont='-size 10'
        )
        self.discharging_voltage_meter.pack()

        # Configure row and column weights to ensure even distribution
        for i in range(2):
            main_frame.grid_rowconfigure(i, weight=1)
        for j in range(3):
            main_frame.grid_columnconfigure(j, weight=1)

        # Pack the content_frame itself
        self.content_frame.pack(fill="both", expand=True)
        update_max_value()
        self.select_button(self.dashboard_button)

    def prompt_manual_cycle_count(self):
        # Create a new top-level window for input
            input_window = tk.Toplevel(self.master)
            input_window.title("Cycle Count")
        
            # Calculate the position for the window to be centered
            screen_width = self.master.winfo_screenwidth()
            screen_height = self.master.winfo_screenheight()
            window_width = 300
            window_height = 150
            x = (screen_width // 2) - (window_width // 2)
            y = (screen_height // 2) - (window_height // 2)
        
            # Set the geometry of the window directly in the center
            input_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
            input_window.transient(self.master)  # Set to always be on top of the parent window
            input_window.grab_set()  # Make the input window modal
        
            # Create and pack the widgets
            label = ttk.Label(input_window, text="Enter Cycle Count:", font=("Helvetica", 12))
            label.pack(pady=10)
        
            cycle_count_var = tk.IntVar()  # Use IntVar to store integer value
            entry = ttk.Entry(input_window, textvariable=cycle_count_var)
            entry.pack(pady=5)
        
            # Function to handle OK button click
            def handle_ok():
                # Get the selected battery from the dropdown
                if self.selected_battery == "Battery 1":
                    # Update cycle count for Battery 1
                    device_data_battery_1['cycle_count'] = cycle_count_var.get()
                    update_cycle_count_in_can_data(device_data_battery_1['serial_number'], device_data_battery_1['cycle_count'])
                elif self.selected_battery == "Battery 2":
                    # Update cycle count for Battery 2
                    device_data_battery_2['cycle_count'] = cycle_count_var.get()
                    update_cycle_count_in_can_data(device_data_battery_2['serial_number'], device_data_battery_2['cycle_count'])

                input_window.destroy()  # Close the input window
                self.show_dashboard()
        
            # OK Button
            ok_button = ttk.Button(input_window, text="OK", command=handle_ok)
            ok_button.pack(pady=10)
        
            # Focus the input window and entry
            input_window.focus()
            entry.focus_set()
        
            # Bind the Return key to trigger the OK button
            input_window.bind('<Return>', lambda event: handle_ok())

    def show_help(self):
        self.clear_content_frame()
        label = ttk.Label(self.content_frame, text="Help", font=("Helvetica", 16))
        label.pack(pady=20)
        self.select_button(self.help_button)
    
    def show_report(self):
        self.clear_content_frame()
        latest_data = {}
        selected_mode = self.mode_var.get()  # Get the selected mode
        latest_data = get_latest_can_data(self.device_data['serial_number'])
        # Container for the content
        main_frame = ttk.Frame(self.content_frame)
        main_frame.grid(row=0, column=0, columnspan=12, rowspan=6, padx=10, pady=10, sticky="nsew")

        # Configure the content frame for expansion and add equal space on both sides
        self.content_frame.grid_columnconfigure(0, weight=1)  # Allow the content frame to expand
        self.content_frame.grid_rowconfigure(0, weight=1)     # Allow the main_frame to expand vertically

        # Configure the main_frame for equal spacing
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=2)
        main_frame.grid_columnconfigure(2, weight=1)

        def open_report_folder():
            documents_folder = os.path.join(os.path.expanduser("~"), "Documents", "Battery Reports")
            os.startfile(documents_folder)

        # Row 0: View Reports Button
        generate_button = ttk.Button(main_frame, text="Old Reports", command=lambda:open_report_folder())
        generate_button.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="ew")

        # Battery Info Frame (row 1-2, col 0-2 inside the main_frame)
        battery_info_frame = ttk.LabelFrame(main_frame, text="Battery Info", borderwidth=5)
        battery_info_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky="nsew")

        # Add a Label for the Project Dropdown
        ttk.Label(battery_info_frame, text="Project").grid(row=0, column=0, padx=10, pady=10, sticky="e")

        # Add the Combobox for the Project Dropdown
        project_options = config.config_values['can_config']['projects']
        project_combobox = ttk.Combobox(battery_info_frame, values=project_options, state="readonly")  # Dropdown with readonly state
        project_combobox.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        project_combobox.set("Select a Project")

        # Add individual fields for Battery Info
        ttk.Label(battery_info_frame, text="Device Name").grid(row=0,  column=2, padx=10, pady=10, sticky="e")
        device_name_entry = ttk.Entry(battery_info_frame)
        device_name_entry.insert(0, str(latest_data.get("Device Name", "")))
        device_name_entry.grid(row=0, column=3, padx=10, pady=10, sticky="w")

        ttk.Label(battery_info_frame, text="Serial Number").grid(row=0,  column=4, padx=10, pady=10, sticky="e")
        serial_number_entry = ttk.Entry(battery_info_frame)
        serial_number_entry.insert(0, str(latest_data.get("Serial Number",0)))
        serial_number_entry.grid(row=0, column=5, padx=10, pady=10, sticky="w")
        serial_number_entry.config(state="readonly")  # Disable editing

        ttk.Label(battery_info_frame, text="Manufacturer Name").grid(row=1,  column=0, padx=10, pady=10, sticky="e")
        manufacturer_entry = ttk.Entry(battery_info_frame)
        manufacturer_entry.insert(0, str(latest_data.get("Manufacturer Name", "")))
        manufacturer_entry.grid(row=1, column=1, padx=10, pady=10, sticky="w")

        ttk.Label(battery_info_frame, text="Cycle Count").grid(row=1,  column=2, padx=10, pady=10, sticky="e")
        cycle_count_entry = ttk.Entry(battery_info_frame)
        cycle_count_entry.insert(0, str(latest_data.get("Cycle Count",0)))
        cycle_count_entry.grid(row=1, column=3, padx=10, pady=10, sticky="w")

        ttk.Label(battery_info_frame, text="Battery Capacity").grid(row=1,  column=4, padx=10, pady=10, sticky="e")
        capacity_entry = ttk.Entry(battery_info_frame)
        capacity_entry.insert(0, str(latest_data.get("Full Charge Capacity",103)))
        capacity_entry.grid(row=1, column=5, padx=10, pady=10, sticky="w")
        capacity_entry.config(state="readonly")  # Disable editing

        # Row 3-4: Charging Info Frame
        charging_frame = ttk.LabelFrame(main_frame, text="Charging Info", borderwidth=5)
        charging_frame.grid(row=2, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

        ttk.Label(charging_frame, text="Date :").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        charging_date_entry = ttk.Entry(charging_frame)
        charging_date_entry.insert(0, "09/09/2024")
        charging_date_entry.grid(row=0, column=0, padx=50, pady=5, sticky="w")

        ttk.Label(charging_frame, text="OCV :").grid(row=0, column=0, padx=200, pady=5, sticky="w")
        charging_ocv_entry = ttk.Entry(charging_frame)
        charging_ocv_entry.insert(0, 0)
        charging_ocv_entry.grid(row=0, column=0, padx=250, pady=5, sticky="w")

        
        def load_can_charging_data():
            # Define the path to the CAN charging data Excel file inside the AppData directory
            can_charging_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "can_charging_data.xlsx")

            # Read the Excel file using pandas
            try:
                df_can_charging = pd.read_excel(can_charging_data_file)
            except FileNotFoundError:
                print(f"Error: {can_charging_data_file} not found.")
                return

            # Retrieve the serial number from self.device_data (assuming it contains a key 'serial_number')
            device_serial_number = self.device_data.get('serial_number')

            # Filter the DataFrame where the serial number matches
            filtered_data = df_can_charging[df_can_charging['Serial Number'] == device_serial_number]

            # If no matching serial number is found, you can add a fallback or a message
            if filtered_data.empty:
                print(f"No data found for Serial Number: {device_serial_number}")
                return

            # Clear any existing data in the Treeview
            for item in charging_table.get_children():
                charging_table.delete(item)

            # Insert the filtered data into the Treeview
            for _, row in filtered_data.iterrows():
                charging_table.insert('', 'end', values=(row['Hour'], row['Voltage'], row['Current'], row['Temperature'], row['State of Charge']))



        # Create Treeview for Charging Info Table
        columns = ('Hour', 'Voltage (V)', 'Current (A)', 'Temperature (°C)', 'State of Charge (%)')
        charging_table = ttk.Treeview(charging_frame, columns=columns, show='headings', height=5)

        charging_table.heading('Hour', text='Mins')
        charging_table.heading('Voltage (V)', text='Voltage (V)')
        charging_table.heading('Current (A)', text='Current (A)')
        charging_table.heading('Temperature (°C)', text='Temperature (°C)')
        charging_table.heading('State of Charge (%)', text='State of Charge (%)')

        charging_table.column('Hour', width=100, anchor='center')
        charging_table.column('Voltage (V)', width=100, anchor='center')
        charging_table.column('Current (A)', width=100, anchor='center')
        charging_table.column('Temperature (°C)', width=150, anchor='center')
        charging_table.column('State of Charge (%)', width=150, anchor='center')

        # Adding a vertical scrollbar to the charging table
        scrollbar_charging = ttk.Scrollbar(charging_frame, orient="vertical", command=charging_table.yview)
        charging_table.configure(yscrollcommand=scrollbar_charging.set)

        charging_table.grid(row=1, column=0, sticky='nsew')
        scrollbar_charging.grid(row=1, column=1, sticky='ns')

        charging_frame.grid_rowconfigure(1, weight=1)
        charging_frame.grid_columnconfigure(0, weight=1)

        # Call the function to load CAN charging data from the Excel file
        load_can_charging_data()

        # Add the Discharge Info Frame based on the mode (Testing vs Maintenance)
        discharge_info_frame = ttk.LabelFrame(main_frame, text="Discharge Info", borderwidth=5)
        discharge_info_frame.grid(row=3, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

        ttk.Label(discharge_info_frame, text="Date :").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        discharging_date_entry = ttk.Entry(discharge_info_frame)
        discharging_date_entry.insert(0, "09/09/2024")
        discharging_date_entry.grid(row=0, column=0, padx=50, pady=5, sticky="w")

        ttk.Label(discharge_info_frame, text="OCV :").grid(row=0, column=0, padx=200, pady=5, sticky="w")
        discharging_ocv_entry = ttk.Entry(discharge_info_frame)
        discharging_ocv_entry.insert(0, 0)
        discharging_ocv_entry.grid(row=0, column=0, padx=250, pady=5, sticky="w")

        def save_values():
            # Collect values from the entry boxes into arrays
            project = project_combobox.get()  # Get the selected project
            device_name = device_name_entry.get()
            serial_number = serial_number_entry.get()
            manufacturer_name = manufacturer_entry.get()
            cycle_count = cycle_count_entry.get()
            full_charge_capacity = capacity_entry.get()
            charging_date = charging_date_entry.get()  # New field
            ocv_before_charging = charging_ocv_entry.get()  # New field
            discharging_date = discharging_date_entry.get()  # New field
            ocv_before_discharging = discharging_ocv_entry.get()  # New field

            # Check if any required field is empty
            if project == "Select a Project" or not project or not device_name or not serial_number or not manufacturer_name or not cycle_count or not full_charge_capacity:
                messagebox.showwarning("Missing Information", "Please fill in all the required fields and select a valid project.")
                return

            # Create an array to hold the values
            data_to_save = [
                project,  # Add the project value at the beginning
                device_name,
                serial_number,
                manufacturer_name,
                cycle_count,
                full_charge_capacity,
                charging_date,
                ocv_before_charging,
                discharging_date,
                ocv_before_discharging
            ]

            # Call the method to update the Excel file
            update_excel_and_download_pdf(data_to_save)


        def load_can_discharging_data():
            """
            Load the CAN discharging data from an Excel file and display it in the Treeview.
            """
            # Define the path to the CAN discharging data Excel file inside the AppData directory
            can_discharging_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "can_discharging_data.xlsx")

            # Read the Excel file using pandas
            try:
                df_can_discharging = pd.read_excel(can_discharging_data_file)
            except FileNotFoundError:
                messagebox.showerror("Error", f"{can_discharging_data_file} not found.")
                return

            # Retrieve the serial number from self.device_data (assuming it contains a key 'serial_number')
            device_serial_number = self.device_data.get('serial_number')

            # Filter the DataFrame where the serial number matches
            filtered_data = df_can_discharging[df_can_discharging['Serial Number'] == device_serial_number]

            # If no matching serial number is found, display a message and return
            if filtered_data.empty:
                messagebox.showwarning("No Data", f"No data found for Serial Number: {device_serial_number}")
                return

            # Clear any existing data in the Treeview
            for item in discharging_table.get_children():
                discharging_table.delete(item)

            # Insert the filtered data into the Treeview
            for _, row in filtered_data.iterrows():
                try:
                    discharging_table.insert('', 'end', values=(
                        row['Hour'], 
                        float(row['Voltage']), 
                        float(row['Current']), 
                        float(row['Temperature']), 
                        float(row['Load Current'])
                    ))
                except ValueError:
                    print(f"Error processing row: {row}")

            print(f"Data for Serial Number {device_serial_number} loaded successfully.")

        # Create Treeview for Charging Info Table
        columns = ('Hour', 'Voltage (V)', 'Current (A)', 'Temperature (°C)', 'Load Current (A)')
        discharging_table = ttk.Treeview(discharge_info_frame, columns=columns, show='headings', height=5)

        discharging_table.heading('Hour', text='Mins')
        discharging_table.heading('Voltage (V)', text='Voltage (V)')
        discharging_table.heading('Current (A)', text='Current (A)')
        discharging_table.heading('Temperature (°C)', text='Temperature (°C)')
        discharging_table.heading('Load Current (A)', text='Load Current (A)')

        discharging_table.column('Hour', width=100, anchor='center')
        discharging_table.column('Voltage (V)', width=100, anchor='center')
        discharging_table.column('Current (A)', width=100, anchor='center')
        discharging_table.column('Temperature (°C)', width=150, anchor='center')
        discharging_table.column('Load Current (A)', width=150, anchor='center')

        # Adding a vertical scrollbar to the discharging table
        scrollbar_discharging = ttk.Scrollbar(discharge_info_frame, orient="vertical", command=discharging_table.yview)
        discharging_table.configure(yscrollcommand=scrollbar_discharging.set)
        discharging_table.grid(row=1, column=0, sticky='nsew')
        scrollbar_discharging.grid(row=1, column=1, sticky='ns')

        discharge_info_frame.grid_rowconfigure(1, weight=1)
        discharge_info_frame.grid_columnconfigure(0, weight=1)

        def load_can_data():
            """
            Load the CAN charging and discharging data from an Excel file and display it.
            """
            # Define the path to the CAN data Excel file inside the AppData directory
            can_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "can_data.xlsx")

            # Read the Excel file using pandas
            try:
                df_can_data = pd.read_excel(can_data_file)
            except FileNotFoundError:
                messagebox.showerror("Error", f"{can_data_file} not found.")
                return

            # Retrieve the serial number from self.device_data
            device_serial_number = self.device_data.get('serial_number')

            # Filter the DataFrame where the serial number matches
            filtered_data = df_can_data[df_can_data['Serial Number'] == device_serial_number]

            # If no matching serial number is found, display a message and return
            if filtered_data.empty:
                messagebox.showwarning("No Data", f"No data found for Serial Number: {device_serial_number}")
                return

            # Assuming there's only one row per serial number, use the first row for display
            first_row = filtered_data.iloc[0]

            # Populate individual fields from the data
            device_name_entry.delete(0, tk.END)
            device_name_entry.insert(0, str(first_row.get("Device Name", "")))

            serial_number_entry.delete(0, tk.END)
            serial_number_entry.insert(0, str(first_row.get("Serial Number", "")))

            manufacturer_entry.delete(0, tk.END)
            manufacturer_entry.insert(0, str(first_row.get("Manufacturer Name", "")))

            cycle_count_entry.delete(0, tk.END)
            cycle_count_entry.insert(0, str(first_row.get("Cycle Count", "")))

            capacity_entry.config(state="normal")  # Enable field for update
            capacity_entry.delete(0, tk.END)
            capacity_entry.insert(0, str(first_row.get("Full Charge Capacity", 0)))
            capacity_entry.config(state="readonly")  # Disable field after update

            # Charging Info fields
            charging_ocv_entry.delete(0, tk.END)
            charging_ocv_entry.insert(0, float(first_row.get("OCV Before Charging", 0)))

            charging_date_entry.delete(0, tk.END)
            charging_date_entry.insert(0, str(first_row.get("Charging Date", '')))
            
            # Discharging Info fields
            discharging_ocv_entry.delete(0, tk.END)
            discharging_ocv_entry.insert(0, float(first_row.get("OCV Before Discharging", 0)))

            discharging_date_entry.delete(0, tk.END)
            discharging_date_entry.insert(0, str(first_row.get("Discharging Date", '')))

        load_can_discharging_data()
        load_can_data()

        main_frame.grid_rowconfigure(3, weight=1)

        # Row 5: Download Report Button at the bottom
        download_button = ttk.Button(main_frame, text="View Report", command=lambda:save_values())
        download_button.grid(row=4, column=0, columnspan=3, padx=10, pady=10, sticky="ew")

        # Ensure the last row is expanded to push the button to the bottom
        main_frame.grid_rowconfigure(4, weight=1)

        # Make the content_frame expand
        self.content_frame.pack(fill="both", expand=True)

        # Select the report button (if necessary for your UI)
        self.select_button(self.report_button)

    def toggle_dynamic_fields(self):
        if self.show_dynamic_fields_var.get():
            self.dynamic_fields_frame.grid()  # Show the dynamic fields
            self.maintenance_button.grid_remove()  # Hide the "Set 100A and Turn ON Load" button
            self.turn_off_button.grid_remove()  # Hide the "Turn OFF Load" button
        else:
            self.dynamic_fields_frame.grid_remove()  # Hide the dynamic fields
            self.maintenance_button.grid()  # Show the "Set 100A and Turn ON Load" button
            self.turn_off_button.grid()  # Show the "Turn OFF Load" button
    
    def save_dynamic_fields(self):
        """Save the dynamic field values entered in the L1, L2, T1, T2, and Repeat fields."""
        # Fetch the values from the entry fields
        l1_value = self.l1_entry.get()
        l2_value = self.l2_entry.get()
        t1_value = self.t1_entry.get()
        t2_value = self.t2_entry.get()
        repeat_value = self.repeat_entry.get()

        # Process the values (for now, we will print them as a placeholder)
        print(f"L1: {l1_value}, L2: {l2_value}, T1: {t1_value}, T2: {t2_value}, Repeat: {repeat_value}")

    def update_mode(self, event=None):
        """Update the UI components based on the selected mode (Testing or Maintenance)."""
        selected_mode = self.mode_var.get()

        if selected_mode == "Testing" and self.selected_button == self.dashboard_button:
            # Hide the checkbox in Testing
            if hasattr(self, 'discharge_max_checkbox'):
                self.discharge_max_checkbox.pack_forget()
        elif selected_mode == "Maintenance" and self.selected_button == self.dashboard_button:
            if hasattr(self, 'discharge_max_checkbox'):
                self.discharge_max_checkbox.pack(pady=(5, 10), before=self.discharging_current_meter)

        # Check if we are in Testing or Maintenance
        if selected_mode == "Maintenance":
            self.limited = False  # Show full data in Maintenance Mode
            # Show the Configuration button
            if not hasattr(self, 'config_button'):  # Ensure the button exists
                self.config_button = ttk.Button(
                    self.side_menu_frame,
                    text=" Configuration",
                    command=lambda: self.select_button(self.config_button,self.show_config),
                    image=self.config_icon,
                    compound="left",  # Place the icon to the left of the text
                    bootstyle="info"
                )
            self.config_button.pack(fill="x", pady=2, before=self.help_button)
        else:
            self.limited = True  # Show limited data in Testing
            # Hide the Configuration button if in Testing
            if hasattr(self, 'config_button'):
                self.config_button.pack_forget()
        """Update the Info page content based on the selected mode."""
        if self.selected_button == self.info_button:
            self.show_info()
        if self.selected_button == self.control_button:
            self.show_new_control()
        if self.selected_button == self.config_button:
            self.show_dashboard()
    
    def configure_styles(self):
            # Configure custom styles
            self.style.configure("Custom.Treeview", font=("Helvetica", 10), rowheight=25, background="#f0f0f0", fieldbackground="#e8f4ec")
            self.style.configure("Custom.Treeview.Heading", font=("Helvetica", 12, "bold"), background="#333", foreground="#e8f4ec", padding=5)
            self.style.map("Custom.Treeview", background=[("selected", "#0078d7")], foreground=[("selected", "white")])

    def center_window(self, width, height):
            """Centers the window on the screen."""
            screen_width = self.master.winfo_screenwidth()
            screen_height = self.master.winfo_screenheight()
            center_x = int(screen_width / 2 - width / 2)
            center_y = int(screen_height / 2 - height / 2)
            self.master.geometry(f'{width}x{height}+{center_x}+{center_y}')
            self.master.update_idletasks()

    def load_icon(self, path, size=(15, 15)):
        image = Image.open(path)
        image = image.resize(size, Image.Resampling.LANCZOS)
        ctk_image = ctk.CTkImage(light_image=image, dark_image=image, size=size)
        return ctk_image
    
    def load_icon_menu(self, path, size=(24, 24)):
        icon = Image.open(path)
        icon = icon.resize(size, Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(icon)

    def update_frame_sizes(self):
        width = self.master.winfo_width()
        self.side_menu_width = int(width * self.side_menu_width_ratio)
        self.content_frame_width = width - self.side_menu_width

    def on_window_resize(self, event):
        self.update_frame_sizes()
        self.side_menu_frame.place(x=0, y=0, width=self.side_menu_width, relheight=1.0)
        self.content_frame.place(x=self.side_menu_width, y=0, width=self.content_frame_width, relheight=1.0)

    def select_button(self, button, command=None):
        # Check if the previously selected button still exists
        if self.selected_button and self.selected_button.winfo_exists():
            # Only reset style if the previously selected button is not the Disconnect button
            if self.selected_button != self.disconnect_button:
                self.selected_button.configure(bootstyle="info")  # Reset the previously selected button's style

        # Highlight the new selected button
        if button != self.disconnect_button:
            button.configure(bootstyle="primary")  # Apply the selected style, unless it's the Disconnect button
        self.selected_button = button

        # Execute the button's command if provided
        if command:
            command()

    def clear_content_frame(self):
        # Clear the content frame before displaying new content
        for widget in self.content_frame.winfo_children():
            widget.destroy()

    def on_disconnect(self):
        self.mode_var.set("Testing")
        if device_data_battery_2['serial_number'] != 0:
            pcan_write_control('both_off',1)
            pcan_write_control('both_off',2)
        else:
            pcan_write_control('both_off',1)
        time.sleep(1)
        # Cancel the scheduled auto-refresh task
        if hasattr(self, 'auto_refresh_task'):
            self.master.after_cancel(self.auto_refresh_task)
            print("Auto-refresh stopped.")
        pcan_uninitialize()
        self.first_time_dashboard = True
        self.caution_alert_shown = False
        self.alarm_alert_shown = False
        self.critical_alert_shown = False
        self.main_frame.pack_forget()
        self.main_window.show_main_window()
        # Perform disconnection logic here (example: print disconnect message)
        print("Disconnecting...")

    def show_info(self, event=None):
        self.clear_content_frame()

        # Determine if we are in Testing or Maintenance
        selected_mode = self.mode_var.get()
        self.limited = selected_mode == "Testing"   
        # Add a new right-side frame for controls     
        self.refresh_icon = self.load_icon(os.path.join(self.assets_path, "refresh.png"))
        self.start_icon = self.load_icon(os.path.join(self.assets_path, "start.png"))
        self.stop_icon = self.load_icon(os.path.join(self.assets_path, "stop.png"))
        self.file_icon = self.load_icon(os.path.join(self.assets_path, "folder.png"))

        # Create a Frame to hold the checkbox and buttons in one row with a white background
        control_frame = ttk.Labelframe(self.content_frame, text="Battery Data Controls", bootstyle="dark", borderwidth=5, relief="solid")
        control_frame.pack(fill="x", expand=True, padx=5, pady=1)


        self.refresh_button = ctk.CTkButton(control_frame, text="Refresh", image=self.refresh_icon, compound="left", command=self.refresh_info, width=6, fg_color="#5188d4",
            hover_color="#4263cc")
        self.refresh_button.pack(side="left", padx=5, pady=1, fill="x", expand=True)

        timer_value_label = ttk.Label(control_frame, text="Timer")
        timer_value_label.pack(side="left", padx=5, pady=1, fill="x", expand=True)

        self.timer_value = tk.StringVar(value=5)
        self.timer = ttk.Spinbox(
            control_frame,
            from_=1,
            to=30,
            width=5,
            values=(1, 5, 10, 15, 20, 30),
            textvariable=self.timer_value,
        )
        self.timer.pack(side="left", padx=5, pady=1, fill="x", expand=True)

        self.start_logging_button = ctk.CTkButton(control_frame, text="Start Log", image=self.start_icon, compound="left", command=self.start_logging, fg_color="#72b043",
            hover_color="#007f4e")
        self.start_logging_button.pack(side="left", padx=5, pady=1, fill="x", expand=True)

        self.stop_logging_button = ctk.CTkButton(control_frame, text="Stop Log", image=self.stop_icon, compound="left", command=self.stop_logging, fg_color="#ff6361",
            hover_color="#d74a49", state=tk.DISABLED)
        self.stop_logging_button.pack(side="left", padx=5, pady=1, fill="x", expand=True)

        file_button = ctk.CTkButton(control_frame, image=self.file_icon, text='', compound="left", command=self.folder_open, fg_color="#72b043",
            hover_color="#007f4e")
        file_button.pack(side="left", padx=5, pady=1, fill="x", expand=True)

        if self.logging_active:
            self.start_logging_button.configure(state=tk.DISABLED)
            self.start_logging_button.configure(state=tk.NORMAL)
        else:
            self.start_logging_button.configure(state=tk.NORMAL)
            self.stop_logging_button.configure(state=tk.DISABLED)   
        # Frame to hold the Treeview and Scrollbar
        table_frame = ttk.Frame(self.content_frame, borderwidth=5, relief="solid")
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)   
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical")
        scrollbar.pack(side="right", fill="y")   
        columns = ('Name', 'Value', 'Units')
        self.info_table = ttk.Treeview(table_frame, columns=columns, show='headings', style="Custom.Treeview", yscrollcommand=scrollbar.set)   
        scrollbar.config(command=self.info_table.yview)   
        self.info_table.column('Name', anchor='center', width=200)
        self.info_table.column('Value', anchor='center', width=200)
        self.info_table.column('Units', anchor='center', width=100)   
        self.info_table.heading('Name', text='Name', anchor='center')
        self.info_table.heading('Value', text='Value', anchor='center')
        self.info_table.heading('Units', text='Units', anchor='center')   
        self.info_table.pack(fill="both", expand=True)   
        if self.limited:
            # Insert limited data for Testing Mode
            limited_data_keys = [
                 'device_name', 
                 'serial_number', 
                 'manufacturer_name', 
                 'cycle_count',
                 'remaining_capacity', 
                 'temperature', 
                 'current', 
                 'voltage', 
                 'charging_current', 
                 'charging_voltage', 
                 'charging_battery_status', 
                 'rel_state_of_charge'
             ]
            for key in limited_data_keys:
                name = ' '.join(word.title() for word in key.split('_'))
                value = self.device_data.get(key, 'N/A')
                unit = unit_mapping.get(key, '')
                self.info_table.insert('', 'end', values=(name, value, unit))
        else:
            # Insert full data for Maintenance Mode
            for key, value in self.device_data.items():
                name = ' '.join(word.title() for word in key.split('_'))
                unit = unit_mapping.get(key, '')
                self.info_table.insert('', 'end', values=(name, value, unit))               
        self.status_frame = ttk.Labelframe(self.content_frame, text="Battery Status", bootstyle="dark", borderwidth=5, relief="solid")
        self.status_frame.pack(fill="both", expand=True, padx=10, pady=10)   
        if self.limited:
            limited_status_flags = {
                 "Over Temperature Alarm": self.battery_status_flags.get('over_temperature_alarm'),
                 "Fully Charged": self.battery_status_flags.get('fully_charged'),
                 "Fully Discharged": self.battery_status_flags.get('fully_discharged')
            }
            self.display_status_labels(self.status_frame, limited_status_flags)
        else:
            self.display_status_labels(self.status_frame, self.battery_status_flags)   
        self.select_button(self.info_button)

    def display_status_labels(self, frame, battery_status_flags):
        max_columns = 3  # Number of columns before wrapping to the next row
        current_column = 0
        current_row = 0

        for label_text, flag_value in battery_status_flags.items():
            transformed_label_text = ' '.join(word.title() for word in label_text.split('_'))
            if transformed_label_text == "Error Codes":
                self.create_error_code_label(frame, flag_value, row=current_row, column=current_column)
            else:
                created = self.create_status_label(frame, transformed_label_text, flag_value, row=current_row, column=current_column)
                if created:
                    current_column += 1
                    if current_column >= max_columns:
                        current_column = 0
                        current_row += 1

        # Ensure that the rows and columns expand evenly
        for i in range(current_row + 1):
            frame.grid_rowconfigure(i, weight=1)
        for j in range(max_columns):
            frame.grid_columnconfigure(j, weight=1)

    def create_status_label(self, frame, label_text, flag_value, row, column):
        # Only show "Fully Charged" and "Fully Discharged" if the value is 1
        if label_text == "Fully Charged" and flag_value == 1:
            status_text = "Fully Charged"
            bootstyle = "inverse-success"
        elif label_text == "Fully Discharged" and flag_value == 1:
            status_text = "Fully Discharged"
            bootstyle = "inverse-danger"
        else:
            # Skip creating the label if the flag is not active (except for Fully Charged/Discharged)
            if flag_value != 1:
                return False
            status_text = ""  # Just show the label without Active/Inactive text
            bootstyle = "inverse-success"

        # Create and place the label
        self.battery_status_label = ttk.Label(frame, text=f"{label_text}", bootstyle=bootstyle, width=35)
        self.battery_status_label.grid(row=row, column=column, padx=5, pady=5, sticky="ew")

        # Make the column expand to fill space
        frame.grid_columnconfigure(column, weight=1)

        return True  # Return True if a label was created
    
    def create_error_code_label(self, frame, error_code, row, column):
        # Determine the bootstyle and text based on the error code
        error_texts = {
            0: ("OK", "success"),
            1: ("Busy", "warning"),
            2: ("Reserved Command", "secondary"),
            3: ("Unsupported Command", "danger"),
            4: ("Access Denied", "info"),
            5: ("Overflow/Underflow", "dark"),
            6: ("Bad Size", "primary"),
            7: ("Unknown Error", "danger"),
        }
        error_text, bootstyle = error_texts.get(error_code, ("Unknown", "danger"))
        self.battery_status_label = ttk.Label(frame, text=f"Error Code: {error_text}", bootstyle=f"inverse-{bootstyle}", width=40)
        self.battery_status_label.grid(row=row, column=column, padx=5, pady=2, sticky="w")

    def auto_refresh(self):
        # asyncio.run(update_device_data())
        for key, value in device_data_battery_1.items():
            if key not in ["serial_number", "cycle_count",'full_charge_capacity','charging_current','rel_state_of_charge']:  # Skip serial_number and cycle_count
                if isinstance(value, (int, float)):  # Only update numeric fields
                    device_data_battery_1[key] = round(value + 0.1,2)
        if device_data_battery_2['serial_number'] > 0:
            for key, value in device_data_battery_2.items():
                if key not in ["serial_number", "cycle_count",'full_charge_capacity',"charging_current","current"]:  # Skip serial_number and cycle_count
                    if isinstance(value, (int, float)):  # Only update numeric fields
                        device_data_battery_2[key] = round(value + 2.51,2)
        if self.selected_battery == "Battery 1":
            self.device_data = device_data_battery_1
            self.battery_status_flags = battery_1_status_flags
        elif self.selected_battery == "Battery 2":
            self.device_data = device_data_battery_2
            self.battery_status_flags = battery_2_status_flags

        self.update_ui()
        current_temp = self.device_data['temperature']
        caution_temp = config.config_values['can_config'].get('temperature_caution')
        alarm_temp = config.config_values['can_config'].get('temperature_alarm')
        critical_temp = config.config_values['can_config'].get('temperature_critical')

        # Check for Caution Alert
        if current_temp == caution_temp and not self.caution_alert_shown:
            messagebox.showwarning("Caution", f"Battery Temperature reached {caution_temp}")
            self.caution_alert_shown = True  # Mark the alert as shown
            # if self.charger_control_var_battery_1:
            #     self.charger_control_var_battery_1.set(False)
            #     self.toggle_button_style(tk.BooleanVar(value=False), self.charge_button_battery_1, 'charge', self.selected_battery)
            # elif self.discharger_control_var_battery_1:
            #     self.discharger_control_var_battery_1.set(False)
            #     self.toggle_button_style(tk.BooleanVar(value=False), self.charge_button_battery_1, 'charge', self.selected_battery)
        # Check for Alarm Alert
        elif current_temp == alarm_temp and not self.alarm_alert_shown:
            messagebox.showwarning("Alarm", f"Battery Temperature reached {alarm_temp}")
            self.alarm_alert_shown = True  # Mark the alert as shown

        # Check for Critical Alert
        elif current_temp >= critical_temp and not self.critical_alert_shown:
            messagebox.showwarning("Critical", f"Battery Temperature reached {current_temp}")
            self.critical_alert_shown = True  # Mark the alert as shown
            turn_load_off()

        # Schedule the next refresh and store the task ID in self.auto_refresh_task
        self.auto_refresh_task = self.master.after(1000, self.auto_refresh)

    def update_ui(self):
        if self.selected_button == self.dashboard_button:
            # Update Charging/Discharging Status
            if self.device_data['charging_current'] > 0:
                self.device_data['charging_battery_status'] = 'Charging'
                self.status_var.set("charging")
                self.status_indicator.configure(text="Charging", bootstyle="success")
                self.charging_frame.configure(bootstyle="success")
                self.discharging_frame.configure(bootstyle="dark")
            elif self.device_data['current'] > 0:
                self.device_data['charging_battery_status'] = 'Discharging'
                self.status_var.set("discharging")
                self.status_indicator.configure(text="Discharging", bootstyle="danger")
                self.discharging_frame.configure(bootstyle="danger")
                self.charging_frame.configure(bootstyle="dark")
            elif self.device_data['current'] == 0 and self.device_data['charging_current'] == 0:
                self.device_data['charging_battery_status'] = 'Off'
                self.status_var.set("off")
                self.status_indicator.configure(text="Off", bootstyle="dark")
                self.charging_frame.configure(bootstyle="dark")
                self.discharging_frame.configure(bootstyle="dark")
            self.charging_current_meter.configure(amountused=self.device_data['charging_current'], bootstyle=self.get_gauge_style(self.device_data['charging_current'], "charging_current"))
            self.charging_voltage_meter.configure(amountused=self.device_data['charging_voltage'], bootstyle=self.get_gauge_style(self.device_data['charging_voltage'], "charging_voltage"))
            self.discharging_current_meter.configure(amountused=self.device_data['current'], bootstyle=self.get_gauge_style(self.device_data['current'], "current"))
            self.discharging_voltage_meter.configure(amountused=self.device_data['voltage'], bootstyle=self.get_gauge_style(self.device_data['voltage'], "voltage"))
            self.temp_meter.configure(amountused=self.device_data['temperature'], bootstyle=self.get_gauge_style(self.device_data['temperature'], "temperature"))
            remaining_capacity = self.device_data.get('remaining_capacity')
            full_charge_capacity = self.device_data.get('full_charge_capacity')
            capacity = (remaining_capacity / full_charge_capacity) * 100
            self.capacity_meter.configure(amountused=round(capacity, 2), bootstyle=self.get_gauge_style(capacity, "capacity"))
        elif self.selected_button == self.info_button:
            selected_mode = self.mode_var.get()
            for item in self.info_table.get_children():
                self.info_table.delete(item)
            if selected_mode == 'Testing':
                # Insert limited data for Testing Mode
                limited_data_keys = [
                    'device_name', 
                    'serial_number', 
                    'manufacturer_name', 
                    'cycle_count',
                    'remaining_capacity', 
                    'temperature', 
                    'current', 
                    'voltage', 
                    'charging_current', 
                    'charging_voltage', 
                    'charging_battery_status', 
                    'rel_state_of_charge'
                ]
                for index, key in enumerate(limited_data_keys):
                    name = ' '.join(word.title() for word in key.split('_'))
                    value = self.device_data.get(key, 'N/A')
                    unit = unit_mapping.get(key, '')
                    self.info_table.insert('', 'end', values=(name, value, unit), tags=('evenrow' if index % 2 == 0 else 'oddrow'))
                limited_status_flags = {
                    "Over Temperature Alarm": battery_status_flags.get('over_temperature_alarm'),
                    "Fully Charged": battery_status_flags.get('fully_charged'),
                    "Fully Discharged": battery_status_flags.get('fully_discharged')
                }
                self.display_status_labels(self.status_frame, limited_status_flags)
            elif selected_mode == 'Maintenance':
                # Insert full data for Maintenance Mode
                for key, value in self.device_data.items():
                    name = ' '.join(word.title() for word in key.split('_'))
                    unit = unit_mapping.get(key, '')
                    self.info_table.insert('', 'end', values=(name, value, unit))
                self.display_status_labels(self.status_frame, battery_status_flags)
    
    def update_charging_log_battery_1(self):
        # Define the directory where your database/excel files are stored
        DATABASE_DIR = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        excel_file_path = os.path.join(DATABASE_DIR, "can_charging_data.xlsx")

        # Define column names for the charging data
        columns = ["Date", "Serial Number", "Hour", "Voltage", "Current", "Temperature", "State of Charge"]

        # Current date and time
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")

        # Retrieve the logging interval (in minutes) from the config
        log_interval_minutes = config.config_values['can_config'].get('logging_time')

        # Retrieve data from battery 1
        battery_1_serial = int(device_data_battery_1['serial_number'])
        battery_1_voltage = float(device_data_battery_1['charging_voltage'])
        battery_1_current = float(device_data_battery_1['charging_current'])
        battery_1_temperature = float(device_data_battery_1['temperature'])
        battery_1_state_of_charge = int(device_data_battery_1['rel_state_of_charge'])

        # Load existing data from the Excel file, or create a new DataFrame if the file doesn't exist
        try:
            df = pd.read_excel(excel_file_path)
        except FileNotFoundError:
            df = pd.DataFrame(columns=columns)

         # Filter rows to only include logs for Battery 1 based on its serial number
        battery_1_logs = df[df["Serial Number"] == battery_1_serial]
        # Check if an old log exists for this serial number and its date is not today
        if not battery_1_logs.empty:
            old_logs = battery_1_logs[battery_1_logs["Date"] != current_date]
            if not old_logs.empty:
                # Remove old logs that do not have today's date
                df = df[~((df["Serial Number"] == battery_1_serial) & (df["Date"] != current_date))]
                battery_1_logs = df[df["Serial Number"] == battery_1_serial]
                # messagebox.showinfo("Old Logs Cleared", f"Old logs for Battery 1 (Serial: {battery_1_serial}) have been cleared.")
        
    
        # Calculate the next "Hour" value to log in intervals like 0:00, 5:00, 10:00, etc.
        total_logs = len(battery_1_logs)
        next_hour = (total_logs * log_interval_minutes) % 60
        next_hour_str = f"{next_hour:02}:00"

        data_battery_1 = [
            current_date,
            battery_1_serial,
            next_hour_str,
            battery_1_voltage,
            battery_1_current,
            battery_1_temperature,
            battery_1_state_of_charge
        ]

        df.loc[len(df)] = data_battery_1

        # Save the updated DataFrame back to the Excel file
        df.to_excel(excel_file_path, index=False)
        remaining_capacity = self.device_data.get('remaining_capacity')
        full_charge_capacity = self.device_data.get('full_charge_capacity')
        battery_capacity = round((remaining_capacity / full_charge_capacity) * 100,2)

        # Add new data for battery 1 if charging conditions are met
        if battery_1_current < config.config_values['can_config'].get('charging_cutoff_curr') and battery_1_voltage > config.config_values['can_config'].get('charging_cutoff_volt') and battery_1_state_of_charge > config.config_values['can_config'].get('charging_cutoff_capacity'):
            messagebox.showinfo("Charging Completed", "Battery 1 charging has been completed. Please disconnect charger")
            self.charger_control_var_battery_1.set(False)
            self.toggle_button_style(tk.BooleanVar(value=False), self.charge_button_battery_1, 'charge', self.selected_battery)
            return

        # Schedule the method to run again after the logging interval (in milliseconds)
        self.master.after(log_interval_minutes * 60 * 1000, self.update_charging_log_battery_1)


    def update_charging_log_battery_2(self):
        # Define the directory where your database/excel files are stored
        DATABASE_DIR = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        excel_file_path = os.path.join(DATABASE_DIR, "can_charging_data.xlsx")

        # Define column names for the charging data
        columns = ["Date", "Serial Number", "Hour", "Voltage", "Current", "Temperature", "State of Charge"]

        # Current date and time
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        # Retrieve the logging interval (in minutes) from the config
        log_interval_minutes = config.config_values['can_config'].get('logging_time')

        # Retrieve data from battery 2
        if device_data_battery_2 and int(device_data_battery_2['serial_number']) > 0:
            battery_2_serial = int(device_data_battery_2['serial_number'])
            battery_2_voltage = float(device_data_battery_2['charging_voltage'])
            battery_2_current = float(device_data_battery_2['charging_current'])
            battery_2_temperature = float(device_data_battery_2['temperature'])
            battery_2_state_of_charge = int(device_data_battery_2['rel_state_of_charge'])

            # Filter rows to only include logs for Battery 1 based on its serial number
            battery_2_logs = df[df["Serial Number"] == battery_2_serial]
            
            if not battery_2_logs.empty:
                old_logs = battery_2_logs[battery_2_logs["Date"] != current_date]
                if not old_logs.empty:
                    # Remove old logs that do not have today's date
                    df = df[~((df["Serial Number"] == battery_2_serial) & (df["Date"] != current_date))]
                    battery_2_logs = df[df["Serial Number"] == battery_2_serial]
                    # messagebox.showinfo("Old Logs Cleared", f"Old logs for Battery 1 (Serial: {battery_2_serial}) have been cleared.")

            # Calculate the next "Hour" value to log in intervals like 0:00, 5:00, 10:00, etc.
            total_logs = len(battery_2_logs)
            next_hour = (total_logs * log_interval_minutes) % 60
            next_hour_str = f"{next_hour:02}:00"

            data_battery_2 = [
                current_date,
                battery_2_serial,
                next_hour_str,
                battery_2_voltage,
                battery_2_current,
                battery_2_temperature,
                battery_2_state_of_charge
            ]

            # Load existing data from the Excel file, or create a new DataFrame if the file doesn't exist
            try:
                df = pd.read_excel(excel_file_path)
            except FileNotFoundError:
                df = pd.DataFrame(columns=columns)

            # Add new data for battery 1
            if battery_2_current < config.config_values['can_config'].get('charging_cutoff_curr') and battery_2_voltage > config.config_values['can_config'].get('charging_cutoff_volt') and battery_2_state_of_charge > config.config_values['can_config'].get('charging_cutoff_capacity'):
                messagebox.showinfo("Charging Completed","Battery 2 charging has been completed. Please disconnect charger")
                self.charger_control_var_battery_2.set(False)
                self.toggle_button_style(tk.BooleanVar(value=False), self.charge_button_battery_2, 'charge', self.selected_battery)
                return
            else:
                df.loc[len(df)] = data_battery_2

            # Save the updated DataFrame back to the Excel file
            df.to_excel(excel_file_path, index=False)

        # Schedule the method to run again after the logging interval (in milliseconds)
        self.master.after(log_interval_minutes * 60 * 1000, self.update_charging_log_battery_2)

    def update_discharging_log_battery_1(self):
        # Define the directory where your database/excel files are stored
        DATABASE_DIR = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")

        # Define the path to the Excel file inside the database directory
        excel_file_path = os.path.join(DATABASE_DIR, "can_discharging_data.xlsx")

        # Define column names for the discharging data
        columns = ["Date", "Serial Number", "Hour", "Voltage", "Current", "Temperature", "Load Current"]

        # Current date and time
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        # Retrieve the logging interval (in minutes) from the config
        log_interval_minutes = config.config_values['can_config'].get('logging_time')


        # Retrieve data from battery 1
        battery_1_serial = int(device_data_battery_1['serial_number'])
        battery_1_voltage = float(device_data_battery_1['voltage'])  # Use appropriate key
        battery_1_current = float(device_data_battery_1['current'])  # Use appropriate key
        battery_1_temperature = float(device_data_battery_1['temperature'])
        print(f"load_current 1 {get_load_current}")
        battery_1_load_current = get_load_current()  # Assuming this key exists
        print(f"load_current 2 {battery_1_load_current}")

        try:
            df = pd.read_excel(excel_file_path)
        except FileNotFoundError:
            df = pd.DataFrame(columns=columns)

        # Filter rows to only include logs for Battery 1 based on its serial number
        battery_1_logs = df[df["Serial Number"] == battery_1_serial]
        # Check if an old log exists for this serial number and its date is not today
        if not battery_1_logs.empty:
            old_logs = battery_1_logs[battery_1_logs["Date"] != current_date]
            if not old_logs.empty:
                # Remove old logs that do not have today's date
                df = df[~((df["Serial Number"] == battery_1_serial) & (df["Date"] != current_date))]
                battery_1_logs = df[df["Serial Number"] == battery_1_serial]
                # messagebox.showinfo("Old Logs Cleared", f"Old logs for Battery 1 (Serial: {battery_1_serial}) have been cleared.")

        # Calculate the next "Hour" value to log in intervals like 0:00, 5:00, 10:00, etc.
        total_logs = len(battery_1_logs)
        next_hour = (total_logs * log_interval_minutes) % 60
        next_hour_str = f"{next_hour:02}:00"

        # Check the discharging cutoff conditions
        if battery_1_voltage > config.config_values['can_config'].get('discharge_cutoff_volt'):
            # Prepare the data to save for battery 1
            data_battery_1 = [
                current_date,
                battery_1_serial,
                next_hour_str,
                battery_1_voltage,
                battery_1_current,
                battery_1_temperature,
                battery_1_load_current
            ]

            # Add new data for battery 1
            df.loc[len(df)] = data_battery_1

            # Save the updated DataFrame back to the Excel file
            df.to_excel(excel_file_path, index=False)

        # Check if the discharging process should stop
        if battery_1_current > 0 and battery_1_voltage <= config.config_values['can_config'].get('discharge_cutoff_volt'):
            # Show messagebox and stop logging
            messagebox.showinfo("Discharging Completed", "Battery 1 discharging completed. Stopping log.")
            turn_load_off()
            pcan_write_control('both_off',1)
            self.discharger_control_var_battery_1.set(False)      
            self.toggle_button_style(tk.BooleanVar(value=False), self.discharge_button_battery_1, 'charge', self.selected_battery)
            return  # Stop the recursion here

        # Schedule the method to run again after the logging interval (in milliseconds)
        self.master.after(log_interval_minutes * 60 * 1000, self.update_discharging_log_battery_1)

    # Similarly, for Battery 2
    def update_discharging_log_battery_2(self):
        # Define the directory where your database/excel files are stored
        DATABASE_DIR = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")

        # Define the path to the Excel file inside the database directory
        excel_file_path = os.path.join(DATABASE_DIR, "can_discharging_data.xlsx")

        # Define column names for the discharging data
        columns = ["Date", "Serial Number", "Hour", "Voltage", "Current", "Temperature", "Load Current"]

        # Current date and time
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        # Retrieve the logging interval (in minutes) from the config
        log_interval_minutes = config.config_values['can_config'].get('logging_time')


        # Retrieve data from battery 2
        battery_2_serial = int(device_data_battery_2['serial_number'])
        battery_2_voltage = float(device_data_battery_2['voltage'])  # Use appropriate key
        battery_2_current = float(device_data_battery_2['current'])  # Use appropriate key
        battery_2_temperature = float(device_data_battery_2['temperature'])
        # battery_2_load_current = float(device_data_battery_2['load_current'])  # Assuming this key exists

        # Load existing data from the Excel file, or create a new DataFrame if the file doesn't exist
        try:
            df = pd.read_excel(excel_file_path)
        except FileNotFoundError:
            df = pd.DataFrame(columns=columns)

        # Filter rows to only include logs for Battery 2 based on its serial number
        battery_2_logs = df[df["Serial Number"] == battery_2_serial]
        # Check if an old log exists for this serial number and its date is not today
        if not battery_2_logs.empty:
            old_logs = battery_2_logs[battery_2_logs["Date"] != current_date]
            if not old_logs.empty:
                # Remove old logs that do not have today's date
                df = df[~((df["Serial Number"] == battery_2_serial) & (df["Date"] != current_date))]
                battery_2_logs = df[df["Serial Number"] == battery_2_serial]
                # messagebox.showinfo("Old Logs Cleared", f"Old logs for Battery 1 (Serial: {battery_2_serial}) have been cleared.")

        # Calculate the next "Hour" value to log in intervals like 0:00, 5:00, 20:00, etc.
        total_logs = len(battery_2_logs)
        next_hour = (total_logs * log_interval_minutes) % 60
        next_hour_str = f"{next_hour:02}:00"

        # Check the discharging cutoff conditions
        if battery_2_voltage > config.config_values['can_config'].get('discharge_cutoff_volt'):
            # Prepare the data to save for battery 2
            data_battery_2 = [
                current_date,
                battery_2_serial,
                next_hour_str,
                battery_2_voltage,
                battery_2_current,
                battery_2_temperature,
                load_current
            ]

            # Add new data for battery 2
            df.loc[len(df)] = data_battery_2

            # Save the updated DataFrame back to the Excel file
            df.to_excel(excel_file_path, index=False)

        # Check if the discharging process should stop
        if battery_2_current > 0 and battery_2_voltage <= config.config_values['can_config'].get('discharge_cutoff_volt'):
            # Show messagebox and stop logging
            messagebox.showinfo("Discharging Completed", "Battery 2 discharging completed. Stopping log.")
            self.discharger_control_var_battery_2.set(False)
            turn_load_off()
            self.toggle_button_style(tk.BooleanVar(value=False), self.discharge_button_battery_2, 'charge', self.selected_battery)
            return  # Stop the recursion here

        # Schedule the method to run again after the logging interval (in milliseconds)
        self.master.after(log_interval_minutes * 60 * 1000, self.update_discharging_log_battery_2)

    def refresh_info(self):
        asyncio.run(update_device_data())
        # Clear the existing table rows
        for item in self.info_table.get_children():
            self.info_table.delete(item)

        # Check if Testing Mode is active
        if self.limited:
            # Insert limited data for Testing Mode
            limited_data_keys = [
                'device_name', 
                'serial_number', 
                'manufacturer_name', 
                'cycle_count',
                'remaining_capacity', 
                'temperature', 
                'current', 
                'voltage', 
                'charging_current', 
                'charging_voltage', 
                'charging_battery_status', 
                'rel_state_of_charge'
            ]
            for index, key in enumerate(limited_data_keys):
                name = ' '.join(word.title() for word in key.split('_'))
                value = self.device_data.get(key, 'N/A')
                unit = unit_mapping.get(key, '')
                self.info_table.insert('', 'end', values=(name, value, unit), tags=('evenrow' if index % 2 == 0 else 'oddrow'))
        else:
            # Insert full data for Maintenance Mode
            for index, (key, value) in enumerate(self.device_data.items()):
                name = ' '.join(word.title() for word in key.split('z_'))
                unit = unit_mapping.get(key, '')
                self.info_table.insert('', 'end', values=(name, value, unit), tags=('evenrow' if index % 2 == 0 else 'oddrow'))
        # Update the battery status flags
        if self.limited:
            limited_status_flags = {
                "Over Temperature Alarm": battery_status_flags.get('over_temperature_alarm'),
                "Fully Charged": battery_status_flags.get('fully_charged'),
                "Fully Discharged": battery_status_flags.get('fully_discharged')
            }
            self.display_status_labels(self.status_frame, limited_status_flags)
        else:
            self.display_status_labels(self.status_frame, battery_status_flags)
    
    def start_logging(self):
        self.logging_active = True
        # Disable the Start Logging button and enable the Stop Logging button
        self.start_logging_button.configure(state=tk.DISABLED)
        self.stop_logging_button.configure(state=tk.NORMAL)

        # Create the folder path
        folder_path = os.path.join(os.path.expanduser("~"), "Documents", "ADE Log Folder")
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Create a timestamped file name
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"CAN_Connector_Log_{self.device_data['charging_battery_status']}_{timestamp}.xlsx"
        self.file_path = os.path.join(folder_path, file_name)

        # Create a new Excel workbook and worksheet
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "BMS Data"

        # Write the headers (first two columns for date and time, then the device names)
        headers = ["Date", "Time"] + [name_mapping.get(key, key) for key in self.device_data.keys()]
        self.sheet.append(headers)

        # Start the logging loop
        self.logging_active = True
        self.log_data()

    def log_data(self):
        if self.logging_active:
            # Update device data
            asyncio.run(update_device_data())

            for item in self.info_table.get_children():
                self.info_table.delete(item)

            # Repopulate the table with updated data
            for index, (key, value) in enumerate(self.device_data.items()):
                name = ' '.join(word.title() for word in key.split('_'))
                unit = unit_mapping.get(key, key)
                self.info_table.insert('', 'end', values=(name, value, unit), tags=('evenrow' if index % 2 == 0 else 'oddrow'))

            # Get the current date and time
            current_datetime = datetime.datetime.now()
            date = current_datetime.strftime("%Y-%m-%d")
            time = current_datetime.strftime("%H:%M:%S")

            # Collect the row data: date, time, and device values
            row_data = [date, time] + [self.device_data[key] for key in self.device_data.keys()]

            # Write the row data to the Excel sheet
            self.sheet.append(row_data)

            # Save the workbook after each update
            self.workbook.save(self.file_path)

            self.log_interval = int(self.timer_value.get()) * 1000

            # Schedule the next log
            self.master.after(self.log_interval, self.log_data)
    
    def stop_logging(self):
        # Stop logging process and update button states
        self.logging_active = False
        # Disable the Stop Logging button and enable the Start Logging button
        self.stop_logging_button.configure(state=tk.DISABLED)
        self.start_logging_button.configure(state=tk.NORMAL)

        self.logging_active = False  # Stop the logging loop
        if hasattr(self, 'workbook'):
            self.workbook.save(self.file_path)  # Ensure the final save
            self.workbook.close()  # Close the workbook
        folder_path = os.path.dirname(self.file_path)
        os.startfile(folder_path)

    def folder_open(self):
        folder_path = os.path.join(os.path.expanduser("~"), "Documents", "ADE Log Folder")
        if os.path.exists(folder_path):
            os.startfile(folder_path)
        else:
            messagebox.showwarning("Folder Not Found", f"The folder '{folder_path}' does not exist.")

    def get_gauge_style(self, value, gauge_type):
        if gauge_type == "voltage":
            if value < 20:
                return "success"
            elif 20 <= value < 30:
                return "warning"
            elif 30 <= value < 35:
                return "danger"
            else:
                return "success"
        elif gauge_type == "current":
            if value < 20:
                return "success"
            elif 20 <= value < 50:
                return "info"
            elif 50 <= value < 120:
                return "warning"
            else:
                return "danger"
        elif gauge_type == "temperature":
            if value < 30:
                return "success"
            elif 30 <= value < 69:
                return "warning"
            elif value >= 70:
                return "danger"
            else:
                return "info"
        elif gauge_type == "capacity":
            if value < 20:
                return "danger"
            elif 20 <= value < 50:
                return "warning"
            elif 50 <= value < 80:
                return "info"
            else:
                return "success"
        elif gauge_type == "charging_current":
            if value < 20:
                return "success"
            elif 20 <= value < 50:
                return "info"
            elif 50 <= value < 80:
                return "warning"
            else:
                return "danger"
        else:
            return "info"  # Default style if gauge_type is unknown
 
    def toggle_button_style(self, var, button, control_type, active_battery):
        """
        Toggle button style and control charging/discharging for the given active battery.

        Parameters:
        - var: The BooleanVar tracking the state of the toggle (True/False).
        - button: The button object whose style will be updated.
        - control_type: 'charge' or 'discharge' to specify the type of control.
        - active_battery: 'Battery 1' or 'Battery 2' to specify which battery is being controlled.
        """
        # Determine which battery's control to operate on
        if active_battery == "Battery 1":
            charge_on = self.charger_control_var_battery_1.get()
            discharge_on = self.discharger_control_var_battery_1.get()
            # charge_button = self.charge_button_battery_1  # Assuming you have defined this button elsewhere
            # discharge_button = self.discharge_button_battery_1  # Assuming you have defined this button elsewhere
        elif active_battery == "Battery 2":
            charge_on = self.charger_control_var_battery_2.get()
            discharge_on = self.discharger_control_var_battery_2.get()
            # charge_button = self.charge_button_battery_2  # Assuming you have defined this button elsewhere
            # discharge_button = self.discharge_button_battery_2  # Assuming you have defined this button elsewhere

        # Ensure only one operation (charge/discharge) is active at a time
        if control_type == 'charge':
            if charge_on and discharge_on:
                messagebox.showwarning("Warning", "Please turn off discharging before enabling charging.")
                if active_battery == "Battery 1":
                    self.charger_control_var_battery_1.set(False)
                elif active_battery == "Battery 2":
                    self.charger_control_var_battery_2.set(False)
                return
            if active_battery == "Battery 1":
                self.charge_fet_status_battery_1 = charge_on
                if charge_on:
                    update_charging_ocv_in_excel(device_data_battery_1['serial_number'],device_data_battery_1['voltage'])
                    pcan_write_control('charge_on', battery_no=1)
                    self.update_charging_log_battery_1()
                    self.discharger_control_var_battery_1.set(False)  # Turn off discharging
                else:
                    pcan_write_control('both_off', battery_no=1)
                    # Stop logging by not calling update_charging_log_battery_1 anymore
                    self.master.after_cancel(self.update_charging_log_battery_1)  # Cancel the next scheduled logging
                    # Show a message to indicate that the charger is turned off
                    print("Charger Turned Off The charger has been turned off for Battery 1. Logging has been stopped.")
            elif active_battery == "Battery 2":
                self.charge_fet_status_battery_2 = charge_on
                if charge_on:
                    update_charging_ocv_in_excel(device_data_battery_2['serial_number'],device_data_battery_2['voltage'])
                    pcan_write_control('charge_on', battery_no=2)
                    self.update_charging_log_battery_2()
                    self.discharger_control_var_battery_2.set(False)  # Turn off discharging
                else:
                    pcan_write_control('both_off', battery_no=2)
                    # Stop logging by not calling update_charging_log_battery_1 anymore
                    self.master.after_cancel(self.update_charging_log_battery_2)  # Cancel the next scheduled logging
                    # Show a message to indicate that the charger is turned off
                    print("Charger Turned Off The charger has been turned off for Battery 2. Logging has been stopped.")

        elif control_type == 'discharge':
            if charge_on and discharge_on:
                messagebox.showwarning("Warning", "Please turn off charging before enabling discharging.")
                if active_battery == "Battery 1":
                    self.discharger_control_var_battery_1.set(False)
                elif active_battery == "Battery 2":
                    self.discharger_control_var_battery_2.set(False)
                return
            if active_battery == "Battery 1":
                self.discharge_fet_status_battery_1 = discharge_on
                if discharge_on:
                    update_discharging_ocv_in_excel(device_data_battery_1['serial_number'],device_data_battery_1['voltage'])
                    pcan_write_control('discharge_on', battery_no=1)
                    self.update_discharging_log_battery_1()
                    self.charger_control_var_battery_1.set(False)  # Turn off charging
                else:
                    pcan_write_control('both_off', battery_no=1)
                    # Stop logging by not calling update_charging_log_battery_1 anymore
                    self.master.after_cancel(self.update_discharging_log_battery_1)  # Cancel the next scheduled logging
                    # Show a message to indicate that the charger is turned off
                    print("Discharger Turned Off The discharging control has been turned off for Battery 1. Logging has been stopped.")
            elif active_battery == "Battery 2":
                self.discharge_fet_status_battery_2 = discharge_on
                if discharge_on:
                    update_discharging_ocv_in_excel(device_data_battery_2['serial_number'],device_data_battery_2['voltage'])
                    pcan_write_control('discharge_on', battery_no=2)
                    self.update_discharging_log_battery_2()
                    self.charger_control_var_battery_2.set(False)  # Turn off charging
                else:
                    pcan_write_control('both_off', battery_no=2)
                     # Stop logging by not calling update_charging_log_battery_1 anymore
                    self.master.after_cancel(self.update_discharging_log_battery_2)  # Cancel the next scheduled logging
                    # Show a message to indicate that the charger is turned off
                    print("Discharger Turned Off The discharging control has been turned off for Battery 2. Logging has been stopped.")

        # Show alert message
        action = "Charging" if control_type == 'charge' else "Discharging"
        if var.get():  # If the button is toggled ON (action started)
            messagebox.showinfo("Action Started", f"{action} function for {active_battery} started.")
        else:  # If the button is toggled OFF (action stopped)
            # Update button styles and show stop alert
            messagebox.showinfo("Action Stopped", f"{action} function for {active_battery} stopped.")

        # Disable the button for 10 seconds using threading
        button.config(state=tk.DISABLED)

        def reenable_button():
            # Wait for 10 seconds in a separate thread
            def wait_and_enable():
                time.sleep(10)
                try:
                    # Ensure the widget is still valid
                    if button.winfo_exists():
                        button.config(state=tk.NORMAL)
                except tk.TclError:
                    # The button no longer exists, so just pass
                    pass

            threading.Thread(target=wait_and_enable).start()

        # Start the cooldown timer
        reenable_button()

    # def bmsreset(self):
    #     """Show a progress bar at the top of the main window for BMS reset process."""

    #     # Create a frame at the top of the main window for the progress bar
    #     progress_frame = ctk.CTkFrame(self.master)
    #     progress_frame.pack(fill="x", padx=20, pady=10)

    #     # Add a label and progress bar to the frame
    #     label = ctk.CTkLabel(progress_frame, text="Resetting BMS, please wait...", font=("Helvetica", 12))
    #     label.pack(pady=5)

    #     progress_bar = ctk.CTkProgressBar(progress_frame, mode="determinate", progress_color="#45a049")
    #     progress_bar.pack(fill="x", padx=20, pady=5)

    #     # Function to update the progress bar
    #     def update_progress_bar():
    #         for i in range(101):  # 0% to 100%
    #             progress_bar.set(i / 100)  # Update progress bar value
    #             self.master.update_idletasks()  # Keep the UI responsive
    #             time.sleep(0.15)  # Simulate progress over time

    #     # Call the function to simulate progress
    #     update_progress_bar()

    #     # Perform the BMS reset command
    #     pcan_write_control('bms_reset')

    #     # Remove the progress bar frame after the reset is complete
    #     progress_frame.destroy()

    #     # Optionally, show a message box or status update after reset
    #     messagebox.showinfo("Success", "BMS Reset successfully completed.")
    
    def activate_heater(self):
        # Change the button to indicate the heater is activated
        self.activate_heater_button.configure(text="Heater Activated", fg_color="#4CAF50", hover_color="#45a049")

        if self.battery_var.get() == "Battery 1":
            # Call pcan_write_control to activate the heater for Battery 1
            pcan_write_control('heater_on',1)

        elif self.battery_var.get() == "Battery 2":
            # Call pcan_write_control to activate the heater for Battery 2
            pcan_write_control('heater_on',2)

        # Start a timer to revert the button after 10 seconds
        self.master.after(10000, self.reset_heater_button)

    def reset_heater_button(self):
        # Revert the button to its original state
        self.activate_heater_button.configure(text="Activate Heater", fg_color="#ff6361", hover_color="#d74a49")

    def check_battery_log_for_cycle_count(self):
        """
        Check the CAN data log file to see if the current battery's serial number exists
        and if the cycle count is 0. If the cycle count is 0, show a popup to update it.
        """
        # Define the folder and file path for the CAN data log
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        file_path = os.path.join(folder_path, "can_data.xlsx")

        # Check if the file exists
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "CAN data log file not found.")
            return False  # Return False if the file doesn't exist

        # Open the existing Excel file
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Loop through the rows to find the correct row by Serial Number
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        cycle_count_column = 8    # Assuming Cycle Count is in column H (8th column)

        battery_serial_number = self.device_data.get('serial_number')

        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == battery_serial_number:
                # Serial number found, check cycle count
                cycle_count = sheet.cell(row=row, column=cycle_count_column).value
                if cycle_count == 0:
                    print('Cycle count is 0, showing update prompt.')
                    self.first_time_dashboard = True
                    self.prompt_manual_cycle_count()  # Show the input dialog for cycle count update
                else:
                    print('Cycle count found and not 0.')
                    self.device_data['cycle_count'] = cycle_count
                    self.first_time_dashboard = False
                return True  # Exit the loop and return True once serial number is found

        # If loop completes without finding the serial number
        messagebox.showwarning("Warning", f"Serial number {battery_serial_number} not found.")
        return False  # Return False if the serial number is not found



