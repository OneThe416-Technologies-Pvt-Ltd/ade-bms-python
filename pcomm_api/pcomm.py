import serial
import threading
import struct
from openpyxl import Workbook, load_workbook
import datetime
import helpers.config as config
import os
from tkinter import messagebox
import pandas as pd
import helpers.pdf_generator as pdf_generator  # Config helper

battery_1_control_232=None
battery_2_control_232=None
battery_1_control_422_c_1=None
battery_1_control_422_c_2=None
battery_2_control_422_c_1=None
battery_2_control_422_c_2=None
rs_232_flag=False
rs_422_flag=False
rs232_interval_event = None
rs422_interval_event = None


rs232_device_1_data ={
            'battery_id': 1,
            'serial_number': 1,
            'hw_version': 1,
            'sw_version': 1,
            'cell_1_voltage': 100,
            'cell_2_voltage': 100,
            'cell_3_voltage': 100,
            'cell_4_voltage': 100,
            'cell_5_voltage': 100,
            'cell_6_voltage': 100,
            'cell_7_voltage': 100,
            'cell_1_temp': 20,
            'cell_2_temp': 20,
            'cell_3_temp': 20,
            'cell_4_temp': 20,
            'cell_5_temp': 20,
            'cell_6_temp': 20,
            'cell_7_temp': 20,
            'ic_temp': 25,
            'bus_1_voltage_before_diode': 100,
            'bus_2_voltage_before_diode': 100,
            'bus_1_voltage_after_diode': 100,
            'bus_2_voltage_after_diode': 100,
            'bus_1_current_sensor1': 100,
            'bus_2_current_sensor1': 100,
            'charger_input_current': 100,
            'charger_output_current': 100,
            'charger_output_voltage': 100,
            'charging_on_off_status': 1,
            'charger_relay_status': 1,
            'constant_voltage_mode': 1,
            'constant_current_mode': 1,
            'input_under_voltage': 100,
            'output_over_current': 100,
            'bus1_status': 0,
            'bus2_status': 1,
            'heater_pad': 1,
}


rs422_device_1_data ={
            'eb_1_relay_status': 0,
            'eb_2_relay_status': 0,
            'heater_pad_charger_relay_status': 0,
            'charger_status': 0,
            'voltage': 0,
            'eb_1_current': 0,
            'eb_2_current': 0,
            'charge_current': 0,
            'temperature': 0,
            'channel_selected': 0
}


rs232_device_2_data ={
            'battery_id': 1,
            'serial_number': 1,
            'hw_version': 1,
            'sw_version': 1,
            'cell_1_voltage': 2,
            'cell_2_voltage': 2,
            'cell_3_voltage': 2,
            'cell_4_voltage': 2,
            'cell_5_voltage': 2,
            'cell_6_voltage': 2,
            'cell_7_voltage': 2,
            'cell_1_temp': 1,
            'cell_2_temp': 1,
            'cell_3_temp': 1,
            'cell_4_temp': 1,
            'cell_5_temp': 1,
            'cell_6_temp': 1,
            'cell_7_temp': 1,
            'ic_temp': 25,
            'bus_1_voltage_before_diode': 100,
            'bus_2_voltage_before_diode': 100,
            'bus_1_voltage_after_diode': 100,
            'bus_2_voltage_after_diode': 100,
            'bus_1_current_sensor1': 100,
            'bus_2_current_sensor1': 100,
            'charger_input_current': 100,
            'charger_output_current': 100,
            'charger_output_voltage': 100,
            'charging_on_off_status': 1,
            'charger_relay_status': 1,
            'constant_voltage_mode': 1,
            'constant_current_mode': 1,
            'input_under_voltage': 100,
            'output_over_current': 100,
            'bus1_status': 1,
            'bus2_status': 0,
            'heater_pad': 1,
}


rs422_device_2_data ={
            'eb_1_relay_status': 0,
            'eb_2_relay_status': 0,
            'heater_pad_charger_relay_status': 0,
            'charger_status': 0,
            'voltage': 0,
            'eb_1_current': 0,
            'eb_2_current': 0,
            'charge_current': 0,
            'temperature': 0,
            'channel_selected': 0
}


rs232_write = {
    'header' : 0xAA,
    'cmd_byte_1' : 0x00,
    'cmd_byte_2' : 0x00,
    'cmd_byte_3' : 0x00,
    'cmd_byte_4' : 0x00,
}


rs422_write = {
    'header' : 0x55,
    'cmd_byte_1' : 0xC0,
    'cmd_byte_2' : 0x08,
    'cmd_byte_3' : 0x00,
}


# Custom setInterval function using threading
def set_interval_async(func, interval):
    """
    Repeatedly calls a function every `interval` seconds using threading.
    
    Args:
    - func: The function to call repeatedly.
    - interval: The time (in seconds) to wait between calls.
    
    Returns:
    - A threading.Event object used to stop the interval function.
    """
    # Event to stop the loop when needed
    stopped = threading.Event()

    def wrapper():
        """Wraps the function and ensures it stops when the stop event is set."""
        while not stopped.is_set():  # Loop until the stop event is triggered
            func()  # Call the function synchronously
            stopped.wait(interval)  # Wait for the interval or until the stop event is set

    # Start the function call in a separate daemon thread (it will stop when the main program exits)
    thread = threading.Thread(target=wrapper)
    thread.daemon = True  # Daemon thread will terminate when the main program ends
    thread.start()

    # Return the event object to allow stopping the interval function later
    return stopped


def connect_to_serial_port(flag):
    """
    Connect to a serial port for RS-232 or RS-422 communication and configure it with default settings.
    
    Args:
    - flag: Determines whether to use "RS-232" or "RS-422" communication.
    
    Returns:
    - True if both connections succeed, False if connection fails, None if there is an error.
    """
    # Define baud rate, data bits, parity, and stop bits for the serial connection
    baud_rate = 19200 if flag == "RS-232" else 9600  # Default baud rate for RS-232 or RS-422
    data_bits = serial.EIGHTBITS  # 8 data bits
    parity = serial.PARITY_NONE  # No parity checking
    stop_bits = serial.STOPBITS_ONE  # One stop bit for the serial communication

    # Retrieve port configurations from the `config` object (assuming it's a dictionary)
    battery_1_port_232 = config.config_values['rs_config'].get('battery_1_rs232')
    battery_2_port_232 = config.config_values['rs_config'].get('battery_2_rs232')

    try:
        if flag == "RS-232":  # Check if RS-232 flag is passed
            # Open serial connections for battery 1 and 2 using RS-232 settings
            battery_1_control_232 = serial.Serial(
                port=battery_1_port_232,
                baudrate=baud_rate,
                bytesize=data_bits,
                parity=parity,
                stopbits=stop_bits,
                timeout=1  # Timeout of 1 second for reading data
            )
            battery_2_control_232 = serial.Serial(
                port=battery_2_port_232,
                baudrate=baud_rate,
                bytesize=data_bits,
                parity=parity,
                stopbits=stop_bits,
                timeout=1  # Timeout of 1 second for reading data
            )

            # Check if both serial ports are open and ready
            if battery_1_control_232.is_open and battery_2_control_232.is_open:
                print(f"Connected to {battery_1_port_232} with baud rate {baud_rate}.")
                # Once connected, we could start periodic communication (uncomment to activate communication loop)
                # start_communication()
                return True  # Return True if both connections succeed
            else:
                return False  # Return False if either connection fails
    except serial.SerialException as e:
        # Handle any serial connection errors
        print(f"Error opening the serial port: {e}")
        return None  # Return None if there's an error opening the port


# Method to send RS232 or RS422 data based on the flag
def send_data():
    """Send RS232 or RS422 data periodically based on the flag."""
    while battery_1_control_232 and battery_1_control_232.is_open:
        if rs_232_flag:
            send_rs232_data()  # Send RS232 data if flag is set
        elif rs_422_flag:
            send_rs422_data()  # Send RS422 data if flag is set
        # battery_1_control_232.flush()  # Uncomment to flush the buffer if necessary
        threading.Event().wait(0.16)  # Wait for 160ms between each send


# Method to read RS232 or RS422 data based on the flag
def read_data():
    """Read RS232 or RS422 data periodically based on the flag."""
    while battery_1_control_232 and battery_1_control_232.is_open:
        if rs_232_flag:
            read_rs232_data()  # Read RS232 data if flag is set
        elif rs_422_flag:
            read_rs422_data()  # Read RS422 data if flag is set
        threading.Event().wait(0.16)  # Adjust the read interval if necessary


# Start a separate thread for sending data
def start_sending():
    """Start a separate thread to handle sending data periodically."""
    send_thread = threading.Thread(target=send_data, daemon=True)  # Create the thread as a daemon
    send_thread.start()  # Start the thread


# Start a separate thread for reading data
def start_reading():
    """Start a separate thread to handle reading data periodically."""
    read_thread = threading.Thread(target=read_data, daemon=True)  # Create the thread as a daemon
    read_thread.start()  # Start the thread


# Send RS232 data
def send_rs232_data():
    """Send RS232 data by creating and sending a packet."""
    # if battery_1_control_232.is_open:  # Uncomment if you need to check if port is open
    data = create_rs232_packet()  # Create the RS232 packet
    print(f"{data} data send")  # Log the data being sent
    battery_1_control_232.write(data)  # Write the data to the serial port
    print(f"Sent RS232 data: {data.hex()}")  # Log the hex format of the sent data


# Send RS422 data
def send_rs422_data():
    """Send RS422 data by creating and sending a packet."""
    # if battery_1_control_232.is_open:  # Uncomment if you need to check if port is open
    data = create_rs422_packet()  # Create the RS422 packet
    battery_1_control_232.write(data)  # Write the data to the serial port
    print(f"Sent RS422 data: {data.hex()}")  # Log the hex format of the sent data


# Read RS232 data, find a valid 64-byte block and update device data
def read_rs232_data():
    """Read 256 bytes of data, find the valid 64-byte block, and update RS232 data."""
    if battery_1_control_232.is_open:
        received_data = battery_1_control_232.read(256)  # Read 256 bytes of data
        if len(received_data) == 256:
            # Log the received data in hex format for debugging purposes
            received_hex = received_data.hex()
            print(f"RS232 Data received: {received_hex}")

            # Search for the start (0x66AA) and end (0x55BB) of the 64-byte data block
            for i in range(0, 192):  # Search up to byte 192 to have room for 64 bytes
                if received_data[i] == 0x66 and received_data[i+1] == 0xAA:
                    if received_data[i+62] == 0x55 and received_data[i+63] == 0xBB:
                        # Extract the 64-byte block
                        valid_block = received_data[i:i+64]
                        print(f"RS232 valid Data received: {valid_block.hex()}")
                        print(f"Valid RS232 data block found from index {i} to {i+63}.")
                        # Pass the valid block to update the device data
                        update_rs232_device_1_data(valid_block)
                        log_rs_data(rs232_device_1_data)
                        return  # Exit after finding and processing the valid block
            print("No valid RS232 data block found in the received 256 bytes.")
        else:
            print("Received data is not 256 bytes long. Discarding data.")


# Read RS422 data, validate checksum, and update device data
def read_rs422_data():
    """Read 40 bytes, validate the first 18-byte RS422 data block, and update the device data."""
    if battery_1_control_232.is_open:
        received_data = battery_1_control_232.read(40)  # Read 40 bytes of data
        if len(received_data) == 40:
            print(f"RS422 Data received (40 bytes): {received_data.hex()}")

            # Search for 0x55 within the received data
            start_index = -1  # Initialize with an invalid index
            for i in range(len(received_data)):
                if received_data[i] == 0x55:
                    start_index = i
                    break

            if start_index != -1:
                print(f"Found 0x55 at index {start_index}. Proceeding with validation.")

                # Ensure we have enough data for a 40-byte block starting from this index
                if start_index + 17 <= len(received_data):
                    block = received_data[start_index:start_index + 19]

                    # Calculate checksum for the first 39 bytes
                    calculated_checksum = 0
                    for byte in block[0:17]:
                        calculated_checksum ^= byte  # XOR each byte

                    # Validate checksum against the 40th byte
                    if calculated_checksum == block[18]:
                        print(f"Valid RS422 data block found: {block.hex()}")
                        # Pass the valid 40-byte block to update the device data
                        update_rs422_device_1_data(block)
                    else:
                        print(f"Checksum mismatch. Calculated: {calculated_checksum}, Received: {block[18]}. Discarding block.")
                else:
                    print(f"Not enough data to process a full 40-byte block starting from index {start_index}.")
            else:
                print("0x55 not found in the received data. Discarding block.")
        else:
            print("Received data is not 40 bytes long. Discarding data.")


# Create an RS232 packet to send
def create_rs232_packet():
    """Create an RS232 packet to send."""
    # Define the packet structure based on the rs232_write dictionary
    packet = [rs232_write['header'], rs232_write['cmd_byte_1'], rs232_write['cmd_byte_2'], rs232_write['cmd_byte_3'], rs232_write['cmd_byte_4']]
    checksum = calculate_checksum(packet[0:4])  # Calculate checksum for the first 4 bytes
    packet.append(checksum)  # Append the checksum to the packet
    packet.append(0x55)  # Footer byte

    # Convert the packet to a hex string for display/logging purposes
    hex_string = ''.join([f'{byte:02X}' for byte in packet])

    # Convert the hex string back to bytes and return
    final_packet = bytes.fromhex(hex_string)
    return final_packet


# Create an RS422 packet to send
def create_rs422_packet():
    """Create an RS422 packet to send."""
    # Define the initial part of the RS422 packet
    packet = [rs422_write['header'], rs422_write['cmd_byte_1'], rs422_write['cmd_byte_2']]
    pmu_heartbeat = 0x0000  # Dummy heartbeat data for the packet
    packet.extend(struct.pack('<H', pmu_heartbeat))  # Add the heartbeat data
    checksum = calculate_checksum(packet)  # Calculate checksum for the packet
    packet.append(checksum)  # Append the checksum

    # Return the packet as a bytearray
    return bytearray(packet)


def update_rs232_device_1_data(data_hex):
    """Update rs232_device_1_data with the received bytes."""
    
    # Helper function to calculate and validate cell voltage
    def calculate_cell_voltage(raw_data):
        try:
            # Raw data + 2 to adjust the value
            cell_voltage = raw_data * 0.01 + 2
            # Ensure the voltage is in the range of 2 to 4.5V
            if 2.0 <= cell_voltage <= 4.5:
                return round(cell_voltage, 2)
            else:
                print(f"Warning: Cell voltage {cell_voltage}V out of range.")
                return None
        except Exception as e:
            print(f"Error in calculating cell voltage: {e}")
            return None
    
    try:
        # Byte 2: Battery ID
        rs232_device_1_data['battery_id'] = data_hex[2]

        # Byte 3: Battery Serial Number (1-255)
        rs232_device_1_data['serial_number'] = data_hex[3]

        # Byte 4: Hardware Version (1-9)
        rs232_device_1_data['hw_version'] = data_hex[4]

        # Byte 5: Software Version (1-9)
        rs232_device_1_data['sw_version'] = data_hex[5]

        # Process each cell voltage (adding 2 to raw value and ensuring it's within range)
        # Bytes 6-7: Cell-1 Voltage (Resolution 0.01V)
        cell_1_voltage_raw = data_hex[6]
        rs232_device_1_data['cell_1_voltage'] = calculate_cell_voltage(cell_1_voltage_raw)
        
        # Bytes 8-9: Cell-2 Voltage (Resolution 0.01V)
        cell_2_voltage_raw = data_hex[7]
        rs232_device_1_data['cell_2_voltage'] = calculate_cell_voltage(cell_2_voltage_raw)
        
        # Bytes 10-11: Cell-3 Voltage (Resolution 0.01V)
        cell_3_voltage_raw = data_hex[8]
        rs232_device_1_data['cell_3_voltage'] = calculate_cell_voltage(cell_3_voltage_raw)
        
        # Bytes 12-13: Cell-4 Voltage (Resolution 0.01V)
        cell_4_voltage_raw = data_hex[9]
        rs232_device_1_data['cell_4_voltage'] = calculate_cell_voltage(cell_4_voltage_raw)
        
        # Bytes 14-15: Cell-5 Voltage (Resolution 0.01V)
        cell_5_voltage_raw = data_hex[10]
        rs232_device_1_data['cell_5_voltage'] = calculate_cell_voltage(cell_5_voltage_raw)
        
        # Bytes 16-17: Cell-6 Voltage (Resolution 0.01V)
        cell_6_voltage_raw = data_hex[11]
        rs232_device_1_data['cell_6_voltage'] = calculate_cell_voltage(cell_6_voltage_raw)
        
        # Bytes 18-19: Cell-7 Voltage (Resolution 0.01V)
        cell_7_voltage_raw = data_hex[12]
        rs232_device_1_data['cell_7_voltage'] = calculate_cell_voltage(cell_7_voltage_raw)

        # Byte 20: Cell-1 Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['cell_1_temp'] = data_hex[13] - 40

        # Byte 21: Cell-2 Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['cell_2_temp'] = data_hex[14] - 40

        # Byte 22: Cell-3 Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['cell_3_temp'] = data_hex[15] - 40

        # Byte 23: Cell-4 Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['cell_4_temp'] = data_hex[16] - 40

        # Byte 24: Cell-5 Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['cell_5_temp'] = data_hex[17] - 40

        # Byte 25: Cell-6 Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['cell_6_temp'] = data_hex[18] - 40

        # Byte 26: Cell-7 Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['cell_7_temp'] = data_hex[19] - 40

        cell_status_byte = data_hex[20]  # Byte 21 in the data (index 20)
        binary_status = format(cell_status_byte, '08b')  # Convert to 8-bit binary string

        # Get the status of each cell from the corresponding bit
        rs232_device_1_data['cell_1_status'] = (cell_status_byte & 0x01)  # Bit 0: Cell 1 Status
        rs232_device_1_data['cell_2_status'] = (cell_status_byte & 0x02) >> 1  # Bit 1: Cell 2 Status
        rs232_device_1_data['cell_3_status'] = (cell_status_byte & 0x04) >> 2  # Bit 2: Cell 3 Status
        rs232_device_1_data['cell_4_status'] = (cell_status_byte & 0x08) >> 3  # Bit 3: Cell 4 Status
        rs232_device_1_data['cell_5_status'] = (cell_status_byte & 0x10) >> 4  # Bit 4: Cell 5 Status
        rs232_device_1_data['cell_6_status'] = (cell_status_byte & 0x20) >> 5  # Bit 5: Cell 6 Status
        rs232_device_1_data['cell_7_status'] = (cell_status_byte & 0x40) >> 6  # Bit 6: Cell 7 Status

        # Byte 27: Monitor IC Temperature (-40 to +125, subtract 40)
        rs232_device_1_data['ic_temp'] = data_hex[22] - 40

        # Bytes 28-29: Bus 1 Voltage (18-33.3V, Resolution 0.06V)
        rs232_device_1_data['bus_1_voltage_before_diode'] = round(data_hex[23] * 0.06, 2)

        # Bytes 30-31: Bus 2 Voltage (18-33.3V, Resolution 0.06V)
        rs232_device_1_data['bus_2_voltage_before_diode'] = round(data_hex[24] * 0.06, 2)

        # Bytes 32-33: Bus 1 Voltage after Diode (Resolution 0.06V)
        rs232_device_1_data['bus_1_voltage_after_diode'] = round(data_hex[25] * 0.06, 2)

        # Bytes 34-35: Bus 2 Voltage after Diode (Resolution 0.06V)
        rs232_device_1_data['bus_2_voltage_after_diode'] = round(data_hex[26] * 0.06, 2)

        # Bytes 36-37: Current / Bus 1 Sensor 1 (0-63.75A, Resolution 0.25A)
        rs232_device_1_data['bus_1_current_sensor1'] = round((data_hex[27] - 128) * 0.25, 2)

        # Bytes 38-39: Current / Bus 2 Sensor 2 (0-63.75A, Resolution 0.25A)
        rs232_device_1_data['bus_2_current_sensor1'] = round((data_hex[29] - 128) * 0.25, 2)

        # Byte 40-41: Charger Input Current (0-15A, Resolution 0.1A)
        rs232_device_1_data['charger_input_current'] = round(data_hex[31] * 0.1, 1)

        # Bytes 32-33: Charger Output Current (Resolution 0.1A)
        rs232_device_1_data['charger_output_current'] = round(data_hex[32] * 0.1, 2)

        # Bytes 34-35: Charger Output Voltage (Resolution 0.1V)
        rs232_device_1_data['charger_output_voltage'] = round(data_hex[33] * 0.06, 2)

        charger_status_byte = data_hex[35]  # Byte 35 in the data

        # Get the charging on/off status (0th bit)
        rs232_device_1_data['charging_on_off_status'] = (charger_status_byte & 0x01)  # Bit 0: Charging On/Off Status

        # Byte 37: Charger Relay Status
        rs232_device_1_data['charger_relay_status'] = (charger_status_byte & 0x02) >> 1

        # Byte 38: Constant Voltage Mode (1 byte)
        rs232_device_1_data['constant_voltage_mode'] = (charger_status_byte & 0x04) >> 2

        # Byte 39: Constant Current Mode (1 byte)
        rs232_device_1_data['constant_current_mode'] = (charger_status_byte & 0x08) >> 3

        # Bus 1 status (Byte 41)
        rs232_device_1_data['bus_1_status'] = (data_hex[41] & 0x01)
        rs232_device_1_data['bus_2_status'] = (data_hex[41] & 0x02) >> 1
        rs232_device_1_data['heater_pad_status'] = (data_hex[41] & 0x04) >> 2

        print("Updated rs232_device_1_data:")
        print(rs232_device_1_data)
    
    except IndexError as e:
        print(f"Error: Data length mismatch. Expected at least {len(rs232_device_1_data)} bytes. Exception: {e}")
    except Exception as e:
        print(f"Unexpected error occurred: {e}")


def update_rs422_device_1_data(data_bytes):
    """Update rs422_device_1_data with the received bytes."""
    
    # Helper function to safely extract data and handle any errors
    def extract_byte_data(index, resolution_factor=1, offset=0):
        try:
            raw_data = data_bytes[index]
            return (raw_data * resolution_factor) + offset
        except IndexError:
            print(f"Error: Missing data at index {index}.")
            return None
        except Exception as e:
            print(f"Unexpected error when extracting data at index {index}: {e}")
            return None

    try:
        # Byte 1: Relay status byte
        raw_status_byte = data_bytes[1]  # First byte (Byte 1)
        
        # Extract and assign relay status using bitwise operations
        rs422_device_1_data['eb_1_relay_status'] = (raw_status_byte & 0x01)  # Bit 0: EB1 Relay Status
        rs422_device_1_data['eb_2_relay_status'] = (raw_status_byte & 0x02) >> 1  # Bit 1: EB2 Relay Status
        rs422_device_1_data['heater_pad_status'] = (raw_status_byte & 0x04) >> 2  # Bit 2: Heater Pad Status
        rs422_device_1_data['charger_output_relay_status'] = (raw_status_byte & 0x08) >> 3  # Bit 3: Charger Output Relay Status
        
        # Byte 3: Channel status byte
        channel_status_byte = data_bytes[3]
        
        # Extract channel packet counts (bits 3:0 for channel 1 and bits 7:4 for channel 2)
        channel_1_count = channel_status_byte & 0x0F  # Mask with 0x0F to get bits 3:0
        channel_2_count = (channel_status_byte & 0xF0) >> 4  # Mask with 0xF0 and shift right by 4 to get bits 7:4
        
        # Determine the health status for channels
        if 0 < channel_1_count < 15:
            rs422_device_1_data['channel_1_status'] = 1  # Healthy
        else:
            rs422_device_1_data['channel_1_status'] = 0  # Not Healthy
        
        if 0 < channel_2_count < 15:
            rs422_device_1_data['channel_2_status'] = 1  # Healthy
        else:
            rs422_device_1_data['channel_2_status'] = 0  # Not Healthy

        # Determine which channel is selected based on status
        if rs422_device_1_data['channel_1_status'] == 1:
            rs422_device_1_data['channel_selected'] = 1  # Channel 1 selected
        elif rs422_device_1_data['channel_2_status'] == 1:
            rs422_device_1_data['channel_selected'] = 2  # Channel 2 selected
        else:
            rs422_device_1_data['channel_selected'] = 0  # No healthy channel
        
        # Byte 4: Voltage (Resolution 0.15V)
        raw_voltage = data_bytes[4]
        rs422_device_1_data['voltage'] = raw_voltage * 0.15
        
        # Byte 5: EB-1 Current (Resolution 0.5A, offset -63.75A)
        raw_eb1_current = data_bytes[5]
        rs422_device_1_data['eb_1_current'] = (raw_eb1_current * 0.5) - 63.75
        
        # Byte 6: EB-2 Current (Resolution 0.5A, offset -63.75A)
        raw_eb2_current = data_bytes[6]
        rs422_device_1_data['eb_2_current'] = (raw_eb2_current * 0.5) - 63.75
        
        # Byte 7: Charge Current (Resolution 0.05A)
        raw_charge_current = data_bytes[7]
        rs422_device_1_data['charge_current'] = raw_charge_current * 0.05
        
        # Byte 8: Battery Temperature (Resolution 0.5°C, offset -40°C)
        raw_temperature = data_bytes[8]
        rs422_device_1_data['temperature'] = raw_temperature - 40
        
        # Byte 9: State of Charge (No resolution conversion, just raw value)
        raw_state_of_charge = data_bytes[9]
        rs422_device_1_data['state_of_charge'] = raw_state_of_charge
        
        # Print updated device data for debugging
        print(f"Updated RS422 device data: {rs422_device_1_data}")
    
    except IndexError as e:
        print(f"Error: Data length mismatch. Expected at least {len(data_bytes)} bytes. Exception: {e}")
    except Exception as e:
        print(f"Unexpected error occurred: {e}")


def calculate_checksum(data):
    """Calculate XOR checksum for a given byte array."""
    checksum = 0x00
    for byte in data:
        checksum ^= byte  # Perform XOR operation on each byte in the data
    return checksum


def start_communication():
    """Start communication by initiating both sending and reading operations."""
    start_sending()  # Placeholder for the function that handles sending data
    start_reading()  # Placeholder for the function that handles reading data


def stop_communication():
    """Stop communication by terminating ongoing RS232/RS422 intervals and closing serial connections."""
    global rs_232_flag, rs_422_flag, rs232_interval_event, rs422_interval_event

    try:
        # Stop RS232 or RS422 intervals if they are running
        if rs232_interval_event:
            rs232_interval_event.set()  # Stop RS232 interval
        if rs422_interval_event:
            rs422_interval_event.set()  # Stop RS422 interval

        # Update flags to indicate communication is stopped
        rs_232_flag = False
        rs_422_flag = False

        # Close the RS232 connection if it's open
        if battery_1_control_232 and battery_1_control_232.is_open:
            battery_1_control_232.close()
            print("Serial connection closed.")
    except Exception as e:
        print(f"Error in stop_communication: {e}")
        messagebox.showerror("Error", f"An error occurred while stopping communication: {e}")


def get_active_protocol():
    """Get the currently active protocol (RS-232 or RS-422)."""
    global rs_422_flag, rs_232_flag
    if rs_232_flag:
        return "RS-232"
    elif rs_422_flag:
        return "RS-422"
    else:
        return "None"


def set_active_protocol(selected_flag):
    """Set the active communication protocol (RS-232 or RS-422)."""
    global rs_422_flag, rs_232_flag
    try:
        if selected_flag == "RS-232":
            rs_232_flag = True
        elif selected_flag == "RS-422":
            rs_422_flag = True
        else:
            print(f"Invalid protocol selected: {selected_flag}")
    except Exception as e:
        print(f"Error setting active protocol: {e}")


def log_rs_data(update_rs_data):
    """
    Log RS232 data to an Excel file. If the serial number exists, do nothing;
    if it doesn't exist, append a new row with default values.
    """
    try:
        # Define the path for the RS232 data Excel file inside the AppData directory
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)  # Create folder if it does not exist

        # Define the log file name
        file_name = "rs_data.xlsx"
        file_path = os.path.join(folder_path, file_name)

        # Create or load workbook
        if os.path.exists(file_path):
            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active
            except Exception as e:
                print(f"Error loading the Excel file: {str(e)}")
                messagebox.showerror("Error", f"Failed to load the Excel file: {str(e)}. Recreating the file.")
                workbook = Workbook()  # Recreate the workbook
                sheet = workbook.active
                sheet.title = "RS232 Data"
                # Add headers for RS232 data
                headers = [
                    "SI No", "Date", "Time", "Project", "Device Name", "Manufacturer Name", 
                    "Serial Number", "Cycle Count", "Full Charge Capacity", "Charging Date", 
                    "OCV Before Charging", "Discharging Date", "OCV Before Discharging"
                ]
                sheet.append(headers)
        else:
            # File does not exist, create a new workbook and add headers
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "RS232 Data"
            headers = [
                "SI No", "Date", "Time", "Project", "Device Name", "Manufacturer Name", 
                "Serial Number", "Cycle Count", "Full Charge Capacity", "Charging Date", 
                "OCV Before Charging", "Discharging Date", "OCV Before Discharging"
            ]
            sheet.append(headers)

        # Log the current RS232 data
        current_datetime = datetime.datetime.now()
        date = current_datetime.strftime("%Y-%m-%d")
        time = current_datetime.strftime("%H:%M:%S")
        serial_number = update_rs_data.get('serial_number', 'N/A')

        # Check if the serial number already exists in the Excel file
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        serial_found = False

        # Loop through rows to check if the serial number already exists
        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                serial_found = True
                print(f"Serial number {serial_number} already exists. No action taken.")
                break

        if not serial_found:
            # If serial number is not found, add a new row with default values
            next_row = sheet.max_row + 1
            rs_data = [
                next_row - 1,  # SI No
                date,  # Date
                time,  # Time
                update_rs_data.get('project', ''),  # Project
                update_rs_data.get('device_name', 'BT-70939APH'),  # Device Name
                update_rs_data.get('manufacturer_name', 'Bren-Tronics'),  # Manufacturer Name
                serial_number,  # Serial Number
                int(update_rs_data.get('cycle_count', 0)),  # Ensure Cycle Count is an int
                103,  # Full Charge Capacity as a default float
                update_rs_data.get('charging_date', date),  # Charging Date
                float(update_rs_data.get('ocv_before_charging', 0.0)),  # OCV Before Charging as float
                update_rs_data.get('discharging_date', date),  # Discharging Date
                float(update_rs_data.get('ocv_before_discharging', 0.0))  # OCV Before Discharging as float
            ]
            sheet.append(rs_data)

        # Save the updated Excel file
        workbook.save(file_path)
        workbook.close()
        print(f"RS232 data logged in {file_path}.")

    except Exception as e:
        print(f"An error occurred while updating the Excel file: {str(e)}")
        messagebox.showerror("Error", f"An error occurred while updating the Excel file: {str(e)}")


def get_latest_rs_data(serial_number):
    """
    Retrieves the most recent RS232 data for the specified serial number.
    """
    rs_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "rs_data.xlsx")

    if not os.path.exists(rs_data_file):
        messagebox.showerror("Error", "No RS232 data file found. Please log RS232 data first.")
        return None

    # Load the existing workbook and access the first sheet
    try:
        workbook = load_workbook(rs_data_file)
        sheet = workbook.active
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load RS232 data file: {e}")
        return None

    serial_number_column = 7  # Serial Number is in column G (7th column)
    serial_found = False
    latest_data = {}

    # Search for the serial number and store the corresponding row data
    for row in range(2, sheet.max_row + 1):
        current_serial_number = sheet.cell(row=row, column=serial_number_column).value
        if current_serial_number == serial_number:
            serial_found = True
            print(f"Serial Number {serial_number} found in row {row}")
            headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
            values = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
            latest_data = dict(zip(headers, values))
            for header, value in latest_data.items():
                print(f"{header}: {value}")  # Debug print for fetched values
            break

    if not serial_found:
        messagebox.showwarning("Warning", f"Serial number {serial_number} not found in the RS232 data.")
        return None

    return latest_data


def update_cycle_count_in_rs_data(serial_number, new_cycle_count):
    """
    Update the cycle count for a specific device identified by its serial number
    in the RS232 Data Excel file.
    """
    folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
    file_path = os.path.join(folder_path, "rs_data.xlsx")

    if not os.path.exists(file_path):
        messagebox.showerror("Error", "RS232 data file not found.")
        return

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load RS232 data file: {e}")
        return

    serial_number_column = 7  # Serial Number is in column G (7th column)
    cycle_count_column = 8    # Cycle Count is in column H (8th column)

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=serial_number_column).value == serial_number:
            sheet.cell(row=row, column=cycle_count_column).value = new_cycle_count
            break
    else:
        messagebox.showwarning("Warning", f"Serial number {serial_number} not found.")
        return

    try:
        workbook.save(file_path)
        messagebox.showinfo("Success", f"Cycle count updated successfully for Serial Number {serial_number}.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save RS232 data: {e}")


def update_excel_and_download_pdf(data):
    """
    Get the updated values from the form, update the Excel file, and generate a PDF.
    """
    rs_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "rs_data.xlsx")

    try:
        df_rs_data = pd.read_excel(rs_data_file)
    except FileNotFoundError:
        messagebox.showerror("Error", f"{rs_data_file} not found.")
        return

    serial_number = data[2]  # Serial Number is assumed to be the 3rd element
    df_rs_data['Serial Number'] = df_rs_data['Serial Number'].astype(str).str.strip()
    df_rs_data['OCV Before Charging'] = df_rs_data['OCV Before Charging'].astype(float)
    df_rs_data['OCV Before Discharging'] = df_rs_data['OCV Before Discharging'].astype(float)

    index = df_rs_data[df_rs_data['Serial Number'] == serial_number].index
    if not index.empty:
        # Update the values in the DataFrame
        df_rs_data.loc[index[0], 'Project'] = str(data[0])
        df_rs_data.loc[index[0], 'Device Name'] = str(data[1])
        df_rs_data.loc[index[0], 'Manufacturer Name'] = str(data[3])
        df_rs_data.loc[index[0], 'Cycle Count'] = int(data[4])
        df_rs_data.loc[index[0], 'Full Charge Capacity'] = float(data[5])
        df_rs_data.loc[index[0], 'Charging Date'] = str(data[6])
        df_rs_data.loc[index[0], 'OCV Before Charging'] = float(data[7])
        df_rs_data.loc[index[0], 'Discharging Date'] = str(data[8])
        df_rs_data.loc[index[0], 'OCV Before Discharging'] = float(data[9])

        # Save the updated DataFrame back to the Excel file
        try:
            df_rs_data.to_excel(rs_data_file, index=False)
            messagebox.showinfo("Success", "Data updated successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save updated data to Excel: {e}")
    else:
        messagebox.showwarning("Warning", "Serial Number not found in the Excel file.")

    # Generate the PDF with the updated values
    pdf_generator.create_rs_report_pdf(serial_number, "RS232")
