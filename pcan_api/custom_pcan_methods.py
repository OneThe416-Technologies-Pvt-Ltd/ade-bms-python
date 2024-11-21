#pcan_methods.py

# Import necessary libraries and modules
from pcan_api.pcan import *  # Import the PCAN API for CAN bus communication
from tkinter import messagebox  # For showing messages to the user
import time  # For time-related functions
import pandas as pd  # For data manipulation and analysis
import asyncio  # For asynchronous programming
from openpyxl import Workbook, load_workbook  # For working with Excel files
import datetime  # For date and time functions
from fpdf import FPDF  # For generating PDF reports
import traceback
import os  # For file system operations
import helpers.pdf_generator as pdf_generator  # Helper for generating PDFs
import helpers.config as config  # Config helper for configuration management
from helpers.logger import logger  # Logger for debugging and error tracking

# Initialize the PCANBasic object for CAN communication
m_objPCANBasic = PCANBasic()
# Set the PCAN handle to a specific value (81 in this case)
m_PcanHandle = 81
# Flag to indicate whether current data is being fetched
is_fetching_current = False

# Path to store the configuration file (e.g., in the user's home directory)
CONFIG_FILE_PATH = os.path.join(os.path.expanduser("~"), "device_config.json")

# Dictionary for battery status flags; each key represents a specific flag status.
battery_status_flags = {
    "overcharged_alarm": 0,  # Bit 15 indicates overcharge alarm
    "terminate_charge_alarm": 0,  # Bit 14 indicates termination of charge alarm
    "over_temperature_alarm": 0,  # Bit 12 indicates over-temperature alarm
    "terminate_discharge_alarm": 0,  # Bit 11 indicates termination of discharge alarm
    "remaining_capacity_alarm": 0,  # Bit 9 indicates remaining capacity alarm
    "remaining_time_alarm": 0,  # Bit 8 indicates remaining time alarm
    "initialization": 0,  # Bit 7 indicates initialization status
    "charge_fet_test": 1,  # Bit 6 indicates Charge FET test status
    "fully_charged": 0,  # Bit 5 indicates if the battery is fully charged
    "fully_discharged": 0,  # Bit 4 indicates if the battery is fully discharged
    "error_codes": 0  # Bits 3:0 indicate specific error codes
}

# Battery 1 status flags
battery_1_status_flags = {
    "overcharged_alarm": 1,  # Overcharge alarm is on
    "terminate_charge_alarm": 1,  # Termination of charge alarm is on
    "over_temperature_alarm": 1,  # Over-temperature alarm is on
    "terminate_discharge_alarm": 1,  # Termination of discharge alarm is on
    "remaining_capacity_alarm": 1,  # Remaining capacity alarm is on
    "remaining_time_alarm": 1,  # Remaining time alarm is on
    "initialization": 1,  # Initialization status is on
    "charge_fet_test": 1,  # Charge FET test is on
    "fully_charged": 0,  # Battery is not fully charged
    "fully_discharged": 1,  # Battery is fully discharged
    "error_codes": 1  # Error code set
}

# Battery 2 status flags (similar to Battery 1 but with different values)
battery_2_status_flags = {
    "overcharged_alarm": 0,  # No overcharge alarm
    "terminate_charge_alarm": 0,  # No termination of charge alarm
    "over_temperature_alarm": 0,  # No over-temperature alarm
    "terminate_discharge_alarm": 0,  # No termination of discharge alarm
    "remaining_capacity_alarm": 0,  # No remaining capacity alarm
    "remaining_time_alarm": 0,  # No remaining time alarm
    "initialization": 0,  # Initialization status is off
    "charge_fet_test": 1,  # Charge FET test is on
    "fully_charged": 0,  # Battery is not fully charged
    "fully_discharged": 0,  # Battery is not fully discharged
    "error_codes": 0  # No error codes
}

# Dictionary for mapping internal names to user-friendly names for data display
name_mapping = {
    "device_name": "Device Name",
    "serial_number": "Serial No",
    "firmware_version": "Firmware Version",
    "manufacturer_date": "Manufacturer Date",
    "manufacturer_name": "Manufacturer Name",
    "battery_status": "Battery Status",
    "cycle_count": "Cycle Count",
    "design_capacity": "Design Capacity",
    "design_voltage": "Design Voltage",
    "at_rate_ok_text": "At Rate OK",
    "at_rate_time_to_full": "At Rate Time To Full",
    "at_rate_time_to_empty": "At Rate Time To Empty",
    "at_rate": "At Rate",
    "rel_state_of_charge": "Rel State of Charge",
    "abs_state_of_charge": "Absolute State of Charge",
    "run_time_to_empty": "Run Time To Empty",
    "avg_time_to_empty": "Avg Time To Empty",
    "avg_time_to_full": "Avg Time To Full",
    "charging_battery_status": "Charging Status",
    "max_error": "Max Error",
    "temperature": "Temperature",
    "current": "Current",
    "remaining_capacity": "Remaining Capacity",
    "voltage": "Voltage",
    "avg_current": "Avg Current",
    "charging_current": "Charging Current",
    "full_charge_capacity": "Full Charge Capacity",
    "charging_voltage": "Charging Voltage"
}

# Dictionary for mapping each device data attribute to its respective unit
unit_mapping = {
    "device_name": "string",
    "firmware_version": "string",
    "serial_number": "number",
    "manufacturer_date": "unsigned int",
    "manufacturer_name": "string",
    "battery_status": "bit flags",
    "cycle_count": "Count",
    "manual_cycle_count": "Count",
    "design_capacity": "Ah",
    "design_voltage": "mV",
    "at_rate_ok_text": "Yes/No",
    "at_rate_time_to_full": "minutes",
    "at_rate_time_to_empty": "minutes",
    "at_rate": "A",
    "rel_state_of_charge": "Percent",
    "abs_state_of_charge": "Percent",
    "run_time_to_empty": "minutes",
    "avg_time_to_empty": "minutes",
    "charging_battery_status": "string",
    "avg_time_to_full": "minutes",
    "max_error": "Percent",
    "temperature": "°C",
    "current": "A",
    "remaining_capacity": "Ah",
    "voltage": "mV",
    "avg_current": "A",
    "charging_current": "A",
    "full_charge_capacity": "Ah",
    "charging_voltage": "mV"
}

# Device data dictionary for storing general device information (default values)
device_data = {
    'device_name': "BT-70939APH",
    'serial_number': 0,
    'manufacturer_name': "Bren-Tronics",
    'firmware_version': "",
    'battery_status': "",
    'cycle_count': 0,
    'design_capacity': 0,
    'design_voltage': 0,
    'remaining_capacity': 0,
    'temperature': 0,
    'current': 0,
    'voltage': 0,
    'avg_current': 0,
    'charging_current': 0,
    'full_charge_capacity': 0,
    'charging_voltage': 0,
    'at_rate_time_to_full': 0,
    'at_rate_time_to_empty': 0,
    'at_rate_ok_text': "",
    'at_rate': 0,
    'charging_battery_status': "Off",
    'rel_state_of_charge': 0,
    'abs_state_of_charge': 0,
    'run_time_to_empty': 0,
    'avg_time_to_empty': 0,
    'avg_time_to_full': 0,
    'max_error': 0
}

# Battery 1 specific data dictionary with values specific to Battery 1
device_data_battery_1 = {
    'device_name': "BT-70939APH",
    'serial_number': 1478,
    'manufacturer_name': "Bren-Tronics",
    'firmware_version': "",
    'battery_status': "",
    'cycle_count': 0,
    'design_capacity': 0,
    'design_voltage': 0,
    'remaining_capacity': 0,
    'temperature': 0,
    'current': 0,
    'voltage': 0,
    'avg_current': 0,
    'charging_current': 100,
    'full_charge_capacity': 103.8,
    'charging_voltage': 0,
    'at_rate_time_to_full': 0,
    'at_rate_time_to_empty': 0,
    'at_rate_ok_text': "",
    'at_rate': 0,
    'charging_battery_status': "Off",
    'rel_state_of_charge': 0,
    'abs_state_of_charge': 0,
    'run_time_to_empty': 0,
    'avg_time_to_empty': 0,
    'avg_time_to_full': 0,
    'max_error': 0
}

# Battery 2 specific data dictionary with values specific to Battery 2
device_data_battery_2 = {
    'device_name': "BT-70939APH",
    'serial_number': 0,
    'manufacturer_name': "Bren-Tronics",
    'firmware_version': "",
    'battery_status': "",
    'cycle_count': 0,
    'design_capacity': 0,
    'design_voltage': 0,
    'remaining_capacity': 0,
    'temperature': 0,
    'current': 0,
    'voltage': 0,
    'avg_current': 0,
    'charging_current': 0,
    'full_charge_capacity': 103,
    'charging_voltage': 0,
    'at_rate_time_to_full': 0,
    'at_rate_time_to_empty': 0,
    'at_rate_ok_text': "",
    'at_rate': 0,
    'charging_battery_status': "Off",
    'rel_state_of_charge': 0,
    'abs_state_of_charge': 0,
    'run_time_to_empty': 0,
    'avg_time_to_empty': 0,
    'avg_time_to_full': 0,
    'max_error': 0
}


# Asynchronous function to update device data concurrently for multiple data points
async def update_device_data():
    try:
        # List of tuples where the first element is the call name, and the second element is the key for device data
        data_points = [
            ('serial_number', 'serial_number'),
            ('design_capacity', 'design_capacity'),
            ('design_voltage', 'design_voltage'),
            ('remaining_capacity', 'remaining_capacity'),
            ('temperature', 'temperature'),
            ('current', 'current'),
            ('voltage', 'voltage'),
            ('battery_status', 'battery_status'),
            ('avg_current', 'avg_current'),
            ('full_charge_capacity', 'full_charge_capacity'),
            ('charging_voltage', 'charging_voltage'),
            ('at_rate_time_to_full', 'at_rate_time_to_full'),
            ('at_rate_time_to_empty', 'at_rate_time_to_empty'),
            ('at_rate_ok_text', 'at_rate_ok_text'),
            ('at_rate', 'at_rate'),
            ('rel_state_of_charge', 'rel_state_of_charge'),
            ('abs_state_of_charge', 'abs_state_of_charge'),
            ('run_time_to_empty', 'run_time_to_empty'),
            ('avg_time_to_empty', 'avg_time_to_empty'),
            ('avg_time_to_full', 'avg_time_to_full'),
            ('max_error', 'max_error')
        ]
        
        # Create a list of tasks to fetch and store data concurrently for all data points
        tasks = []
        for call_name, key in data_points:
            task = asyncio.create_task(fetch_and_store_data(call_name, key))  # Create async task for each data point
            tasks.append(task)
        
        # Await completion of all tasks
        await asyncio.gather(*tasks)   
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


# Function to fetch and store data asynchronously
async def fetch_and_store_data(call_name, key):
    try:
        # If battery 2 is initialized (serial number not 0), fetch data for both batteries
        if device_data_battery_2['serial_number'] != 0:
            await asyncio.to_thread(pcan_write_read, call_name, 1)  # Fetch data from battery 1
            await asyncio.to_thread(pcan_write_read, call_name, 2)  # Fetch data from battery 2
        else:
            await asyncio.to_thread(pcan_write_read, call_name, 1)  # Only fetch data from battery 1 
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


# Function to initialize the PCAN device with specified parameters
def pcan_initialize(baudrate, hwtype, ioport, interrupt):
    try:
        # Initialize the PCAN device with the given parameters (baudrate, hardware type, etc.)
        result = m_objPCANBasic.Initialize(m_PcanHandle, baudrate, hwtype, ioport, interrupt)
        
        # If initialization fails, log data and display error message
        if result != PCAN_ERROR_OK:
            log_can_data(device_data_battery_1)  # Log data for battery 1
            if result == 5120:
                result = 512  # Adjust error code if necessary
            messagebox.showerror("Error!", GetFormatedError(result))  # Show formatted error message
            return True  # Indicate failure
        
        # If initialization succeeds, fetch data for both batteries if available
        else:
            pcan_write_read('serial_number', 1)  # Read serial number for battery 1
            pcan_write_read('serial_number', 2)  # Read serial number for battery 2
            
            # If both batteries are connected (serial numbers are non-zero), fetch and log data for both
            if device_data_battery_2['serial_number'] != 0 and device_data_battery_1['serial_number'] != 0:
                pcan_write_read('temperature', 2)
                pcan_write_read('firmware_version', 2)       
                pcan_write_read('voltage', 2)            
                pcan_write_read('battery_status', 2)       
                pcan_write_read('current', 2)     
                pcan_write_read('remaining_capacity', 2)       
                pcan_write_read('full_charge_capacity', 2)
                log_can_data(device_data_battery_2)  # Log data for battery 2
                pcan_write_read('temperature', 1)
                pcan_write_read('firmware_version', 1)       
                pcan_write_read('voltage', 1)            
                pcan_write_read('battery_status', 1)       
                pcan_write_read('current', 1)      
                pcan_write_read('remaining_capacity', 1)       
                pcan_write_read('full_charge_capacity', 1)
                log_can_data(device_data_battery_1)  # Log data for battery 1
                # messagebox.showinfo("Info!", "Device Connected with 2 batteries")  # Display info message
                return True  # Indicate success
            
            # If only battery 1 is connected, fetch and log data for battery 1
            elif device_data_battery_1['serial_number'] != 0:
                pcan_write_read('temperature', 1)
                pcan_write_read('firmware_version', 1)       
                pcan_write_read('voltage', 1)            
                pcan_write_read('battery_status', 1)       
                pcan_write_read('current', 1)      
                pcan_write_read('remaining_capacity', 1)       
                pcan_write_read('full_charge_capacity', 1)
                log_can_data(device_data_battery_1)  # Log data for battery 1
                # messagebox.showinfo("Info!", "Device Connected with 1 battery")  # Display info message
                return True  # Indicate success
            
            # If no batteries are connected, show an error message and uninitialize the PCAN device
            else:
                messagebox.showinfo("Error!", "Connection Failed! Wait for 10 seconds after disconnecting")
                m_objPCANBasic.Uninitialize(m_PcanHandle)  # Uninitialize the PCAN device
                return False  # Indicate failure   
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
        return False


# PCAN Uninitialize API Call with exception handling
def pcan_uninitialize():
    try:
        result = m_objPCANBasic.Uninitialize(m_PcanHandle)  # Uninitialize the PCAN handle
        if result != PCAN_ERROR_OK:
            # If uninitialization fails, reset device data and battery flags to defaults
            update_device_data_to_default()
            update_battery_status_flags_to_default()
            # Show error message with the formatted error code
            messagebox.showerror("Error!", GetFormatedError(result))
            return False  # Return failure status
        else:
            # If uninitialization is successful, reset device data and battery flags
            update_device_data_to_default()
            update_battery_status_flags_to_default()
            # Show info message for successful disconnection
            messagebox.showinfo("Info!", "Connection Disconnect!")
            return True  # Return success status
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
        return False  # Return failure status


# PCAN Write API Call with added exception handling and comments for clarity
def pcan_write_read(call_name, battery_no):
    try:
        CANMsg = TPCANMsg()  # Create a new CAN message instance
        
        # Set the CAN message ID based on the battery number
        if battery_no == 1:
            CANMsg.ID = int('18EFC0D0', 16)  # Battery 1
        elif battery_no == 2:
            CANMsg.ID = int('18EFC1D0', 16)  # Battery 2

        # Set the message length and message type
        CANMsg.LEN = int(8)  # Length of the CAN message
        CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED  # Set the message type to extended
        
        # Setting data byte 0 based on the call_name (for different types of queries)
        if call_name == 'firmware_version':
            CANMsg.DATA[0] = int('00', 16)  # Firmware version query
        else:
            CANMsg.DATA[0] = int('01', 16)  # Other queries

        CANMsg.DATA[1] = int('00', 16)  # Fixed value for byte 1
        
        # Set data byte 2 based on the call_name (for different types of queries)
        if call_name == 'serial_number':
            CANMsg.DATA[2] = int('1c', 16)
        elif call_name == 'at_rate':
            CANMsg.DATA[2] = int('04', 16)
        elif call_name == 'at_rate_time_to_full':
            CANMsg.DATA[2] = int('05', 16)
        elif call_name == 'at_rate_time_to_empty':
            CANMsg.DATA[2] = int('06', 16)
        elif call_name == 'at_rate_ok_text':
            CANMsg.DATA[2] = int('07', 16)
        elif call_name == 'temperature':
            CANMsg.DATA[2] = int('08', 16)
        elif call_name == 'voltage':
            CANMsg.DATA[2] = int('09', 16)
        elif call_name == 'current':
            CANMsg.DATA[2] = int('0a', 16)
        elif call_name == 'avg_current':
            CANMsg.DATA[2] = int('0b', 16)
        elif call_name == 'max_error':
            CANMsg.DATA[2] = int('0c', 16)
        elif call_name == 'rel_state_of_charge':
            CANMsg.DATA[2] = int('0d', 16)
        elif call_name == 'abs_state_of_charge':
            CANMsg.DATA[2] = int('0e', 16)
        elif call_name == 'remaining_capacity':
            CANMsg.DATA[2] = int('0f', 16)
        elif call_name == 'full_charge_capacity':
            CANMsg.DATA[2] = int('10', 16)
        elif call_name == 'run_time_to_empty':
            CANMsg.DATA[2] = int('11', 16)
        elif call_name == 'avg_time_to_empty':
            CANMsg.DATA[2] = int('12', 16)
        elif call_name == 'avg_time_to_full':
            CANMsg.DATA[2] = int('13', 16)
        elif call_name == 'charging_current':
            CANMsg.DATA[2] = int('14', 16)
        elif call_name == 'charging_voltage':
            CANMsg.DATA[2] = int('15', 16)
        elif call_name == 'battery_status':
            CANMsg.DATA[2] = int('16', 16)
        elif call_name == 'cycle_count':
            CANMsg.DATA[2] = int('17', 16)
        elif call_name == 'design_capacity':
            CANMsg.DATA[2] = int('18', 16)
        elif call_name == 'design_voltage':
            CANMsg.DATA[2] = int('19', 16)
        elif call_name == 'manufacturer_date':
            CANMsg.DATA[2] = int('1b', 16)
        elif call_name == 'manufacturer_name':
            CANMsg.DATA[2] = int('20', 16)
        elif call_name == 'device_name':
            CANMsg.DATA[2] = int('21', 16)
        elif call_name == 'firmware_version':
            CANMsg.DATA[2] = int('00', 16)
        else:
            messagebox.showinfo("Error!", "Write operation not found!")  # Handle invalid call_name
        
        # Set data byte 3 based on the call_name (manufacturer_name and device_name get a special value)
        if call_name == 'manufacturer_name' or call_name == 'device_name':
            CANMsg.DATA[3] = int('02', 16)
        else:
            CANMsg.DATA[3] = int('01', 16)  # Default value for byte 3
        
        # Set other fixed bytes 4 to 6
        CANMsg.DATA[4] = int('08', 16)
        CANMsg.DATA[5] = int('02', 16)
        CANMsg.DATA[6] = int('00', 16)
        
        # Set data byte 7 based on the battery number and call_name
        if battery_no == 1:
            # Battery 1-specific call names
            if call_name == 'serial_number':
                CANMsg.DATA[7] = int('01',16)
            elif call_name == 'at_rate':
                CANMsg.DATA[7] = int('02',16)
            elif call_name == 'at_rate_time_to_full':
                CANMsg.DATA[7] = int('03',16)
            elif call_name == 'at_rate_time_to_empty':
                CANMsg.DATA[7] = int('04',16)
            elif call_name == 'at_rate_ok_text':
                CANMsg.DATA[7] = int('05',16)
            elif call_name == 'temperature':
                CANMsg.DATA[7] = int('06',16)
            elif call_name == 'voltage':
                CANMsg.DATA[7] = int('07',16)
            elif call_name == 'current':
                CANMsg.DATA[7] = int('08',16)
            elif call_name == 'avg_current':
                CANMsg.DATA[7] = int('09',16)
            elif call_name == 'max_error':
                CANMsg.DATA[7] = int('0A',16)
            elif call_name == 'rel_state_of_charge':
                CANMsg.DATA[7] = int('0B',16)
            elif call_name == 'abs_state_of_charge':
                CANMsg.DATA[7] = int('0C',16)
            elif call_name == 'remaining_capacity':
                CANMsg.DATA[7] = int('0D',16)
            elif call_name == 'full_charge_capacity':
                CANMsg.DATA[7] = int('0E',16)
            elif call_name == 'run_time_to_empty':
                CANMsg.DATA[7] = int('0F',16)
            elif call_name == 'avg_time_to_empty':
                CANMsg.DATA[7] = int('10',16)
            elif call_name == 'avg_time_to_full':
                CANMsg.DATA[7] = int('11',16)
            elif call_name == 'charging_current':
                CANMsg.DATA[7] = int('12',16)
            elif call_name == 'charging_voltage':
                CANMsg.DATA[7] = int('13',16)
            elif call_name == 'battery_status':
                CANMsg.DATA[7] = int('14',16)
            elif call_name == 'design_capacity':
                CANMsg.DATA[7] = int('16',16)
            elif call_name == 'design_voltage':
                CANMsg.DATA[7] = int('17',16)
            elif call_name == 'firmware_version':
                CANMsg.DATA[7] = int('1B',16)
        elif battery_no == 2:
            # Battery 2-specific call names
            if call_name == 'serial_number':
                CANMsg.DATA[7] = int('1C',16)
            elif call_name == 'at_rate':
                CANMsg.DATA[7] = int('1D',16)
            elif call_name == 'at_rate_time_to_full':
                CANMsg.DATA[7] = int('1E',16)
            elif call_name == 'at_rate_time_to_empty':
                CANMsg.DATA[7] = int('1F',16)
            elif call_name == 'at_rate_ok_text':
                CANMsg.DATA[7] = int('20',16)
            elif call_name == 'temperature':
                CANMsg.DATA[7] = int('21',16)
            elif call_name == 'voltage':
                CANMsg.DATA[7] = int('22',16)
            elif call_name == 'current':
                CANMsg.DATA[7] = int('23',16)
            elif call_name == 'avg_current':
                CANMsg.DATA[7] = int('24',16)
            elif call_name == 'max_error':
                CANMsg.DATA[7] = int('25',16)
            elif call_name == 'rel_state_of_charge':
                CANMsg.DATA[7] = int('26',16)
            elif call_name == 'abs_state_of_charge':
                CANMsg.DATA[7] = int('27',16)
            elif call_name == 'remaining_capacity':
                CANMsg.DATA[7] = int('28',16)
            elif call_name == 'full_charge_capacity':
                CANMsg.DATA[7] = int('29',16)
            elif call_name == 'run_time_to_empty':
                CANMsg.DATA[7] = int('2A',16)
            elif call_name == 'avg_time_to_empty':
                CANMsg.DATA[7] = int('2B',16)
            elif call_name == 'avg_time_to_full':
                CANMsg.DATA[7] = int('2C',16)
            elif call_name == 'charging_current':
                CANMsg.DATA[7] = int('2D',16)
            elif call_name == 'charging_voltage':
                CANMsg.DATA[7] = int('2E',16)
            elif call_name == 'battery_status':
                CANMsg.DATA[7] = int('2F',16)
            elif call_name == 'design_capacity':
                CANMsg.DATA[7] = int('30',16)
            elif call_name == 'design_voltage':
                CANMsg.DATA[7] = int('31',16)
            elif call_name == 'firmware_version':
                CANMsg.DATA[7] = int('32',16)

        
        # Write the message to the PCAN bus
        result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
        
        # Log the result of the write operation
        logger.info(f"{call_name}: {result}")
        
        # Check if the result is an error and display an error message if necessary
        if result != PCAN_ERROR_OK:
            messagebox.showerror(f"Error! {call_name}", GetFormatedError(result))  # Show error message
            return -2  # Return failure status
        
        time.sleep(0.1)  # Small delay before reading the response
        
        # Call pcan_read() to read the response from the bus
        result_code = pcan_read()
        return result_code  # Return the result code from the read operation   
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
        return -1  # Indicate failure if an exception occurs


# PCAN Read API Call with added comments and exception handling
def pcan_read():
    try:
        # Read data from the PCAN bus
        result = m_objPCANBasic.Read(m_PcanHandle)
        
        # Check if the result is OK or if there was an error
        if result[0] != PCAN_ERROR_OK:
            # Return error code if there was an issue
            return result[0]
        else:
            # Extract the message from the result
            args = result[1:]  # Skipping the first element (error code)
            theMsg = args[0]  # Extract the first message in the result
            
            # Create a new CAN FD message to hold the data
            newMsg = TPCANMsgFD()
            newMsg.ID = theMsg.ID  # Set the message ID
            newMsg.DLC = theMsg.LEN  # Set the message length
            
            # Copy the data from the original message to the new message
            for i in range(8 if theMsg.LEN > 8 else theMsg.LEN):
                newMsg.DATA[i] = theMsg.DATA[i]
            
            # Convert the data to hexadecimal representation for easier viewing
            resulthex = [hex(theMsg.DATA[i]) for i in range(8 if theMsg.LEN > 8 else theMsg.LEN)]
            
            # Extract the first and second bytes from the data
            first_byte = newMsg.DATA[0]
            second_byte = newMsg.DATA[1]
            
            # Swap the bytes to create the correct hex value
            swapped_hex = (second_byte << 8) | first_byte
            # Convert the swapped hex value to a decimal integer
            decimal_value = int(swapped_hex)
            
            # Call the convert_data function with the new message and decimal value
            convert_data(newMsg, theMsg, decimal_value)

            # Return the error code from the read operation
            return result[0] 
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
        return -1  # Indicate failure if an exception occurs


def convert_data(newMsg, theMsg, decimal_value):
    try:
        # Conversion rules
        if newMsg.DATA[4] == 0x04:  # AtRate: mA / 40 unsigned
            if newMsg.DATA[7] == 0x02:
                device_data_battery_1['at_rate'] = (decimal_value*40)/1000
            elif newMsg.DATA[7] == 0x1D:
                device_data_battery_2['at_rate'] = (decimal_value*40)/1000
            else:
                logger.info("AtRate Not Found")
        elif newMsg.DATA[4] == 0x05:  # AtRateTimeToFull: minutes unsigned
            if newMsg.DATA[7] == 0x03:
                device_data_battery_1['at_rate_time_to_full'] = round((decimal_value / 1000),1)
            elif newMsg.DATA[7] == 0x1E:
                device_data_battery_2['at_rate_time_to_full'] = round((decimal_value / 1000),1)
            else:
                logger.info("AtRateTimeToFull Not Found")
        elif newMsg.DATA[4] == 0x06:  # AtRateTimeToEmpty: minutes unsigned
            if newMsg.DATA[7] == 0x04:
                device_data_battery_1['at_rate_time_to_empty'] = round((decimal_value / 1000),1)
            elif newMsg.DATA[7] == 0x1F:
                device_data_battery_2['at_rate_time_to_empty'] = round((decimal_value / 1000),1)
            else:
                logger.info("AtRateTimeToEmpty Not Found")
        elif newMsg.DATA[4] == 0x07:  # AtRateOK: Boolean
            if newMsg.DATA[7] == 0x05:
                device_data_battery_1['at_rate_ok_text'] = "Yes" if decimal_value != 0 else "No" 
            elif newMsg.DATA[7] == 0x20:
                device_data_battery_2['at_rate_ok_text'] = "Yes" if decimal_value != 0 else "No" 
            else:
                logger.info("AtRateOK Not Found")
        elif newMsg.DATA[4] == 0x08: # Temperature: Boolean
            temperature_k = decimal_value / 10.0
            temperature_c = temperature_k - 273.15
            if newMsg.DATA[7] == 0x06:
                device_data_battery_1['temperature'] = round(temperature_c,1) 
            elif newMsg.DATA[7] == 0x21:
                device_data_battery_2['temperature'] = round(temperature_c,1)
            else:
                logger.info("Temperature Not Found")
        elif newMsg.DATA[4] == 0x09:  # Voltage: mV unsigned
            if newMsg.DATA[7] == 0x07:
                device_data_battery_1['voltage'] = round((decimal_value / 1000),1)
            elif newMsg.DATA[7] == 0x22:
                device_data_battery_2['voltage'] = round((decimal_value / 1000),1)
            else:
                logger.info("Voltage Not Found")
        elif newMsg.DATA[4] == 0x0a:  # Current: mA / 40 signed
            if decimal_value == 0:
                if newMsg.DATA[7] == 0x08:
                    device_data_battery_1['current'] = decimal_value
                elif newMsg.DATA[7] == 0x23:
                    device_data_battery_2['current'] = decimal_value
                else:
                    logger.info("Current Not Found")
            else:
                if decimal_value > 32767:
                    decimal_value -= 65536
                currentmA = decimal_value*40
                currentA = currentmA / 1000
                if currentA > 1:
                    if newMsg.DATA[7] == 0x08:
                        device_data_battery_1['charging_battery_status'] = "Charging"
                        device_data_battery_1['current'] = 0
                        device_data_battery_1['charging_current'] = abs(currentA)
                    elif newMsg.DATA[7] == 0x23:
                        device_data_battery_2['charging_battery_status'] = "Charging"
                        device_data_battery_2['current'] = 0
                        device_data_battery_2['charging_current'] = abs(currentA)
                    else:
                        logger.info("Current Not Found")
                elif currentA < 1:
                    if newMsg.DATA[7] == 0x08:
                        device_data_battery_1['charging_battery_status'] = "Discharging"
                        device_data_battery_1['current'] = abs(currentA)
                        device_data_battery_1['charging_current'] = 0
                    elif newMsg.DATA[7] == 0x23:
                        device_data_battery_2['charging_battery_status'] = "Discharging"
                        device_data_battery_2['current'] = abs(currentA)
                        device_data_battery_2['charging_current'] = 0
                    else:
                        logger.info("Current Not Found")
                elif currentA == 0:
                    if newMsg.DATA[7] == 0x08:
                        device_data_battery_1['charging_battery_status'] = "Off"
                        device_data_battery_1['current'] = abs(currentA)
                        device_data_battery_1['charging_current'] = 0
                    elif newMsg.DATA[7] == 0x23:
                        device_data_battery_2['charging_battery_status'] = "Off"
                        device_data_battery_2['current'] = abs(currentA)
                        device_data_battery_2['charging_current'] = 0
                    else:
                        logger.info("Current Not Found")
                else:
                    if newMsg.DATA[7] == 0x08:
                        device_data_battery_1['charging_battery_status'] = "Off"
                        device_data_battery_1['current'] = 0
                        device_data_battery_1['charging_current'] = 0
                    elif newMsg.DATA[7] == 0x23:
                        device_data_battery_2['current'] = decimal_value
                        device_data_battery_2['current'] = decimal_value
                        device_data_battery_2['current'] = decimal_value
                    else:
                        logger.info("Current Not Found")
        elif newMsg.DATA[4] == 0x0b:  # Avg Current: mA / 40 signed
            if decimal_value == 0:
                if newMsg.DATA[7] == 0x09:
                    device_data_battery_1['avg_current'] = decimal_value
                elif newMsg.DATA[7] == 0x24:
                    device_data_battery_2['avg_current'] = decimal_value
                else:
                    logger.info("Avg Current Not Found")
            else:
                if decimal_value > 32767:
                    decimal_value -= 65536
                currentmA = decimal_value*40
                currentA = currentmA / 1000
                if newMsg.DATA[7] == 0x09:
                    device_data_battery_1['avg_current'] = currentA
                elif newMsg.DATA[7] == 0x24:
                    device_data_battery_2['avg_current'] = currentA
                else:
                    logger.info("Avg Current Not Found")
        elif newMsg.DATA[4] == 0x0c:  # MaxError: Percent unsigned
            if newMsg.DATA[7] == 0x0A:
                device_data_battery_1['max_error'] = decimal_value
            elif newMsg.DATA[7] == 0x25:
                device_data_battery_2['max_error'] = decimal_value
            else:
                logger.info("MaxError Not Found")
        elif newMsg.DATA[4] == 0x0d:  # RelStateofCharge: Percent unsigned
            if newMsg.DATA[7] == 0x0B:
                device_data_battery_1['rel_state_of_charge'] = decimal_value
            elif newMsg.DATA[7] == 0x26:
                device_data_battery_2['rel_state_of_charge'] = decimal_value
            else:
                logger.info("RelStateofCharge Not Found")
        elif newMsg.DATA[4] == 0x0e:  # AbsoluteStateofCharge: Percent unsigned
            if newMsg.DATA[7] == 0x0C:
                device_data_battery_1['abs_state_of_charge'] = decimal_value
            elif newMsg.DATA[7] == 0x27:
                device_data_battery_2['abs_state_of_charge'] = decimal_value
            else:
                logger.info("AbsoluteStateofCharge Not Found")
        elif newMsg.DATA[4] == 0x0f:  # RemainingCapacity: mAh / 40 unsigned
            if newMsg.DATA[7] == 0x0D:
                device_data_battery_1['remaining_capacity'] = (decimal_value*40)/1000
            elif newMsg.DATA[7] == 0x28:
                device_data_battery_2['remaining_capacity'] = (decimal_value*40)/1000
            else:
                logger.info("MaxError Not Found")
        elif newMsg.DATA[4] == 0x10:  # FullChargeCapacity: mAh / 40 unsigned
            if newMsg.DATA[7] == 0x0E:
                device_data_battery_1['full_charge_capacity'] = (decimal_value*40)/1000
            elif newMsg.DATA[7] == 0x29:
                device_data_battery_2['full_charge_capacity'] = (decimal_value*40)/1000
            else:
                logger.info("FullChargeCapacity Not Found")
        elif newMsg.DATA[4] == 0x11:  # RunTimeToEmpty: minutes unsigned
            if newMsg.DATA[7] == 0x0F:
                device_data_battery_1['run_time_to_empty'] = round((decimal_value / 10),1)
            elif newMsg.DATA[7] == 0x2A:
                device_data_battery_2['run_time_to_empty'] = round((decimal_value / 10),1)
            else:
                logger.info("RunTimeToEmpty Not Found")
        elif newMsg.DATA[4] == 0x12:  # AvgTimeToEmpty: minutes unsigned
            if newMsg.DATA[7] == 0x10:
                device_data_battery_1['avg_time_to_empty'] = round((decimal_value / 10),1)
            elif newMsg.DATA[7] == 0x2B:
                device_data_battery_2['avg_time_to_empty'] = round((decimal_value / 10),1)
            else:
                logger.info("AvgTimeToEmpty Not Found")
            device_data['avg_time_to_empty'] = round((decimal_value / 10),1)
        elif newMsg.DATA[4] == 0x13:  # AvgTimeToFull: minutes unsigned
            if newMsg.DATA[7] == 0x11:
                device_data_battery_1['avg_time_to_full'] = round((decimal_value / 1000),1)
            elif newMsg.DATA[7] == 0x2C:
                device_data_battery_2['avg_time_to_full'] = round((decimal_value / 1000),1)
            else:
                logger.info("AvgTimeToFull Not Found")
        elif newMsg.DATA[4] == 0x15:  # ChargingVoltage: mV unsigned
            if newMsg.DATA[7] == 0x13:
                device_data_battery_1['charging_voltage'] = round((decimal_value / 1000),1)
                logger.info(f"Charging Voltage {device_data_battery_1['charging_voltage']}")
            elif newMsg.DATA[7] == 0x2E:
                device_data_battery_2['charging_voltage'] = round((decimal_value / 1000),1)
            else:
                logger.info("ChargingVoltage Not Found")
        elif newMsg.DATA[4] == 0x16:  # BatteryStatus: bit flags unsigned
            binary_value = format(decimal_value, '016b')  # Convert to 16-bit binary string

            # Swap the bytes before interpreting the bits
            swapped_value = binary_value[8:] + binary_value[:8]
            if newMsg.DATA[7] == 0x14:
                device_data_battery_1['battery_status'] = swapped_value
                battery_1_status_flags.update({
                        "overcharged_alarm": int(swapped_value[0]),  # Bit 15
                        "terminate_charge_alarm": int(swapped_value[1]),  # Bit 14
                        "over_temperature_alarm": int(swapped_value[3]),  # Bit 12
                        "terminate_discharge_alarm": int(swapped_value[4]),  # Bit 11
                        "remaining_capacity_alarm": int(swapped_value[6]),  # Bit 9
                        "remaining_time_alarm": int(swapped_value[7]),  # Bit 8
                        "initialization": int(swapped_value[8]),  # Bit 7
                        "charge_fet_test": int(swapped_value[9]),  # Bit 6
                        "fully_charged": int(swapped_value[10]),  # Bit 5
                        "fully_discharged": int(swapped_value[11]),  # Bit 4
                        "error_codes": int(swapped_value[12:], 2)  # Bits 3:0, convert remaining bits to an integer
                    })
            elif newMsg.DATA[7] == 0x2F:
                device_data_battery_2['battery_status'] = swapped_value
                battery_2_status_flags.update({
                        "overcharged_alarm": int(swapped_value[0]),  # Bit 15
                        "terminate_charge_alarm": int(swapped_value[1]),  # Bit 14
                        "over_temperature_alarm": int(swapped_value[3]),  # Bit 12
                        "terminate_discharge_alarm": int(swapped_value[4]),  # Bit 11
                        "remaining_capacity_alarm": int(swapped_value[6]),  # Bit 9
                        "remaining_time_alarm": int(swapped_value[7]),  # Bit 8
                        "initialization": int(swapped_value[8]),  # Bit 7
                        "charge_fet_test": int(swapped_value[9]),  # Bit 6
                        "fully_charged": int(swapped_value[10]),  # Bit 5
                        "fully_discharged": int(swapped_value[11]),  # Bit 4
                        "error_codes": int(swapped_value[12:], 2)  # Bits 3:0, convert remaining bits to an integer
                    })
            else:
                logger.info("Battery Status Not Found")          
        elif newMsg.DATA[4] == 0x18:  # DesignCapacity: mAh / 40 unsigned
            if newMsg.DATA[7] == 0x16:
                device_data_battery_1['design_capacity'] = (decimal_value*40)/1000
            elif newMsg.DATA[7] == 0x2F:
                device_data_battery_2['design_capacity'] = (decimal_value*40)/1000
            else:
                logger.info("Design Capacity Not Found")
            device_data['design_capacity'] = (decimal_value*40)/1000
        elif newMsg.DATA[4] == 0x19:  # DesignVoltage: mV unsigned
            if newMsg.DATA[7] == 0x17:
                device_data_battery_1['design_voltage'] = round((decimal_value / 1000),1)
            elif newMsg.DATA[7] == 0x31:
                device_data_battery_2['design_voltage'] = round((decimal_value / 1000),1)
            else:
                logger.info("Design Voltage Not Found")
        elif newMsg.DATA[4] == 0x1c:  # SerialNumber: number unsigned
            if newMsg.DATA[7] == 0x01:
                device_data_battery_1['serial_number'] = decimal_value
            elif newMsg.DATA[7] == 0x1C:
                device_data_battery_2['serial_number'] = decimal_value
            else:
                logger.info("Serial Number Not Found")
        else:
            if newMsg.DATA[7] == 0x1B:
                data_packet = [int(hex(newMsg.DATA[i]), 16) for i in range(8 if (theMsg.LEN > 8) else theMsg.LEN)]
                major_version = data_packet[0]
                minor_version = data_packet[1]
                patch_number = (data_packet[3] << 8) | data_packet[2]
                build_number = (data_packet[5] << 8) | data_packet[4]
                version_string = f"{major_version}.{minor_version}.{patch_number}.{build_number}"
                device_data_battery_1['firmware_version'] = version_string
            elif newMsg.DATA[7] == 0x32:
                data_packet = [int(hex(newMsg.DATA[i]), 16) for i in range(8 if (theMsg.LEN > 8) else theMsg.LEN)]
                major_version = data_packet[0]
                minor_version = data_packet[1]
                patch_number = (data_packet[3] << 8) | data_packet[2]
                build_number = (data_packet[5] << 8) | data_packet[4]
                version_string = f"{major_version}.{minor_version}.{patch_number}.{build_number}"
                device_data_battery_2['firmware_version'] = version_string
            else:
                logger.info("Firmware version is not found")
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


def GetFormatedError(error):
        # Gets the text using the GetErrorText API function
        # If the function success, the translated error is returned. If it fails,
        # a text describing the current error is returned.
        #
        try:
            stsReturn = m_objPCANBasic.GetErrorText(error, 0)
            if stsReturn[0] != PCAN_ERROR_OK:
                return "An error occurred. Error-code's text ({0:X}h) couldn't be retrieved".format(error)
            else:
                return stsReturn[1]
        except Exception as e:
            error_details = traceback.format_exc()
            logger.error(f"An unexpected error occurred: {e}\n{error_details}")
            messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


# PCAN Write Control API Call with added comments and exception handling
def pcan_write_control(call_name, battery_no):
    try:
        # Create a new CAN message
        CANMsg = TPCANMsg()
        
        # Set the CAN message ID based on the battery number
        if battery_no == 1:
            CANMsg.ID = int('18EFC0D0', 16)
        elif battery_no == 2:
            CANMsg.ID = int('18EFC1D0', 16)

        # Set the message length and type
        CANMsg.LEN = int(8)
        CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED
        
        # Define the data to be sent in the CAN message (first step of the control process)
        CANMsg.DATA[0] = int('03', 16)
        CANMsg.DATA[1] = int('03', 16)
        CANMsg.DATA[2] = int('37', 16)
        CANMsg.DATA[3] = int('30', 16)
        CANMsg.DATA[4] = int('39', 16)
        CANMsg.DATA[5] = int('33', 16)
        CANMsg.DATA[6] = int('39', 16)
        CANMsg.DATA[7] = int('00', 16)
        
        # Write the CAN message to the bus
        result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
        if result == PCAN_ERROR_OK:
            logger.info("Enter Diag State : Success")
            
            # Prepare a second CAN message based on the 'call_name' parameter
            CANMsg = TPCANMsg()
            if battery_no == 1:
                CANMsg.ID = int('18EFC0D0', 16)
            elif battery_no == 2:
                CANMsg.ID = int('18EFC1D0', 16)
            
            # Set the length and message type for the new CAN message
            CANMsg.LEN = int(8)
            CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED
            CANMsg.DATA[0] = int('03', 16)

            # Set the control state based on the 'call_name' parameter
            if call_name == 'both_off':
                CANMsg.DATA[1] = int('01', 16)
            elif call_name == 'heater_on':
                CANMsg.DATA[1] = int('05', 16)
            else:
                CANMsg.DATA[1] = int('00', 16)

            # Set the second byte based on 'call_name'
            if call_name == 'both_off':
                CANMsg.DATA[2] = int('00', 16)
            elif call_name == 'charge_on':
                CANMsg.DATA[2] = int('01', 16)
            elif call_name == 'discharge_on':
                CANMsg.DATA[2] = int('02', 16)
            elif call_name == 'both_on':
                CANMsg.DATA[2] = int('03', 16)
            elif call_name == 'bms_reset':
                CANMsg.DATA[2] = int('01', 16)
            elif call_name == 'heater_on':
                CANMsg.DATA[2] = int('00', 16)
            else:
                messagebox.showinfo("Error!", "Write operation not found!")
            
            # Set the remaining bytes in the message
            CANMsg.DATA[3] = int('01', 16)
            CANMsg.DATA[4] = int('08', 16)
            CANMsg.DATA[5] = int('02', 16)
            CANMsg.DATA[6] = int('00', 16)
            CANMsg.DATA[7] = int('00', 16)

            # Write the FET control state to the bus
            result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
            if result == PCAN_ERROR_OK:
                logger.info("FET Control State : Success")
                
                # Prepare a final CAN message to exit the diagnostic state
                CANMsg = TPCANMsg()
                if battery_no == 1:
                    CANMsg.ID = int('18EFC0D0', 16)
                elif battery_no == 2:
                    CANMsg.ID = int('18EFC1D0', 16)
                
                # Set the message parameters for exit state
                CANMsg.LEN = int(8)
                CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED
                CANMsg.DATA[0] = int('03', 16)
                CANMsg.DATA[1] = int('04', 16)
                CANMsg.DATA[2] = int('00', 16)
                CANMsg.DATA[3] = int('00', 16)
                CANMsg.DATA[4] = int('00', 16)
                CANMsg.DATA[5] = int('00', 16)
                CANMsg.DATA[6] = int('00', 16)
                CANMsg.DATA[7] = int('00', 16)

                # Write the exit message to the bus
                result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
                if result == PCAN_ERROR_OK:
                    logger.info("Exit the Diag State : Success")
                    return 0
                else:
                    logger.info("Exit the Diag State : Failed")
                    return 0
            else:
                logger.info("FET Control State : Failed")
                return 0
        else:
            logger.info("Enter Diag State : Failed")
            return 0         
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
        return -1  # Indicate failure if an exception occurs


# Update device data with default values
def update_device_data_to_default():
    def set_default_data(device_data_dict):
        """Helper function to set default values for a device's data."""
        try:
            for key in device_data_dict.keys():
                if key == 'device_name':
                    device_data_dict[key] = "BT-70939APH"
                elif key == 'manufacturer_name':
                    device_data_dict[key] = "Bren-Tronics"
                elif key == 'serial_number':
                    device_data_dict[key] = 1476
                elif key == 'charging_battery_status':
                    device_data_dict[key] = "Off"
                elif isinstance(device_data_dict[key], str):
                    device_data_dict[key] = ""
                elif isinstance(device_data_dict[key], (int, float)):
                    device_data_dict[key] = 0
        except Exception as e:
            error_details = traceback.format_exc()
            logger.error(f"An unexpected error occurred: {e}\n{error_details}")
            messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
    try:
        # Update both batteries' data
        set_default_data(device_data_battery_1)
        set_default_data(device_data_battery_2)
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


# Update battery status flag with default values
def update_battery_status_flags_to_default():
    """Reset battery status flags to default (0)."""
    try:
        for key in battery_1_status_flags.keys():
            battery_1_status_flags[key] = 0
        for key in battery_2_status_flags.keys():
            battery_2_status_flags[key] = 0
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


# Log CAN data to an Excel file
def log_can_data(update_can_data):
    """
    Log CAN data to an Excel file. If the serial number exists, do nothing;
    if it doesn't exist, append a new row with default values.
    """
    try:
        # Define the path for the CAN data Excel file inside the AppData directory
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Define the log file name
        file_name = "can_data.xlsx"
        file_path = os.path.join(folder_path, file_name)

        # Create or load workbook
        if os.path.exists(file_path):
            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active
            except Exception as e:
                logger.error(f"Error loading the Excel file: {str(e)}")
                messagebox.showerror("Error", f"Failed to load the Excel file: {str(e)}. Recreating the file.")
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "CAN Data"
                # Add headers for CAN data
                headers = [
                    "SI No", "Date", "Time", "Project", "Device Name", "Manufacturer Name", 
                    "Serial Number", "Cycle Count", "Full Charge Capacity", "Charging Date", 
                    "OCV Before Charging", "Discharging Date", "OCV Before Discharging"
                ]
                sheet.append(headers)
        else:
            # File does not exist, so create it
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "CAN Data"
            # Add headers for CAN data
            headers = [
                "SI No", "Date", "Time", "Project", "Device Name", "Manufacturer Name", 
                "Serial Number", "Cycle Count", "Full Charge Capacity", "Charging Date", 
                "OCV Before Charging", "Discharging Date", "OCV Before Discharging"
            ]
            sheet.append(headers)

        # Log the current CAN data
        current_datetime = datetime.datetime.now()
        date = current_datetime.strftime("%Y-%m-%d")
        time = current_datetime.strftime("%H:%M:%S")
        serial_number = update_can_data.get('serial_number', 'N/A')

        # Check if the serial number exists in the file
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        serial_found = False

        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                serial_found = True
                logger.info(f"Serial number {serial_number} already exists. No action taken.")
                break

        if not serial_found:
            # If serial number is not found, add a new row with default values
            next_row = sheet.max_row + 1
            can_data = [
                next_row - 1,  # SI No
                date,  # Date
                time,  # Time
                update_can_data.get('project', ''),  # Project
                update_can_data.get('device_name', 'BT-70939APH'),  # Device Name
                update_can_data.get('manufacturer_name', 'Bren-Tronics'),  # Manufacturer Name
                serial_number,  # Serial Number
                int(update_can_data.get('cycle_count', 0)),  # Ensure Cycle Count is an int
                103,  # Ensure Full Charge Capacity is a float
                update_can_data.get('charging_date', date),  # Charging Date
                float(update_can_data.get('ocv_before_charging', 0.0)),  # Ensure OCV Before Charging is a float
                update_can_data.get('discharging_date', date),  # Discharging Date
                float(update_can_data.get('ocv_before_discharging', 0.0))  # Ensure OCV Before Discharging is a float
            ]
            sheet.append(can_data)

        # Save the Excel file
        workbook.save(file_path)
        workbook.close()
        logger.info(f"CAN data logged in {file_path}.")
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")

# Retrieves the most recent CAN data for the specified serial number.
def get_latest_can_data(serial_number):
    """
    Retrieves the most recent CAN data for the specified serial number.
    Returns a dictionary of the latest data if found, otherwise returns an empty dictionary.
    """
    try:
        # Path to the CAN data Excel file (stored in AppData directory)
        can_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "can_data.xlsx")

        # If the CAN data file doesn't exist, show an error message and return an empty dictionary
        if not os.path.exists(can_data_file):
            messagebox.showerror("Error", "No CAN data file found. Please log CAN data first.")
            return {}

        # Open the Excel file in read-only mode for efficiency
        workbook = load_workbook(can_data_file, read_only=True)
        sheet = workbook.active

        # Serial Number is stored in the 7th column (G column)
        serial_number_column = 7

        # Iterate through the rows in the sheet starting from row 2 (to skip header)
        for row in range(2, sheet.max_row + 1):
            current_serial_number = sheet.cell(row=row, column=serial_number_column).value
            if current_serial_number == serial_number:
                # Log the row where the serial number was found
                logger.info(f"Serial Number {serial_number} found in row {row}")

                # Retrieve the header values (first row) and data values (current row)
                headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
                values = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]

                # Create a dictionary by mapping headers to corresponding values
                latest_data = dict(zip(headers, values))
                
                # Debugging: Log the retrieved data
                for header, value in latest_data.items():
                    logger.info(f"{header}: {value}")

                # Return the latest data as a dictionary
                return latest_data

        # If the serial number wasn't found in the sheet, show a warning
        messagebox.showwarning("Warning", f"Serial number {serial_number} not found in the CAN data.")
        return {}
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
        return {}


# Update the cycle count for a specific device identified by its serial number
def update_cycle_count_in_can_data(serial_number, new_cycle_count):
    """
    Update the cycle count for a specific device identified by its serial number
    in the CAN Data Excel file.
    """
    try:
        # Define the folder and file path for the CAN data file
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        file_path = os.path.join(folder_path, "can_data.xlsx")

        # Check if the file exists, otherwise show an error
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "CAN data file not found.")
            return

        # Open the existing Excel file
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Define the columns for serial number and cycle count
        serial_number_column = 7  # Serial Number in column G
        cycle_count_column = 8    # Cycle Count in column H

        # Loop through the rows to find the row with the specified serial number
        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                # Update the Cycle Count in the matching row
                sheet.cell(row=row, column=cycle_count_column).value = new_cycle_count
                break
        else:
            # If the serial number is not found, show a warning
            messagebox.showwarning("Warning", f"Serial number {serial_number} not found.")
            return

        # Save the updated Excel file
        workbook.save(file_path)
        messagebox.showinfo("Success", f"Cycle count updated successfully for Serial Number {serial_number}.")
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


def update_excel_and_download_pdf(data):
    """
    Get the updated values from the form, update the Excel file, and generate a PDF.
    """
    try:
        # Define the path to the CAN data Excel file
        can_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "can_data.xlsx")

        # Load the existing data
        try:
            df_can_data = pd.read_excel(can_data_file)
        except FileNotFoundError:
            messagebox.showerror("Error", f"{can_data_file} not found.")
            return

        # Assuming the serial number is unique, update the corresponding row
        serial_number = int(data[2])  # serial_number is the second element in the array
        df_can_data['Serial Number'] = df_can_data['Serial Number'].astype(int)
        df_can_data['OCV Before Charging'] = df_can_data['OCV Before Charging'].astype(float)
        df_can_data['OCV Before Discharging'] = df_can_data['OCV Before Discharging'].astype(float)
        index = df_can_data[df_can_data['Serial Number'] == serial_number].index
        logger.info(f"{index} index")

        if not index.empty:
            # Update the values in the DataFrame
            df_can_data.loc[index[0], 'Project'] = str(data[0])  # Ensure it's a string
            df_can_data.loc[index[0], 'Device Name'] = str(data[1])  # Ensure it's a string
            df_can_data.loc[index[0], 'Manufacturer Name'] = str(data[3])  # Ensure it's a string
            df_can_data.loc[index[0], 'Serial Number'] = serial_number
            df_can_data.loc[index[0], 'Cycle Count'] = int(data[4])  # Convert to int
            df_can_data.loc[index[0], 'Full Charge Capacity'] = float(data[5])  # Convert to float
            df_can_data.loc[index[0], 'Charging Date'] = str(data[6])  # Ensure it's a string
            df_can_data.loc[index[0], 'OCV Before Charging'] = float(data[7])  # Convert to float
            df_can_data.loc[index[0], 'Discharging Date'] = str(data[8])  # Ensure it's a string
            df_can_data.loc[index[0], 'OCV Before Discharging'] = float(data[9])  # Convert to float

            # Save the updated DataFrame back to the Excel file
            df_can_data.to_excel(can_data_file, index=False)
            messagebox.showinfo("Success", "Data updated successfully.")
        else:
            messagebox.showwarning("Warning", "Serial Number not found in the Excel file.")

        # Generate the PDF with the updated values
        pdf_generator.create_can_report_pdf(serial_number, "CAN")
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


def open_pdf_folder():
    try:
        # Get the folder path where the PDFs are saved
        folder_path = os.path.join(os.path.expanduser("~"), "Documents", "Battery_PDFs")

        # Check if the folder exists, if not, show an error
        if not os.path.exists(folder_path):
            messagebox.showerror("Error", "No PDF files found. Please generate a PDF first.")
            return

        # Open the folder
        try:
            os.startfile(folder_path)  # This will open the folder in Windows Explorer
        except Exception as e:
            error_details = traceback.format_exc()
            logger.error(f"An unexpected error occurred: {e}\n{error_details}")
            messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


def update_charging_ocv_in_excel(serial_number, ocv_value):
    """
    Updates the charging OCV and charging date fields in the CAN data Excel file.
    
    Parameters:
    - serial_number (str): The serial number of the battery.
    - ocv_value (float): The OCV value before charging.
    """
    try:
        # Define the folder and file path for the CAN data log
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        file_path = os.path.join(folder_path, "can_data.xlsx")

        # Load the Excel file
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "CAN data log file not found.")
            return False

        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Loop through rows to find the correct row by Serial Number
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        charging_date_column = 10   # Assuming Charging Date is in column J (10th column)
        charging_ocv_column = 11   # Assuming OCV Before Charging is in column K (11th column)

        serial_found = False
        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                serial_found = True
                # Update the OCV before charging
                sheet.cell(row=row, column=charging_ocv_column).value = ocv_value
                # Set the current date as the charging date
                sheet.cell(row=row, column=charging_date_column).value = datetime.datetime.now().strftime("%Y-%m-%d")
                logger.info(f"Updated Charging OCV and Date for Serial: {serial_number}")
                break

        if not serial_found:
            messagebox.showwarning("Warning", f"Serial number {serial_number} not found.")
            return False

        # Save the changes to the Excel file
        workbook.save(file_path)
        workbook.close()
        return True
    except FileNotFoundError:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", "The Excel file could not be found.")
    except PermissionError:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", "The Excel file is currently open. Please close it and try again.")
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}\nCheck logs for details.")


def update_discharging_ocv_in_excel(serial_number, ocv_value):
    """
    Updates the discharging OCV and discharging date fields in the CAN data Excel file.
    
    Parameters:
    - serial_number (str): The serial number of the battery.
    - ocv_value (float): The OCV value before discharging.
    """
    try:
        # Define the folder and file path for the CAN data log
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        file_path = os.path.join(folder_path, "can_data.xlsx")

        # Load the Excel file
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "CAN data log file not found.")
            return False

        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Loop through rows to find the correct row by Serial Number
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        discharging_date_column = 12  # Assuming Discharging Date is in column M (13th column)
        discharging_ocv_column = 13  # Assuming OCV Before Discharging is in column L (12th column)
        
        serial_found = False
        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                serial_found = True
                # Update the OCV before discharging
                sheet.cell(row=row, column=discharging_ocv_column).value = ocv_value
                # Set the current date as the discharging date
                sheet.cell(row=row, column=discharging_date_column).value = datetime.datetime.now().strftime("%Y-%m-%d")
                logger.info(f"Updated Discharging OCV and Date for Serial: {serial_number}")
                break

        if not serial_found:
            messagebox.showwarning("Warning", f"Serial number {serial_number} not found.")
            return False

        # Save the changes to the Excel file
        workbook.save(file_path)
        workbook.close()
        return True  
    except FileNotFoundError:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", "The Excel file could not be found.")
    except PermissionError:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", "The Excel file is currently open. Please close it and try again.")
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"An unexpected error occurred: {e}\n{error_details}")
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

