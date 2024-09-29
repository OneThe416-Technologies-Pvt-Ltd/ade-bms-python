#pcan_methods.py

from pcan_api.pcan import *
from tkinter import messagebox
import time
import pandas as pd
import asyncio
from openpyxl import Workbook, load_workbook
import datetime
from fpdf import FPDF
import os
import helpers.pdf_generator as pdf_generator  # Config helper
import helpers.config as config  # Config helper

m_objPCANBasic = PCANBasic()
m_PcanHandle = 81
is_fetching_current=False

# Path to store the configuration file (e.g., in the user's home directory)
CONFIG_FILE_PATH = os.path.join(os.path.expanduser("~"), "device_config.json")

battery_status_flags = {
    "overcharged_alarm": 0,  # Bit 15
    "terminate_charge_alarm": 0,  # Bit 14
    "over_temperature_alarm": 0,  # Bit 12
    "terminate_discharge_alarm": 0,  # Bit 11
    "remaining_capacity_alarm": 0,  # Bit 9
    "remaining_time_alarm": 0,  # Bit 8
    "initialization": 0,  # Bit 7
    "charge_fet_test": 1,  # Bit 6
    "fully_charged": 0,  # Bit 5
    "fully_discharged": 0,  # Bit 4
    "error_codes": 0  # Bits 3:0
}

battery_1_status_flags = {
    "overcharged_alarm": 1,  # Bit 15
    "terminate_charge_alarm": 1,  # Bit 14
    "over_temperature_alarm": 1,  # Bit 12
    "terminate_discharge_alarm": 1,  # Bit 11
    "remaining_capacity_alarm":1,  # Bit 9
    "remaining_time_alarm": 1,  # Bit 8
    "initialization": 1,  # Bit 7
    "charge_fet_test": 1,  # Bit 6
    "fully_charged": 0,  # Bit 5
    "fully_discharged": 1,  # Bit 4
    "error_codes": 1  # Bits 3:0
}

battery_2_status_flags = {
    "overcharged_alarm": 0,  # Bit 15
    "terminate_charge_alarm": 0,  # Bit 14
    "over_temperature_alarm": 0,  # Bit 12
    "terminate_discharge_alarm": 0,  # Bit 11
    "remaining_capacity_alarm": 0,  # Bit 9
    "remaining_time_alarm": 0,  # Bit 8
    "initialization": 0,  # Bit 7
    "charge_fet_test": 1,  # Bit 6
    "fully_charged": 0,  # Bit 5
    "fully_discharged": 0,  # Bit 4
    "error_codes": 0  # Bits 3:0
}

name_mapping = {
            "device_name": "Device Name",
            "serial_number": "Serial No",
            "firmware_version": "Firmware Version",
            "manufacturer_date": "Manufacturer Date",
            "manufacturer_name": "Manufacturer Name",
            "battery_status": "Battery Status",
            "cycle_count": "Cycle Count",
            "design_capacity": "Design Capacity",
            "design_voltage": "Design Voltage",
            "at_rate_ok_text": "At Rate OK",
            "at_rate_time_to_full": "At Rate Time To Full",
            "at_rate_time_to_empty": "At Rate Time To Empty",
            "at_rate": "At Rate",
            "rel_state_of_charge": "Rel State of Charge",
            "abs_state_of_charge": "Absolute State of Charge",
            "run_time_to_empty": "Run Time To Empty",
            "avg_time_to_empty": "Avg Time To Empty",
            "avg_time_to_full": "Avg Time To Full",
            "charging_battery_status": "Charging Status",
            "max_error": "Max Error",
            "temperature": "Temperature",
            "current": "Current",
            "remaining_capacity": "Remaining Capacity",
            "voltage": "Voltage",
            "avg_current": "Avg Current",
            "charging_current": "Charging Current",
            "full_charge_capacity": "Full Charge Capacity",
            "charging_voltage": "Charging Voltage"
        }

unit_mapping = {
    "device_name": "string",
    "firmware_version": "string",
    "serial_number": "number",
    "manufacturer_date": "unsigned int",
    "manufacturer_name": "string",
    "battery_status": "bit flags",
    "cycle_count": "Count",
    "manual_cycle_count": "Count",
    "design_capacity": "Ah",
    "design_voltage": "mV",
    "at_rate_ok_text": "Yes/No",
    "at_rate_time_to_full": "minutes",
    "at_rate_time_to_empty": "minutes",
    "at_rate":"A",
    "rel_state_of_charge": "Percent",
    "abs_state_of_charge": "Percent",
    "run_time_to_empty": "minutes",
    "avg_time_to_empty": "minutes",
    "charging_battery_status": "string",
    "avg_time_to_full": "minutes",
    "max_error": "Percent",
    "temperature": "°C",
    "current": "A",
    "remaining_capacity": "Ah",
    "voltage": "mV",
    "avg_current": "A",
    "charging_current": "A",
    "full_charge_capacity": "Ah",
    "charging_voltage": "mV"
}

device_data = {
            'device_name': "BT-70939APH",
            'serial_number': 0,
            'manufacturer_name': "Bren-Tronics",
            'firmware_version': "",
            'battery_status': "",
            'cycle_count': 0,
            'design_capacity': 0,
            'design_voltage': 0,
            'remaining_capacity': 0,
            'temperature': 0,
            'current': 0,
            'voltage': 0,
            'avg_current': 0,
            'charging_current': 0,
            'full_charge_capacity': 0,
            'charging_voltage': 0,
            'at_rate_time_to_full': 0,
            'at_rate_time_to_empty': 0,
            'at_rate_ok_text': "",
            'at_rate':0,
            'charging_battery_status':"Off",
            'rel_state_of_charge': 0,
            'abs_state_of_charge': 0,
            'run_time_to_empty': 0,
            'avg_time_to_empty': 0,
            'avg_time_to_full': 0,
            'max_error': 0
        }

device_data_battery_1 = {
            'device_name': "BT-70939APH",
            'serial_number': 1478,
            'manufacturer_name': "Bren-Tronics",
            'firmware_version': "",
            'battery_status': "",
            'cycle_count': 0,
            'design_capacity': 0,
            'design_voltage': 0,
            'remaining_capacity': 0,
            'temperature': 0,
            'current': 30,
            'voltage': 26,
            'avg_current': 0,
            'charging_current': 0,
            'full_charge_capacity': 0,
            'charging_voltage': 0,
            'at_rate_time_to_full': 0,
            'at_rate_time_to_empty': 0,
            'at_rate_ok_text': "",
            'at_rate':0,
            'charging_battery_status':"Off",
            'rel_state_of_charge': 0,
            'abs_state_of_charge': 0,
            'run_time_to_empty': 0,
            'avg_time_to_empty': 0,
            'avg_time_to_full': 0,
            'max_error': 0
        }

device_data_battery_2 = {
            'device_name': "BT-70939APH",
            'serial_number': 1477,
            'manufacturer_name': "Bren-Tronics",
            'firmware_version': "",
            'battery_status': "",
            'cycle_count': 0,
            'design_capacity': 0,
            'design_voltage': 0,
            'remaining_capacity': 0,
            'temperature': 0,
            'current': 0,
            'voltage': 24,
            'avg_current': 0,
            'charging_current': 0,
            'full_charge_capacity': 0,
            'charging_voltage': 0,
            'at_rate_time_to_full': 0,
            'at_rate_time_to_empty': 0,
            'at_rate_ok_text': "",
            'at_rate':0,
            'charging_battery_status':"Off",
            'rel_state_of_charge': 0,
            'abs_state_of_charge': 0,
            'run_time_to_empty': 0,
            'avg_time_to_empty': 0,
            'avg_time_to_full': 0,
            'max_error': 0
        } 


async def update_device_data():
    data_points = [
        ('serial_number', 'serial_number'),
        ('design_capacity', 'design_capacity'),
        ('design_voltage', 'design_voltage'),
        ('remaining_capacity', 'remaining_capacity'),
        ('temperature', 'temperature'),
        ('current', 'current'),
        ('voltage', 'voltage'),
        ('battery_status', 'battery_status'),
        ('avg_current', 'avg_current'),
        ('full_charge_capacity', 'full_charge_capacity'),
        ('charging_voltage', 'charging_voltage'),
        ('at_rate_time_to_full', 'at_rate_time_to_full'),
        ('at_rate_time_to_empty', 'at_rate_time_to_empty'),
        ('at_rate_ok_text', 'at_rate_ok_text'),
        ('at_rate', 'at_rate'),
        ('rel_state_of_charge', 'rel_state_of_charge'),
        ('abs_state_of_charge', 'abs_state_of_charge'),
        ('run_time_to_empty', 'run_time_to_empty'),
        ('avg_time_to_empty', 'avg_time_to_empty'),
        ('avg_time_to_full', 'avg_time_to_full'),
        ('max_error', 'max_error')
    ]

    # Create a list of tasks to be run concurrently
    tasks = []
    for call_name, key in data_points:
        task = asyncio.create_task(fetch_and_store_data(call_name, key))
        tasks.append(task)

    # Wait for all tasks to complete
    await asyncio.gather(*tasks)


async def fetch_and_store_data(call_name, key):
    if device_data_battery_2['serial_number'] != 0:
        await asyncio.to_thread(pcan_write_read, call_name,1)
        await asyncio.to_thread(pcan_write_read, call_name,2)
    else:
        await asyncio.to_thread(pcan_write_read, call_name,1)


def pcan_initialize(baudrate, hwtype, ioport, interrupt):
    result = m_objPCANBasic.Initialize(m_PcanHandle, baudrate, hwtype, ioport, interrupt)
    if result != PCAN_ERROR_OK:
        log_can_data(device_data_battery_1)
        if result == 5120:
            result = 512
        messagebox.showerror("Error!", GetFormatedError(result))
        return True
    else:
        pcan_write_read('serial_number',1)
        pcan_write_read('serial_number',2)
        if device_data_battery_2['serial_number'] !=0 and device_data_battery_1['serial_number'] !=0:
            pcan_write_read('temperature',2)
            pcan_write_read('firmware_version',2)       
            pcan_write_read('voltage',2)            
            pcan_write_read('battery_status',2)       
            pcan_write_read('current',2)     
            pcan_write_read('remaining_capacity',2)       
            pcan_write_read('full_charge_capacity',2)
            log_can_data(device_data_battery_2)
            pcan_write_read('temperature',1)
            pcan_write_read('firmware_version',1)       
            pcan_write_read('voltage',1)            
            pcan_write_read('battery_status',1)       
            pcan_write_read('current',1)      
            pcan_write_read('remaining_capacity',1)       
            pcan_write_read('full_charge_capacity',1)
            log_can_data(device_data_battery_1)
            messagebox.showinfo("Info!", "Device Connected with 2 batterys")
            return True
        elif device_data_battery_1['serial_number'] != 0:
            pcan_write_read('temperature',1)
            pcan_write_read('firmware_version',1)       
            pcan_write_read('voltage',1)            
            pcan_write_read('battery_status',1)       
            pcan_write_read('current',1)      
            pcan_write_read('remaining_capacity',1)       
            pcan_write_read('full_charge_capacity',1)
            log_can_data(device_data_battery_1)
            messagebox.showinfo("Info!", "Device Connected with 1 batterys")
            return True
        else:
            messagebox.showinfo("Error!", "Connection Failed! Wait for 10 seconds after disconnecting")
            m_objPCANBasic.Uninitialize(m_PcanHandle)
            return False


#PCAN Uninitialize API Call
def pcan_uninitialize(): 
        result =  m_objPCANBasic.Uninitialize(m_PcanHandle)
        if result != PCAN_ERROR_OK:
            update_device_data_to_default()
            update_battery_status_flags_to_default()
            messagebox.showerror("Error!", GetFormatedError(result))
            return False
        else:
            # Reset device_data and battery_status_flags using the update methods
            update_device_data_to_default()
            update_battery_status_flags_to_default()
            messagebox.showinfo("Info!", "Connection Disconnect!")
            return True


#PCAN Write API Call
def pcan_write_read(call_name,battery_no):
        CANMsg = TPCANMsg()
        if battery_no == 1:
            CANMsg.ID = int('18EFC0D0',16)
        elif battery_no == 2:
            CANMsg.ID = int('18EFC1D0',16)
        CANMsg.LEN = int(8)
        CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED
        if call_name == 'firmware_version':
            CANMsg.DATA[0] = int('00',16)
        else:
            CANMsg.DATA[0] = int('01',16)
        CANMsg.DATA[1] = int('00',16)
        if call_name == 'serial_number':
            CANMsg.DATA[2] = int('1c',16)
        elif call_name == 'at_rate':
            CANMsg.DATA[2] = int('04',16)
        elif call_name == 'at_rate_time_to_full':
            CANMsg.DATA[2] = int('05',16)
        elif call_name == 'at_rate_time_to_empty':
            CANMsg.DATA[2] = int('06',16)
        elif call_name == 'at_rate_ok_text':
            CANMsg.DATA[2] = int('07',16)
        elif call_name == 'temperature':
            CANMsg.DATA[2] = int('08',16)
        elif call_name == 'voltage':
            CANMsg.DATA[2] = int('09',16)
        elif call_name == 'current':
            CANMsg.DATA[2] = int('0a',16)
        elif call_name == 'avg_current':
            CANMsg.DATA[2] = int('0b',16)
        elif call_name == 'max_error':
            CANMsg.DATA[2] = int('0c',16)
        elif call_name == 'rel_state_of_charge':
            CANMsg.DATA[2] = int('0d',16)
        elif call_name == 'abs_state_of_charge':
            CANMsg.DATA[2] = int('0e',16)
        elif call_name == 'remaining_capacity':
            CANMsg.DATA[2] = int('0f',16)
        elif call_name == 'full_charge_capacity':
            CANMsg.DATA[2] = int('10',16)
        elif call_name == 'run_time_to_empty':
            CANMsg.DATA[2] = int('11',16)
        elif call_name == 'avg_time_to_empty':
            CANMsg.DATA[2] = int('12',16)
        elif call_name == 'avg_time_to_full':
            CANMsg.DATA[2] = int('13',16)
        elif call_name == 'charging_current':
            CANMsg.DATA[2] = int('14',16)
        elif call_name == 'charging_voltage':
            CANMsg.DATA[2] = int('15',16)
        elif call_name == 'battery_status':
            CANMsg.DATA[2] = int('16',16)
        elif call_name == 'cycle_count':
            CANMsg.DATA[2] = int('17',16)
        elif call_name == 'design_capacity':
            CANMsg.DATA[2] = int('18',16)
        elif call_name == 'design_voltage':
            CANMsg.DATA[2] = int('19',16)
        elif call_name == 'manufacturer_date':
            CANMsg.DATA[2] = int('1b',16)
        elif call_name == 'manufacturer_name':
            CANMsg.DATA[2] = int('20',16)
        elif call_name == 'device_name':
            CANMsg.DATA[2] = int('21',16)
        elif call_name == 'firmware_version':
            CANMsg.DATA[2] = int('00',16)
        else:
            messagebox.showinfo("Error!", "Write operation not found!")   
        if call_name == 'manufacturer_name':
            CANMsg.DATA[3] = int('02',16)
        elif call_name == 'device_name':
            CANMsg.DATA[3] = int('02',16)
        else:
            CANMsg.DATA[3] = int('01',16)  
        CANMsg.DATA[4] = int('08',16)
        CANMsg.DATA[5] = int('02',16)
        CANMsg.DATA[6] = int('00',16)
        if battery_no == 1:
            if call_name == 'serial_number':
                CANMsg.DATA[7] = int('01',16)
            elif call_name == 'at_rate':
                CANMsg.DATA[7] = int('02',16)
            elif call_name == 'at_rate_time_to_full':
                CANMsg.DATA[7] = int('03',16)
            elif call_name == 'at_rate_time_to_empty':
                CANMsg.DATA[7] = int('04',16)
            elif call_name == 'at_rate_ok_text':
                CANMsg.DATA[7] = int('05',16)
            elif call_name == 'temperature':
                CANMsg.DATA[7] = int('06',16)
            elif call_name == 'voltage':
                CANMsg.DATA[7] = int('07',16)
            elif call_name == 'current':
                CANMsg.DATA[7] = int('08',16)
            elif call_name == 'avg_current':
                CANMsg.DATA[7] = int('09',16)
            elif call_name == 'max_error':
                CANMsg.DATA[7] = int('0A',16)
            elif call_name == 'rel_state_of_charge':
                CANMsg.DATA[7] = int('0B',16)
            elif call_name == 'abs_state_of_charge':
                CANMsg.DATA[7] = int('0C',16)
            elif call_name == 'remaining_capacity':
                CANMsg.DATA[7] = int('0D',16)
            elif call_name == 'full_charge_capacity':
                CANMsg.DATA[7] = int('0E',16)
            elif call_name == 'run_time_to_empty':
                CANMsg.DATA[7] = int('0F',16)
            elif call_name == 'avg_time_to_empty':
                CANMsg.DATA[7] = int('10',16)
            elif call_name == 'avg_time_to_full':
                CANMsg.DATA[7] = int('11',16)
            elif call_name == 'charging_current':
                CANMsg.DATA[7] = int('12',16)
            elif call_name == 'charging_voltage':
                CANMsg.DATA[7] = int('13',16)
            elif call_name == 'battery_status':
                CANMsg.DATA[7] = int('14',16)
            elif call_name == 'design_capacity':
                CANMsg.DATA[7] = int('16',16)
            elif call_name == 'design_voltage':
                CANMsg.DATA[7] = int('17',16)
            elif call_name == 'firmware_version':
                CANMsg.DATA[7] = int('1B',16)
        elif battery_no == 2:
            if call_name == 'serial_number':
                CANMsg.DATA[7] = int('1C',16)
            elif call_name == 'at_rate':
                CANMsg.DATA[7] = int('1D',16)
            elif call_name == 'at_rate_time_to_full':
                CANMsg.DATA[7] = int('1E',16)
            elif call_name == 'at_rate_time_to_empty':
                CANMsg.DATA[7] = int('1F',16)
            elif call_name == 'at_rate_ok_text':
                CANMsg.DATA[7] = int('20',16)
            elif call_name == 'temperature':
                CANMsg.DATA[7] = int('21',16)
            elif call_name == 'voltage':
                CANMsg.DATA[7] = int('22',16)
            elif call_name == 'current':
                CANMsg.DATA[7] = int('23',16)
            elif call_name == 'avg_current':
                CANMsg.DATA[7] = int('24',16)
            elif call_name == 'max_error':
                CANMsg.DATA[7] = int('25',16)
            elif call_name == 'rel_state_of_charge':
                CANMsg.DATA[7] = int('26',16)
            elif call_name == 'abs_state_of_charge':
                CANMsg.DATA[7] = int('27',16)
            elif call_name == 'remaining_capacity':
                CANMsg.DATA[7] = int('28',16)
            elif call_name == 'full_charge_capacity':
                CANMsg.DATA[7] = int('29',16)
            elif call_name == 'run_time_to_empty':
                CANMsg.DATA[7] = int('2A',16)
            elif call_name == 'avg_time_to_empty':
                CANMsg.DATA[7] = int('2B',16)
            elif call_name == 'avg_time_to_full':
                CANMsg.DATA[7] = int('2C',16)
            elif call_name == 'charging_current':
                CANMsg.DATA[7] = int('2D',16)
            elif call_name == 'charging_voltage':
                CANMsg.DATA[7] = int('2E',16)
            elif call_name == 'battery_status':
                CANMsg.DATA[7] = int('2F',16)
            elif call_name == 'design_capacity':
                CANMsg.DATA[7] = int('30',16)
            elif call_name == 'design_voltage':
                CANMsg.DATA[7] = int('31',16)
            elif call_name == 'firmware_version':
                CANMsg.DATA[7] = int('32',16)
        
        result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
        print(f"{call_name}:{result}")
        if result != PCAN_ERROR_OK:
            messagebox.showerror(f"Error! {call_name}", GetFormatedError(result))
            return -2
        else:
            time.sleep(0.1)
            result_code = pcan_read()
            return result_code


#PCAN Read API Call
def pcan_read():
    result = m_objPCANBasic.Read(m_PcanHandle)
    if result[0] != PCAN_ERROR_OK:
        return result[0]
    else:
        args = result[1:]
        theMsg = args[0]
        newMsg = TPCANMsgFD()
        newMsg.ID = theMsg.ID
        newMsg.DLC = theMsg.LEN
        for i in range(8 if (theMsg.LEN > 8) else theMsg.LEN):
            newMsg.DATA[i] = theMsg.DATA[i]
        
        resulthex = [hex(theMsg.DATA[i]) for i in range(8 if (theMsg.LEN > 8) else theMsg.LEN)]
        
        first_byte = newMsg.DATA[0]
        second_byte = newMsg.DATA[1]

        # Swap the bytes to create the correct hex value
        swapped_hex = (second_byte << 8) | first_byte
        decimal_value = int(swapped_hex)

        if newMsg.DATA[4] ==  0x00:
            data_packet = [int(hex(newMsg.DATA[i]), 16) for i in range(8 if (theMsg.LEN > 8) else theMsg.LEN)]
            major_version = data_packet[0]
            minor_version = data_packet[1]
            patch_number = (data_packet[3] << 8) | data_packet[2]
            build_number = (data_packet[5] << 8) | data_packet[4]

            version_string = f"{major_version}.{minor_version}.{patch_number}.{build_number}"
            if newMsg.DATA[7] == 0x01:
                device_data_battery_1['firmware_version'] = version_string
            elif newMsg.DATA[7] == 0x1C:
                device_data_battery_2['firmware_version'] = version_string
            else:
                print("Firmware version is not found")
        else:
            convert_data(newMsg, decimal_value)

        return result[0]


def convert_data(newMsg, decimal_value):
    # Conversion rules
    if newMsg.DATA[4] == 0x04:  # AtRate: mA / 40 unsigned
        if newMsg.DATA[7] == 0x02:
            device_data_battery_1['at_rate'] = (decimal_value*40)/1000
        elif newMsg.DATA[7] == 0x1D:
            device_data_battery_2['at_rate'] = (decimal_value*40)/1000
        else:
            print("AtRate Not Found")
    elif newMsg.DATA[4] == 0x05:  # AtRateTimeToFull: minutes unsigned
        if newMsg.DATA[7] == 0x03:
            device_data_battery_1['at_rate_time_to_full'] = round((decimal_value / 1000),1)
        elif newMsg.DATA[7] == 0x1E:
            device_data_battery_2['at_rate_time_to_full'] = round((decimal_value / 1000),1)
        else:
            print("AtRateTimeToFull Not Found")
    elif newMsg.DATA[4] == 0x06:  # AtRateTimeToEmpty: minutes unsigned
        if newMsg.DATA[7] == 0x04:
            device_data_battery_1['at_rate_time_to_empty'] = round((decimal_value / 1000),1)
        elif newMsg.DATA[7] == 0x1F:
            device_data_battery_2['at_rate_time_to_empty'] = round((decimal_value / 1000),1)
        else:
            print("AtRateTimeToEmpty Not Found")
    elif newMsg.DATA[4] == 0x07:  # AtRateOK: Boolean
        if newMsg.DATA[7] == 0x05:
            device_data_battery_1['at_rate_ok_text'] = "Yes" if decimal_value != 0 else "No" 
        elif newMsg.DATA[7] == 0x20:
            device_data_battery_2['at_rate_ok_text'] = "Yes" if decimal_value != 0 else "No" 
        else:
            print("AtRateOK Not Found")
    elif newMsg.DATA[4] == 0x08: # Temperature: Boolean
        temperature_k = decimal_value / 10.0
        temperature_c = temperature_k - 273.15
        if newMsg.DATA[7] == 0x06:
            device_data_battery_1['temperature'] = round(temperature_c,1) 
        elif newMsg.DATA[7] == 0x21:
            device_data_battery_2['temperature'] = round(temperature_c,1)
        else:
            print("Temperature Not Found")
    elif newMsg.DATA[4] == 0x09:  # Voltage: mV unsigned
        if newMsg.DATA[7] == 0x07:
            device_data_battery_1['voltage'] = round((decimal_value / 1000),1)
        elif newMsg.DATA[7] == 0x22:
            device_data_battery_2['voltage'] = round((decimal_value / 1000),1)
        else:
            print("Voltage Not Found")
    elif newMsg.DATA[4] == 0x0a:  # Current: mA / 40 signed
        if decimal_value == 0:
            if newMsg.DATA[7] == 0x08:
                device_data_battery_1['current'] = decimal_value
            elif newMsg.DATA[7] == 0x23:
                device_data_battery_2['current'] = decimal_value
            else:
                print("Current Not Found")
        else:
            if decimal_value > 32767:
                decimal_value -= 65536
            currentmA = decimal_value*40
            currentA = currentmA / 1000
            if currentA > 0:
                if newMsg.DATA[7] == 0x08:
                    device_data_battery_1['charging_battery_status'] = "Charging"
                    device_data_battery_1['current'] = 0
                    device_data_battery_1['charging_current'] = abs(currentA)
                elif newMsg.DATA[7] == 0x23:
                    device_data_battery_2['charging_battery_status'] = "Charging"
                    device_data_battery_2['current'] = 0
                    device_data_battery_2['charging_current'] = abs(currentA)
                else:
                    print("Current Not Found")
            elif currentA < 0:
                if newMsg.DATA[7] == 0x08:
                    device_data_battery_1['charging_battery_status'] = "Discharging"
                    device_data_battery_1['current'] = abs(currentA)
                    device_data_battery_1['charging_current'] = 0
                elif newMsg.DATA[7] == 0x23:
                    device_data_battery_2['charging_battery_status'] = "Discharging"
                    device_data_battery_2['current'] = abs(currentA)
                    device_data_battery_2['charging_current'] = 0
                else:
                    print("Current Not Found")
            else:
                if newMsg.DATA[7] == 0x08:
                    device_data_battery_1['charging_battery_status'] = "Off"
                    device_data_battery_1['current'] = 0
                    device_data_battery_1['charging_current'] = 0
                elif newMsg.DATA[7] == 0x23:
                    device_data_battery_2['current'] = decimal_value
                    device_data_battery_2['current'] = decimal_value
                    device_data_battery_2['current'] = decimal_value
                else:
                    print("Current Not Found")
    elif newMsg.DATA[4] == 0x0b:  # Avg Current: mA / 40 signed
        if decimal_value == 0:
            if newMsg.DATA[7] == 0x09:
                device_data_battery_1['avg_current'] = decimal_value
            elif newMsg.DATA[7] == 0x24:
                device_data_battery_2['avg_current'] = decimal_value
            else:
                print("Avg Current Not Found")
        else:
            if decimal_value > 32767:
                decimal_value -= 65536
            currentmA = decimal_value*40
            currentA = currentmA / 1000
            if newMsg.DATA[7] == 0x09:
                device_data_battery_1['avg_current'] = currentA
            elif newMsg.DATA[7] == 0x24:
                device_data_battery_2['avg_current'] = currentA
            else:
                print("Avg Current Not Found")
    elif newMsg.DATA[4] == 0x0c:  # MaxError: Percent unsigned
        if newMsg.DATA[7] == 0x0A:
            device_data_battery_1['max_error'] = decimal_value
        elif newMsg.DATA[7] == 0x25:
            device_data_battery_2['max_error'] = decimal_value
        else:
            print("MaxError Not Found")
    elif newMsg.DATA[4] == 0x0d:  # RelStateofCharge: Percent unsigned
        if newMsg.DATA[7] == 0x0B:
            device_data_battery_1['rel_state_of_charge'] = decimal_value
        elif newMsg.DATA[7] == 0x26:
            device_data_battery_2['rel_state_of_charge'] = decimal_value
        else:
            print("RelStateofCharge Not Found")
    elif newMsg.DATA[4] == 0x0e:  # AbsoluteStateofCharge: Percent unsigned
        if newMsg.DATA[7] == 0x0C:
            device_data_battery_1['abs_state_of_charge'] = decimal_value
        elif newMsg.DATA[7] == 0x27:
            device_data_battery_2['abs_state_of_charge'] = decimal_value
        else:
            print("AbsoluteStateofCharge Not Found")
    elif newMsg.DATA[4] == 0x0f:  # RemainingCapacity: mAh / 40 unsigned
        if newMsg.DATA[7] == 0x0D:
            device_data_battery_1['remaining_capacity'] = (decimal_value*40)/1000
        elif newMsg.DATA[7] == 0x28:
            device_data_battery_2['remaining_capacity'] = (decimal_value*40)/1000
        else:
            print("MaxError Not Found")
    elif newMsg.DATA[4] == 0x10:  # FullChargeCapacity: mAh / 40 unsigned
        if newMsg.DATA[7] == 0x0E:
            device_data_battery_1['full_charge_capacity'] = (decimal_value*40)/1000
        elif newMsg.DATA[7] == 0x29:
            device_data_battery_2['full_charge_capacity'] = (decimal_value*40)/1000
        else:
            print("FullChargeCapacity Not Found")
    elif newMsg.DATA[4] == 0x11:  # RunTimeToEmpty: minutes unsigned
        if newMsg.DATA[7] == 0x0F:
            device_data_battery_1['run_time_to_empty'] = round((decimal_value / 10),1)
        elif newMsg.DATA[7] == 0x2A:
            device_data_battery_2['run_time_to_empty'] = round((decimal_value / 10),1)
        else:
            print("RunTimeToEmpty Not Found")
    elif newMsg.DATA[4] == 0x12:  # AvgTimeToEmpty: minutes unsigned
        if newMsg.DATA[7] == 0x10:
            device_data_battery_1['avg_time_to_empty'] = round((decimal_value / 10),1)
        elif newMsg.DATA[7] == 0x2B:
            device_data_battery_2['avg_time_to_empty'] = round((decimal_value / 10),1)
        else:
            print("AvgTimeToEmpty Not Found")
        device_data['avg_time_to_empty'] = round((decimal_value / 10),1)
    elif newMsg.DATA[4] == 0x13:  # AvgTimeToFull: minutes unsigned
        if newMsg.DATA[7] == 0x11:
            device_data_battery_1['avg_time_to_full'] = round((decimal_value / 1000),1)
        elif newMsg.DATA[7] == 0x2C:
            device_data_battery_2['avg_time_to_full'] = round((decimal_value / 1000),1)
        else:
            print("AvgTimeToFull Not Found")
    elif newMsg.DATA[4] == 0x13:  # ChargingVoltage: mV unsigned
        if newMsg.DATA[7] == 0x13:
            device_data_battery_1['charging_voltage'] = round((decimal_value / 1000),1)
        elif newMsg.DATA[7] == 0x2E:
            device_data_battery_2['charging_voltage'] = round((decimal_value / 1000),1)
        else:
            print("ChargingVoltage Not Found")
    elif newMsg.DATA[4] == 0x16:  # BatteryStatus: bit flags unsigned
        binary_value = format(decimal_value, '016b')  # Convert to 16-bit binary string
        
        # Swap the bytes before interpreting the bits
        swapped_value = binary_value[8:] + binary_value[:8]
        if newMsg.DATA[7] == 0x14:
            device_data_battery_1['battery_status'] = swapped_value
            battery_1_status_flags.update({
                    "overcharged_alarm": int(swapped_value[0]),  # Bit 15
                    "terminate_charge_alarm": int(swapped_value[1]),  # Bit 14
                    "over_temperature_alarm": int(swapped_value[3]),  # Bit 12
                    "terminate_discharge_alarm": int(swapped_value[4]),  # Bit 11
                    "remaining_capacity_alarm": int(swapped_value[6]),  # Bit 9
                    "remaining_time_alarm": int(swapped_value[7]),  # Bit 8
                    "initialization": int(swapped_value[8]),  # Bit 7
                    "charge_fet_test": int(swapped_value[9]),  # Bit 6
                    "fully_charged": int(swapped_value[10]),  # Bit 5
                    "fully_discharged": int(swapped_value[11]),  # Bit 4
                    "error_codes": int(swapped_value[12:], 2)  # Bits 3:0, convert remaining bits to an integer
                })
        elif newMsg.DATA[7] == 0x2F:
            device_data_battery_2['battery_status'] = swapped_value
            battery_2_status_flags.update({
                    "overcharged_alarm": int(swapped_value[0]),  # Bit 15
                    "terminate_charge_alarm": int(swapped_value[1]),  # Bit 14
                    "over_temperature_alarm": int(swapped_value[3]),  # Bit 12
                    "terminate_discharge_alarm": int(swapped_value[4]),  # Bit 11
                    "remaining_capacity_alarm": int(swapped_value[6]),  # Bit 9
                    "remaining_time_alarm": int(swapped_value[7]),  # Bit 8
                    "initialization": int(swapped_value[8]),  # Bit 7
                    "charge_fet_test": int(swapped_value[9]),  # Bit 6
                    "fully_charged": int(swapped_value[10]),  # Bit 5
                    "fully_discharged": int(swapped_value[11]),  # Bit 4
                    "error_codes": int(swapped_value[12:], 2)  # Bits 3:0, convert remaining bits to an integer
                })
        else:
            print("BatteryStatus Not Found")          
    elif newMsg.DATA[4] == 0x18:  # DesignCapacity: mAh / 40 unsigned
        if newMsg.DATA[7] == 0x16:
            device_data_battery_1['design_capacity'] = (decimal_value*40)/1000
        elif newMsg.DATA[7] == 0x2F:
            device_data_battery_2['design_capacity'] = (decimal_value*40)/1000
        else:
            print("DesignCapacity Not Found")
        device_data['design_capacity'] = (decimal_value*40)/1000
    elif newMsg.DATA[4] == 0x19:  # DesignVoltage: mV unsigned
        if newMsg.DATA[7] == 0x17:
            device_data_battery_1['design_voltage'] = round((decimal_value / 1000),1)
        elif newMsg.DATA[7] == 0x31:
            device_data_battery_2['design_voltage'] = round((decimal_value / 1000),1)
        else:
            print("DesignVoltage Not Found")
    elif newMsg.DATA[4] == 0x1c:  # SerialNumber: number unsigned
        if newMsg.DATA[7] == 0x01:
            device_data_battery_1['serial_number'] = decimal_value
        elif newMsg.DATA[7] == 0x1C:
            device_data_battery_2['serial_number'] = decimal_value
        else:
            print("Serial Number Not Found")
    else:
        print(f"Result Command Code Not Found : {hex(newMsg.DATA[4])}")


def GetFormatedError(error):
        # Gets the text using the GetErrorText API function
        # If the function success, the translated error is returned. If it fails,
        # a text describing the current error is returned.
        #
        stsReturn = m_objPCANBasic.GetErrorText(error, 0)
        if stsReturn[0] != PCAN_ERROR_OK:
            return "An error occurred. Error-code's text ({0:X}h) couldn't be retrieved".format(error)
        else:
            return stsReturn[1]


def pcan_write_control(call_name,battery_no):
        CANMsg = TPCANMsg()
        if battery_no == 1:
            CANMsg.ID = int('18EFC0D0',16)
        elif battery_no == 2:
            CANMsg.ID = int('18EFC1D0',16)
        CANMsg.LEN = int(8)
        CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED
        CANMsg.DATA[0] = int('03',16)
        CANMsg.DATA[1] = int('03',16)
        CANMsg.DATA[2] = int('37',16)
        CANMsg.DATA[3] = int('30',16)
        CANMsg.DATA[4] = int('39',16)
        CANMsg.DATA[5] = int('33',16)
        CANMsg.DATA[6] = int('39',16)
        CANMsg.DATA[7] = int('00',16)
        result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
        if result == PCAN_ERROR_OK:
            # pcan_write_read('current',battery_no)
            print("Enter Diag State : Success")
            CANMsg = TPCANMsg()
            if battery_no == 1:
                CANMsg.ID = int('18EFC0D0',16)
            elif battery_no == 2:
                CANMsg.ID = int('18EFC1D0',16)
            CANMsg.LEN = int(8)
            CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED
            CANMsg.DATA[0] = int('03',16)
            if call_name == 'both_off':
                CANMsg.DATA[1] = int('01',16)
            elif call_name == 'heater_on':
                CANMsg.DATA[1] = int('05',16)
            else:
                CANMsg.DATA[1] = int('00',16)
            if call_name == 'both_off':
                CANMsg.DATA[2] = int('00',16)
            elif call_name == 'charge_on':
                CANMsg.DATA[2] = int('01',16)
            elif call_name == 'discharge_on':
                CANMsg.DATA[2] = int('02',16)
            elif call_name == 'both_on':
                CANMsg.DATA[2] = int('03',16)
            elif call_name == 'bms_reset':
                CANMsg.DATA[2] = int('01',16)
            elif call_name == 'heater_on':
                CANMsg.DATA[2] = int('00',16)
            else:
                messagebox.showinfo("Error!", "Write operation not found!")     
            CANMsg.DATA[3] = int('01',16)
            CANMsg.DATA[4] = int('08',16)
            CANMsg.DATA[5] = int('02',16)
            CANMsg.DATA[6] = int('00',16)
            CANMsg.DATA[7] = int('00',16)
            # m_objPCANBasic.Write(m_PcanHandle, CANMsg)
            result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
            if result == PCAN_ERROR_OK:
                print("FET Control State : Success")
                CANMsg = TPCANMsg()
                if battery_no == 1:
                    CANMsg.ID = int('18EFC0D0',16)
                elif battery_no == 2:
                    CANMsg.ID = int('18EFC1D0',16)
                CANMsg.LEN = int(8)
                CANMsg.MSGTYPE = PCAN_MESSAGE_EXTENDED
                CANMsg.DATA[0] = int('03',16)
                CANMsg.DATA[1] = int('04',16)
                CANMsg.DATA[2] = int('00',16)   
                CANMsg.DATA[3] = int('00',16)
                CANMsg.DATA[4] = int('00',16)
                CANMsg.DATA[5] = int('00',16)
                CANMsg.DATA[6] = int('00',16)
                CANMsg.DATA[7] = int('00',16)
                result = m_objPCANBasic.Write(m_PcanHandle, CANMsg)
                if result == PCAN_ERROR_OK:
                    print("Exit the Diag State : Success")
                    return 0
                else:
                    print("Exit the Diag State : Failed")
                    return 0
            else:
                print("FET Control State : Failed")
                return 0
        else:
            print("Enter Diag State : Failed")
            return 0


def update_device_data_to_default():
    for key in device_data_battery_1.keys():
        if key == 'device_name':
            device_data_battery_1[key] = "BT-70939APH"  # Example: Set a specific default device name
        elif key == 'manufacturer_name':
            device_data_battery_1[key] = "Bren-Tronics"  # Example: Set a specific manufacturer name
        elif key == 'charging_battery_status':
            device_data_battery_1[key] = "Off"  # Example: Set a specific charging status
        elif isinstance(device_data[key], str):
            device_data_battery_1[key] = ""  # Reset other string values to empty strings
        elif isinstance(device_data[key], (int, float)):
            device_data_battery_1[key] = 0  # Reset numeric values to zero
    for key in device_data_battery_2.keys():
        if key == 'device_name':
            device_data_battery_2[key] = "BT-70939APH"  # Example: Set a specific default device name
        elif key == 'manufacturer_name':
            device_data_battery_2[key] = "Bren-Tronics"  # Example: Set a specific manufacturer name
        elif key == 'charging_battery_status':
            device_data_battery_2[key] = "Off"  # Example: Set a specific charging status
        elif isinstance(device_data[key], str):
            device_data_battery_2[key] = ""  # Reset other string values to empty strings
        elif isinstance(device_data[key], (int, float)):
            device_data_battery_2[key] = 0  # Reset numeric values to zero


def update_battery_status_flags_to_default():
    for key in battery_1_status_flags.keys():
        battery_1_status_flags[key] = 0
    for key in battery_2_status_flags.keys():
        battery_2_status_flags[key] = 0


def log_can_data(update_can_data):
    """
    Log CAN data to an Excel file. If the serial number exists, do nothing;
    if it doesn't exist, append a new row with default values.
    """
    try:
        # Define the path for the CAN data Excel file inside the AppData directory
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Define the log file name
        file_name = "can_data.xlsx"
        file_path = os.path.join(folder_path, file_name)

        # Create or load workbook
        if os.path.exists(file_path):
            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active
            except Exception as e:
                print(f"Error loading the Excel file: {str(e)}")
                messagebox.showerror("Error", f"Failed to load the Excel file: {str(e)}. Recreating the file.")
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "CAN Data"
                # Add headers for CAN data
                headers = [
                    "SI No", "Date", "Time", "Project", "Device Name", "Manufacturer Name", 
                    "Serial Number", "Cycle Count", "Full Charge Capacity", "Charging Date", 
                    "OCV Before Charging", "Discharging Date", "OCV Before Discharging"
                ]
                sheet.append(headers)
        else:
            # File does not exist, so create it
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "CAN Data"
            # Add headers for CAN data
            headers = [
                "SI No", "Date", "Time", "Project", "Device Name", "Manufacturer Name", 
                "Serial Number", "Cycle Count", "Full Charge Capacity", "Charging Date", 
                "OCV Before Charging", "Discharging Date", "OCV Before Discharging"
            ]
            sheet.append(headers)

        # Log the current CAN data
        current_datetime = datetime.datetime.now()
        date = current_datetime.strftime("%Y-%m-%d")
        time = current_datetime.strftime("%H:%M:%S")
        serial_number = update_can_data.get('serial_number', 'N/A')

        # Check if the serial number exists in the file
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        serial_found = False

        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                serial_found = True
                print(f"Serial number {serial_number} already exists. No action taken.")
                break

        if not serial_found:
            # If serial number is not found, add a new row with default values
            next_row = sheet.max_row + 1
            can_data = [
                next_row - 1,  # SI No
                date,  # Date
                time,  # Time
                update_can_data.get('project', ''),  # Project
                update_can_data.get('device_name', 'BT-70939APH'),  # Device Name
                update_can_data.get('manufacturer_name', 'Bren-Tronics'),  # Manufacturer Name
                serial_number,  # Serial Number
                int(update_can_data.get('cycle_count', 0)),  # Ensure Cycle Count is an int
                103,  # Ensure Full Charge Capacity is a float
                update_can_data.get('charging_date', date),  # Charging Date
                float(update_can_data.get('ocv_before_charging', 0.0)),  # Ensure OCV Before Charging is a float
                update_can_data.get('discharging_date', date),  # Discharging Date
                float(update_can_data.get('ocv_before_discharging', 0.0))  # Ensure OCV Before Discharging is a float
            ]
            sheet.append(can_data)

        # Save the Excel file
        workbook.save(file_path)
        workbook.close()
        print(f"CAN data logged in {file_path}.")

    except Exception as e:
        print(f"An error occurred while updating the Excel file: {str(e)}")
        messagebox.showerror("Error", f"An error occurred while updating the Excel file: {str(e)}")


def get_latest_can_data(serial_number):
    """
    Retrieves the most recent CAN data for the specified serial number.
    """
    # Path to the CAN data Excel file
    can_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "can_data.xlsx")

    if not os.path.exists(can_data_file):
        messagebox.showerror("Error", "No CAN data file found. Please log CAN data first.")
        return None

    # Load the existing workbook and access the first sheet
    workbook = load_workbook(can_data_file)
    sheet = workbook.active

    # Find the serial number column (assuming it's the 7th column, adjust if necessary)
    serial_number_column = 7  # Serial Number is the 7th column (G column)

    # Search for the serial number in the rows and store the data if found
    serial_found = False
    latest_data = {}

    for row in range(2, sheet.max_row + 1):  # Start from the second row to skip headers
        current_serial_number = sheet.cell(row=row, column=serial_number_column).value
        if current_serial_number == serial_number:
            serial_found = True
            print(f"Serial Number {serial_number} found in row {row}")

            # Retrieve the headers from the first row
            headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
            # Retrieve the values for the matching row
            values = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
            
            # Map headers to values and print for debugging
            latest_data = dict(zip(headers, values))
            for header, value in latest_data.items():
                print(f"{header}: {value}")  # Debug print to ensure correct values are fetched
            break

    if not serial_found:
        messagebox.showwarning("Warning", f"Serial number {serial_number} not found in the CAN data.")
        return None

    return latest_data



def update_cycle_count_in_can_data(serial_number, new_cycle_count):
    """
    Update the cycle count for a specific device identified by its serial number
    in the CAN Data Excel file.
    """
    # Define the folder and file path for the CAN data file
    folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
    file_path = os.path.join(folder_path, "can_data.xlsx")

    # Check if the file exists
    if not os.path.exists(file_path):
        messagebox.showerror("Error", "CAN data file not found.")
        return

    # Open the existing Excel file
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Loop through the rows to find the correct row by Serial Number
    serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
    cycle_count_column = 8    # Assuming Cycle Count is in column H (8th column)

    for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
        if sheet.cell(row=row, column=serial_number_column).value == serial_number:
            # Update the Cycle Count value in the corresponding row
            sheet.cell(row=row, column=cycle_count_column).value = new_cycle_count
            break
    else:
        messagebox.showwarning("Warning", f"Serial number {serial_number} not found.")
        return

    # Save the updated Excel file
    workbook.save(file_path)
    messagebox.showinfo("Success", f"Cycle count updated successfully for Serial Number {serial_number}.")


def update_excel_and_download_pdf(data):
    """
    Get the updated values from the form, update the Excel file, and generate a PDF.
    """
    # Define the path to the CAN data Excel file
    can_data_file = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database", "can_data.xlsx")

    # Load the existing data
    try:
        df_can_data = pd.read_excel(can_data_file)
    except FileNotFoundError:
        messagebox.showerror("Error", f"{can_data_file} not found.")
        return
    
    # Assuming the serial number is unique, update the corresponding row
    serial_number = data[2]  # serial_number is the second element in the array
    df_can_data['Serial Number'] = df_can_data['Serial Number'].astype(str).str.strip()
    df_can_data['OCV Before Charging'] = df_can_data['OCV Before Charging'].astype(float)
    df_can_data['OCV Before Discharging'] = df_can_data['OCV Before Discharging'].astype(float)
    index = df_can_data[df_can_data['Serial Number'] == serial_number].index
    print(f"{df_can_data['Serial Number']} test {serial_number}")
    print(f"{index} index")
    if not index.empty:
        # Update the values in the DataFrame
        df_can_data.loc[index[0], 'Project'] = str(data[0])  # Ensure it's a string
        df_can_data.loc[index[0], 'Device Name'] = str(data[1])  # Ensure it's a string
        df_can_data.loc[index[0], 'Manufacturer Name'] = str(data[3])  # Ensure it's a string
        df_can_data.loc[index[0], 'Serial Number'] = int(data[2])  # Ensure it's a string
        df_can_data.loc[index[0], 'Cycle Count'] = int(data[4])  # Convert to int
        df_can_data.loc[index[0], 'Full Charge Capacity'] = float(data[5])  # Convert to float
        df_can_data.loc[index[0], 'Charging Date'] = str(data[6])  # Ensure it's a string
        df_can_data.loc[index[0], 'OCV Before Charging'] = float(data[7])  # Convert to float
        df_can_data.loc[index[0], 'Discharging Date'] = str(data[8])  # Ensure it's a string
        df_can_data.loc[index[0], 'OCV Before Discharging'] = float(data[9])  # Convert to float

        # Save the updated DataFrame back to the Excel file
        df_can_data.to_excel(can_data_file, index=False)
        messagebox.showinfo("Success", "Data updated successfully.")
    else:
        messagebox.showwarning("Warning", "Serial Number not found in the Excel file.")

    # # Generate the PDF with the updated values
    pdf_generator.create_can_report_pdf(serial_number,"CAN")


def open_pdf_folder():
    # Get the folder path where the PDFs are saved
    folder_path = os.path.join(os.path.expanduser("~"), "Documents", "Battery_PDFs")
    
    # Check if the folder exists, if not, show an error
    if not os.path.exists(folder_path):
        messagebox.showerror("Error", "No PDF files found. Please generate a PDF first.")
        return
    
    # Open the folder
    try:
        os.startfile(folder_path)  # This will open the folder in Windows Explorer
    except Exception as e:
        messagebox.showerror("Error", f"Unable to open folder: {str(e)}")


def log_charging_duration_to_excel(serial_number, charging_start_time, charging_end_time, duration):
    # Create or open the Excel file for charging logs
    folder_path = os.path.join(os.path.expanduser("~"), "Documents", "Battery_Logs")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    file_path = os.path.join(folder_path, "Charging_Log.xlsx")

    # Load existing workbook or create a new one if the file does not exist
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Charging Log"
        headers = ["Serial Number", "Rel State of Charge", "Start Date", "Start Time", "End Date", "End Time", "Duration (minutes)"]
        sheet.append(headers)

    sheet = workbook.active

    # Format date and time for both start and end
    start_date = charging_start_time.strftime("%Y-%m-%d")
    start_time = charging_start_time.strftime("%H:%M:%S")
    end_date = charging_end_time.strftime("%Y-%m-%d")
    end_time = charging_end_time.strftime("%H:%M:%S")
    duration_in_minutes = round(duration.total_seconds() / 60, 2)  # Convert duration to minutes

    # Log the session
    rel_state_of_charge = device_data['rel_state_of_charge']  # Assuming device_data has this key
    sheet.append([serial_number, rel_state_of_charge, start_date, start_time, end_date, end_time, duration_in_minutes])

    # Save the workbook and close it
    workbook.save(file_path)
    workbook.close()


def fetch_charging_info(serial_number):
    # Path to the Charging Log Excel file
    folder_path = os.path.join(os.path.expanduser("~"), "Documents", "Battery_Logs")
    file_path = os.path.join(folder_path, "Charging_Log.xlsx")

    # Check if the file exists
    if not os.path.exists(file_path):
        return None, None  # No log found, return None for both duration and date

    # Load the workbook and access the first sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Iterate over the rows to find the most recent entry for the serial number
    for row in sheet.iter_rows(values_only=True):
        if row[0] == serial_number:  # Assuming serial number is in the first column
            date = row[2]  # Assuming date is in the 3rd column
            duration = row[6]  # Assuming duration is in the 5th column
            return duration, date

    # If no matching entry is found
    return None, None  # No previous session found


async def fetch_current(battery_no):
        global is_fetching_current  # Declare it as a global variable
        # Keep fetching the current value asynchronously
        while is_fetching_current:
            try:
                # Call the pcan_write_read method asynchronously
                current_value = await asyncio.to_thread(pcan_write_read, 'current', battery_no)
                
                # Check if the current value is valid
                if current_value is not None:  # Adjust this based on what you consider a valid value
                    print(f"Current value for battery {battery_no}: {current_value}")
                    is_fetching_current = False  # Stop fetching once a valid value is obtained
                    break
            except Exception as e:
                print(f"Error fetching current for battery {battery_no}: {e}")
            
            await asyncio.sleep(config.config_values['can_config']['logging_time'])  # Sleep for 500ms asynchronously before trying again


def start_fetching_current(battery_no):
    global is_fetching_current  # Declare it as a global variable
    # Start fetching in the event loop
    if not is_fetching_current:
        if battery_no == 1:
            update_charging_ocv_in_excel(device_data_battery_1['serial_number'],device_data_battery_1['voltage'])
        elif battery_no == 2:
            update_charging_ocv_in_excel(device_data_battery_2['serial_number'],device_data_battery_2['voltage'])
        is_fetching_current = True
        asyncio.create_task(fetch_current(battery_no))


def stop_fetching_current():
    global is_fetching_current  # Declare it as a global variable
    # Stop fetching current
    is_fetching_current = False

def is_excel_file_open(file_path):
    """Check if the Excel file is open by trying to open it in write mode."""
    try:
        # Attempt to open the file in write mode (this will raise an error if the file is open)
        with open(file_path, 'a'):
            return False  # File is not open
    except PermissionError:
        return True  # File is open 


def update_charging_ocv_in_excel(serial_number, ocv_value):
    """
    Updates the charging OCV and charging date fields in the CAN data Excel file.
    
    Parameters:
    - serial_number (str): The serial number of the battery.
    - ocv_value (float): The OCV value before charging.
    """
    try:
        # Define the folder and file path for the CAN data log
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        file_path = os.path.join(folder_path, "can_data.xlsx")

        # Load the Excel file
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "CAN data log file not found.")
            return False

        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Loop through rows to find the correct row by Serial Number
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        charging_ocv_column = 11   # Assuming OCV Before Charging is in column K (11th column)
        charging_date_column = 10   # Assuming Charging Date is in column J (10th column)

        serial_found = False
        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                serial_found = True
                # Update the OCV before charging
                sheet.cell(row=row, column=charging_ocv_column).value = ocv_value
                # Set the current date as the charging date
                sheet.cell(row=row, column=charging_date_column).value = datetime.datetime.now().strftime("%Y-%m-%d")
                print(f"Updated Charging OCV and Date for Serial: {serial_number}")
                break

        if not serial_found:
            messagebox.showwarning("Warning", f"Serial number {serial_number} not found.")
            return False

        # Save the changes to the Excel file
        workbook.save(file_path)
        workbook.close()
        return True
    
    except FileNotFoundError:
        messagebox.showerror("Error", "The Excel file could not be found.")
    except PermissionError:
        messagebox.showerror("Error", "The Excel file is currently open. Please close it and try again.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

def update_discharging_ocv_in_excel(serial_number, ocv_value):
    """
    Updates the discharging OCV and discharging date fields in the CAN data Excel file.
    
    Parameters:
    - serial_number (str): The serial number of the battery.
    - ocv_value (float): The OCV value before discharging.
    """
    try:
        # Define the folder and file path for the CAN data log
        folder_path = os.path.join(os.getenv('LOCALAPPDATA'), "ADE BMS", "database")
        file_path = os.path.join(folder_path, "can_data.xlsx")

        # Load the Excel file
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "CAN data log file not found.")
            return False

        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Loop through rows to find the correct row by Serial Number
        serial_number_column = 7  # Assuming Serial Number is in column G (7th column)
        discharging_ocv_column = 12  # Assuming OCV Before Discharging is in column L (12th column)
        discharging_date_column = 13  # Assuming Discharging Date is in column M (13th column)

        serial_found = False
        for row in range(2, sheet.max_row + 1):  # Start at 2 to skip the header
            if sheet.cell(row=row, column=serial_number_column).value == serial_number:
                serial_found = True
                # Update the OCV before discharging
                sheet.cell(row=row, column=discharging_ocv_column).value = ocv_value
                # Set the current date as the discharging date
                sheet.cell(row=row, column=discharging_date_column).value = datetime.datetime.now().strftime("%Y-%m-%d")
                print(f"Updated Discharging OCV and Date for Serial: {serial_number}")
                break

        if not serial_found:
            messagebox.showwarning("Warning", f"Serial number {serial_number} not found.")
            return False

        # Save the changes to the Excel file
        workbook.save(file_path)
        workbook.close()
        return True
    
    except FileNotFoundError:
        messagebox.showerror("Error", "The Excel file could not be found.")
    except PermissionError:
        messagebox.showerror("Error", "The Excel file is currently open. Please close it and try again.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

